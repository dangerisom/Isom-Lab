# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# Copyright (c) 2026 Daniel G. Isom

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import os
from datetime import datetime

import cv2
import numpy as np
from PIL import Image, ImageTk

class ImageTriangulatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bridge & Projection & Protrusion Quantifier")
        self.geometry("1500x900")      # widened window
        self.minsize(1400, 800)

        # Debugging...
        self.enable_interface_islands = True  # or False by default
        self.enable_green_fragment_to_interface = False  # or True if you want it on by default

        # track shutdown / after callbacks
        self._after_ids = []
        self._is_closing = False
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.main_container = tk.Frame(self)
        self.main_container.pack(fill="both", expand=True)

        # Left panel: do NOT expand fully; let it size to content
        self.left_panel = tk.Frame(self.main_container)
        self.left_panel.pack(side="left", fill="y", expand=False)

        # Right panel: take remaining width
        self.right_panel = tk.Frame(self.main_container)
        self.right_panel.pack(side="right", fill="both", expand=True)

        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.default_dir = self.script_dir

        # base directory chosen by user (parent only)
        self.base_save_root: str | None = None
        # per-run directory (set inside on_threshold_and_contours)
        self.save_dir: str | None = None

        # Threshold / contour defaults
        self.threshold_min = 10
        self.threshold_max = 255

        # Separate parameters for hole vs non-hole paths
        self.min_hole_area_holes = 1000

        # Hard minimum area for "large" green blobs in 4c
        self.hard_min_large_green_area = 2000

        # Edge margin (px) for ignoring border and drawing inner box
        self.edge_margin = 10

        # Max radius for detecting connections and projections from skeleton midline
        self.max_r: int | None = None  # radial search radius; None means "auto first time"

        # === Original Overlay Image ===
        self.frame_overlay = ttk.LabelFrame(
            self.left_panel, text="Original Overlay Image"
        )
        self.frame_overlay.pack(padx=20, pady=(20, 10), fill="x")

        self.original_overlay_path = ""

        self.btn_overlay = ttk.Button(
            self.frame_overlay,
            text="Load Overlay",
            command=self.load_original_overlay,
        )
        self.btn_overlay.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.lbl_overlay = ttk.Label(
            self.frame_overlay,
            text="No file selected",
            foreground="gray",
        )
        self.lbl_overlay.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # === Save Output ===
        self.frame_save = ttk.LabelFrame(
            self.left_panel, text="Save Output Directory"
        )
        self.frame_save.pack(padx=20, pady=(0, 20), fill="x")

        self.btn_save = ttk.Button(
            self.frame_save,
            text="Select Save Location",
            command=self.select_save_directory,
        )
        self.btn_save.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.lbl_save = ttk.Label(
            self.frame_save,
            text="No directory selected",
            foreground="gray",
        )
        self.lbl_save.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # === Threshold & Contour ===
        self.frame_params = ttk.LabelFrame(
            self.left_panel, text="Threshold & Contour"
        )
        self.frame_params.pack(padx=20, pady=(0, 10), fill="x")

        # Row 0: Threshold Min / Max
        tk.Label(self.frame_params, text="Threshold Min:").grid(
            row=0, column=0, padx=10, pady=5, sticky="w"
        )
        self.entry_threshold_min = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_min.insert(0, str(self.threshold_min))
        self.entry_threshold_min.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Threshold Max:").grid(
            row=0, column=2, padx=10, pady=5, sticky="w"
        )
        self.entry_threshold_max = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_max.insert(0, str(self.threshold_max))
        self.entry_threshold_max.grid(row=0, column=3, padx=10, pady=5, sticky="w")

        # Row 1: Edge margin
        tk.Label(self.frame_params, text="Edge Margin (px):").grid(
            row=1, column=0, padx=10, pady=5, sticky="w"
        )
        self.entry_edge_margin = ttk.Entry(self.frame_params, width=8)
        self.entry_edge_margin.insert(0, str(self.edge_margin))
        self.entry_edge_margin.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        # Row 2: Hole-specific parameter
        ttk.Label(self.frame_params, text="Min Hole Area:").grid(
            row=2, column=0, padx=10, pady=5, sticky="w"
        )
        self.entry_min_hole_area_holes = ttk.Entry(self.frame_params, width=8)
        self.entry_min_hole_area_holes.insert(0, str(self.min_hole_area_holes))
        self.entry_min_hole_area_holes.grid(
            row=2, column=1, padx=5, pady=5, sticky="w"
        )

        # Row 3: Hard min large green area (4c)
        ttk.Label(self.frame_params, text="Min Cell Area:").grid(
            row=3, column=0, padx=10, pady=5, sticky="w"
        )
        self.entry_hard_min_large_green_area = ttk.Entry(self.frame_params, width=8)
        self.entry_hard_min_large_green_area.insert(
            0, str(self.hard_min_large_green_area)
        )
        self.entry_hard_min_large_green_area.grid(
            row=3, column=1, padx=5, pady=5, sticky="w"
        )

        # Row 4: max_r value for detecting connections and projections from skeleton midline
        ttk.Label(self.frame_params, text="P/C Radius:").grid(
            row=4, column=0, padx=10, pady=5, sticky="w"
        )

        self.spin_max_r = tk.Spinbox(
            self.frame_params,
            from_=1,
            to=200,        # adjust upper bound as you like
            width=6,
            increment=1
        )
        self.spin_max_r.delete(0, "end")
        self.spin_max_r.insert(0, "")  # blank until first run
        self.spin_max_r.grid(row=4, column=1, padx=10, pady=5, sticky="w")

        # Row 5: Threshold / contours button
        self.btn_threshold_and_contours = ttk.Button(
            self.frame_params,
            text="Threshold & Contours & Sampling Points",
            command=self.on_threshold_and_contours,
        )
        self.btn_threshold_and_contours.grid(
            row=5, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5)
        )

        # === Progress / Status on right ===
        self.status_frame = ttk.LabelFrame(
            self.right_panel, text="Progress / Status"
        )
        self.status_frame.pack(padx=10, pady=10, fill="both", expand=True)

        log_container = tk.Frame(self.status_frame)
        log_container.pack(fill="both", expand=True)

        self.status_text = tk.Text(
            log_container,
            height=40,
            width=50,
            state="disabled",
            wrap="word",
        )
        self.status_text.pack(
            side="left", fill="both", expand=True, padx=(5, 0), pady=5
        )

        scrollbar = ttk.Scrollbar(
            log_container, orient="vertical", command=self.status_text.yview
        )
        scrollbar.pack(side="right", fill="y", padx=(0, 5), pady=5)
        self.status_text.configure(yscrollcommand=scrollbar.set)

    # --- window helpers ---

    def safe_after(self, delay_ms, func, *args):
        aid = self.after(delay_ms, func, *args)
        self._after_ids.append(aid)
        return aid

    def on_close(self):
        self._is_closing = True
        for aid in self._after_ids:
            try:
                self.after_cancel(aid)
            except Exception:
                pass
        self.destroy()

    # ------------------ IO callbacks ------------------
    def load_original_overlay(self):
        path = filedialog.askopenfilename(
            title="Select Original Overlay Image",
            filetypes=[("Image files", "*.png *.jpg *.tif *.tiff *.bmp")],
            initialdir=self.default_dir,
        )

        if path:
            self.original_overlay_path = path

            # Update default_dir so future dialogs start here
            overlay_dir = os.path.dirname(path)
            self.default_dir = overlay_dir

            # auto-set base save directory to the folder of the overlay
            self.base_save_root = overlay_dir

            # ---- reset P/C radius state for new overlay ----
            self.max_r = None
            if hasattr(self, "spin_max_r"):
                try:
                    self.spin_max_r.delete(0, "end")
                    # leave blank; threshold_and_contours will auto-fill next run
                except Exception:
                    pass
            self.log_status("P/C radius reset; will be re-estimated on next run.")

            # update labels
            self.lbl_overlay.config(
                text=os.path.basename(path),
                foreground="green",
            )
            self.lbl_save.config(
                text=os.path.basename(overlay_dir),
                foreground="green",
            )

            self.log_status(f"Loaded original overlay: {path}")
            self.log_status(f"Base output directory (auto): {overlay_dir}")

    def select_save_directory(self):
        """
        Optional manual override of the base directory.
        """
        directory = filedialog.askdirectory(initialdir=self.default_dir)
        if directory:
            self.base_save_root = directory
            self.lbl_save.config(
                text=os.path.basename(directory),
                foreground="green",
            )
            self.log_status(f"Base output directory (manual): {directory}")


    def log_status(self, message: str):
        """
        Append a line to the right-hand status text box and
        also print to stdout if you want.
        """
        print(message)

        if getattr(self, "_is_closing", False):
            return
        if not self.winfo_exists():
            return
        if not hasattr(self, "status_text") or not self.status_text.winfo_exists():
            return

        self.status_text.configure(state="normal")
        self.status_text.insert("end", message + "\n")
        self.status_text.see("end")
        self.status_text.configure(state="disabled")

    def show_cv2_image(self, img, title: str = "Preview"):
        """
        Display a cv2 BGR image in a Tk Toplevel window, scaled to fit
        inside the main window while preserving aspect ratio.
        """
        if img is None:
            self.log_status(f"IMAGE {title}: None")
            return
        if getattr(self, "_is_closing", False) or not self.winfo_exists():
            return

        im_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        im = Image.fromarray(im_rgb)

        self.update_idletasks()
        main_w = max(self.winfo_width(), 1)
        main_h = max(self.winfo_height(), 1)

        w, h = im.size
        scale = min(main_w / w, main_h / h)
        new_w, new_h = int(w * scale), int(h * scale)
        im = im.resize((new_w, new_h), Image.Resampling.LANCZOS)

        photo = ImageTk.PhotoImage(im)

        win = tk.Toplevel(self)
        win.title(title)

        label = ttk.Label(win, image=photo)
        label.image = photo
        label.pack(fill="both", expand=True)

        main_x = self.winfo_x()
        main_y = self.winfo_y()
        x = main_x + max((main_w - new_w) // 2, 0)
        y = main_y + max((main_h - new_h) // 2, 0)
        win.geometry(f"{new_w}x{new_h}+{x}+{y}")

        win.transient(self)
        win.grab_set()
        self.wait_window(win)

    def _draw_inner_margin_box(self, overlay, color=(0, 165, 255), thickness=1):
        """
        Draw the analysis margin box defined by self.edge_margin on `overlay`,
        and dim pixels outside the box.
        """

        edge_margin = int(getattr(self, "edge_margin", 0))

        # Get H, W from positive_mask if available, else from overlay
        if getattr(self, "positive_mask", None) is not None:
            H, W = self.positive_mask.shape[:2]
        else:
            H, W = overlay.shape[:2]

        inner_x_min = edge_margin
        inner_y_min = edge_margin
        inner_x_max = W - edge_margin
        inner_y_max = H - edge_margin

        # Clamp in case edge_margin is too large
        inner_x_min = max(0, min(inner_x_min, W))
        inner_y_min = max(0, min(inner_y_min, H))
        inner_x_max = max(0, min(inner_x_max, W))
        inner_y_max = max(0, min(inner_y_max, H))

        if inner_x_min >= inner_x_max or inner_y_min >= inner_y_max:
            # Degenerate box: just return without modifying
            return

        # 1) dim outside region
        dim_factor = 0.3  # 0..1, smaller = darker

        mask = np.zeros((H, W), dtype=np.uint8)
        mask[inner_y_min:inner_y_max, inner_x_min:inner_x_max] = 1  # inside

        inside = mask.astype(bool)
        outside = ~inside

        # Ensure overlay is 3-channel uint8 for safe in-place ops
        if overlay.ndim == 2:
            overlay_rgb = cv2.cvtColor(overlay, cv2.COLOR_GRAY2BGR)
        else:
            overlay_rgb = overlay

        # dim only outside pixels
        overlay_rgb[outside] = (
            overlay_rgb[outside].astype(np.float32) * dim_factor
        ).astype(np.uint8)

        # 2) draw inner rectangle
        top_left = (inner_x_min, inner_y_min)
        bot_right = (inner_x_max - 1, inner_y_max - 1)

        cv2.rectangle(
            overlay_rgb,
            top_left,
            bot_right,
            color=color,
            thickness=thickness,
            lineType=cv2.LINE_AA,
        )

        # If we converted from gray, copy back into the original
        if overlay.ndim == 2:
            overlay[:, :] = cv2.cvtColor(overlay_rgb, cv2.COLOR_BGR2GRAY)
        else:
            overlay[:, :] = overlay_rgb

    # ------------------ Threshold + Contours Callbacks ------------------
    def on_threshold_and_contours(self):
        if not self.original_overlay_path:
            messagebox.showwarning(
                "Missing Overlay",
                "Please load the overlay image before processing.",
            )
            return
        if not self.base_save_root:
            messagebox.showwarning(
                "No Save Directory",
                "Please select a save directory before processing.",
            )
            return

        # Fresh per-run directory in the chosen base directory
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.save_dir = os.path.join(self.base_save_root, f"output_{timestamp}")
        os.makedirs(self.save_dir, exist_ok=True)

        # Copy original overlay into this run directory
        if self.original_overlay_path:
            src = self.original_overlay_path
            dst = os.path.join(self.save_dir, os.path.basename(self.original_overlay_path))
            try:
                import shutil
                shutil.copy2(src, dst)
                self.log_status(f"Copied original overlay to: {dst}")
            except Exception as e:
                self.log_status(f"[WARN] Could not copy original overlay: {e}")

        # self.csv_paths.clear()

        self.log_status(
            "Starting thresholding, contour extraction, and superpixels..."
        )
        self.log_status(f"Run output directory: {self.save_dir}")

        try:
            self.threshold_min = int(self.entry_threshold_min.get())
            self.threshold_max = int(self.entry_threshold_max.get())
            self.edge_margin = int(self.entry_edge_margin.get())
            self.min_hole_area_holes = float(self.entry_min_hole_area_holes.get())
            self.min_large_green_area = float(
                self.entry_hard_min_large_green_area.get()
            )
        except ValueError as e:
            messagebox.showerror(
                "Invalid Input",
                f"Please enter valid numeric values:\n{e}",
            )
            self.log_status(f"[ERROR] Invalid input: {e}")
            return

        try:
            # Save directly into the per-run directory (no 'processed' subfolder)
            self.threshold_and_contours(
                self.original_overlay_path,
                self.save_dir,
                title="threshold overlay",
                short_tag="overlay",
                min_hole_area_holes=self.min_hole_area_holes,
                edge_margin=self.edge_margin,
            )
            self.log_status("Thresholding + Surface + Cavity Paths Complete.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_status(f"[ERROR] Thresholding + Surface + Cavity Paths: {e}")
        finally:
            cv2.destroyAllWindows()

    def estimate_closing_radius(self, pos_mask,
                                frac=0.4,
                                min_radius=2,
                                max_radius=20):
        """
        Estimate a morphological closing radius from object thickness.
        pos_mask: uint8 0/255, foreground = 255.
        """
        pos_bin = (pos_mask > 0).astype(np.uint8)
        if pos_bin.sum() == 0:
            return min_radius

        dist = cv2.distanceTransform(pos_bin, cv2.DIST_L2, 5)  # [web:35][web:72]
        vals = dist[dist > 0]
        if vals.size == 0:
            return min_radius

        median_half_width = float(np.median(vals))
        r = int(round(frac * median_half_width))
        r = max(min_radius, min(r, max_radius))
        return r

    def compute_combined_far_pink_score(
        self,
        output_dir,
        short_tag="score",
        projection_axis_ratio_cutoff: float = 1.4,
        contact_radius: int = 10,
    ):
        """
        Cluster-first bridge/projection/protrusion classifier.

        Rules:
        - bridge: n_reachable >= 2
        - projection: n_reachable < 2 and axis_ratio >= projection_axis_ratio_cutoff
        - protrusion: n_reachable < 2 and axis_ratio < projection_axis_ratio_cutoff

        Scoring:
        - bridge = 3 points
        - projection = 1 point
        - protrusion = 0 points

        Also computes per-component:
        - axis_ratio
        - skeleton midline length in pixels
        - skeleton midline length in mm
        - farthest distance of that component's far-pink pixels to reachable green (px, mm)
        """

        far_bin = (self.far_pink_mask > 0).astype(np.uint8)
        green_bin = (self.positive_no_interface_mask > 0).astype(np.uint8)

        pink_area = int((self.interface_mask > 0).astype(np.uint8).sum())
        far_area = int(far_bin.sum())
        fraction_far_pink_interface = far_area / float(pink_area) if pink_area > 0 else 0.0

        bridge_mask = np.zeros_like(far_bin, dtype=np.uint8)
        projection_mask = np.zeros_like(far_bin, dtype=np.uint8)
        protrusion_mask = np.zeros_like(far_bin, dtype=np.uint8)

        if far_area == 0:
            self.component_info = []
            self.bridge_projection_score_points = 0
            self.max_far_pink_distance_px = 0.0
            self.max_far_pink_distance_mm = 0.0
            return (
                fraction_far_pink_interface,
                0,
                0,
                0,
                0,
                bridge_mask,
                projection_mask,
                protrusion_mask,
            )

        dpi = getattr(self, "dpi_x", None) or getattr(self, "dpi_y", None)
        mm_per_pixel = 25.4 / float(dpi) if dpi else 1.0

        contact_kernel = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (2 * contact_radius + 1, 2 * contact_radius + 1)
        )

        num_green, green_labels, green_stats, _ = cv2.connectedComponentsWithStats(
            green_bin, connectivity=8
        )

        green_area = np.zeros(num_green, dtype=int)
        for lbl in range(1, num_green):
            green_area[lbl] = green_stats[lbl, cv2.CC_STAT_AREA]

        nonzero = green_area[1:][green_area[1:] > 0]
        median_area = float(np.median(nonzero)) if nonzero.size > 0 else 0.0
        hard_min = float(getattr(self, "min_large_green_area", 0))
        large_thresh = max(hard_min, 2.0 * median_area if median_area > 0 else 0.0)

        is_large_green = np.zeros(num_green, dtype=bool)
        n_is_large_green = 0
        for lbl in range(1, num_green):
            is_large_green[lbl] = green_area[lbl] >= large_thresh
            if is_large_green[lbl]:
                print("Positive mask area for cell (or cell cluster):", green_area[lbl], large_thresh)
                n_is_large_green += 1

        # visualize which green components are considered "large"
        large_green_mask = np.zeros_like(green_bin, dtype=np.uint8)
        for lbl in range(1, num_green):
            if is_large_green[lbl]:
                large_green_mask[green_labels == lbl] = 255

        # optional overlay for visual context
        large_overlay = self.img_color.copy()
        large_overlay[large_green_mask > 0] = (0, 255, 0)  # bright green

        # out_large = os.path.join(output_dir, f"{short_tag}_large_green_mask.png")
        out_large_ov = os.path.join(output_dir, f"3-cell_overlay.png")
        # cv2.imwrite(out_large, large_green_mask)
        cv2.imwrite(out_large_ov, large_overlay)

        if hasattr(self, "show_cv2_image"):
            # self.show_cv2_image(large_green_mask, title="Large green mask")
            self.show_cv2_image(large_overlay, title="Large green overlay")

        # is_large_green = np.zeros(num_green, dtype=bool)
        # for lbl in range(1, num_green):
        #     is_large_green[lbl] = green_area[lbl] >= 200

        # ---------- helper: skeletonization ----------
        def skeletonize_bin(bin_u8):
            skel = np.zeros_like(bin_u8, dtype=np.uint8)
            element = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
            work = bin_u8.copy()
            while True:
                eroded = cv2.erode(work, element)
                opened = cv2.morphologyEx(eroded, cv2.MORPH_OPEN, element)
                temp = cv2.subtract(eroded, opened)
                skel = cv2.bitwise_or(skel, temp)
                work = eroded.copy()
                if cv2.countNonZero(work) == 0:
                    break
            return skel

        # ---------- helper: find skeleton endpoints for a component ----------
        def skeleton_endpoints(skel):
            """
            Given a binary skeleton mask (0/1), return a list of endpoints
            (pixels with exactly one 8-connected skeleton neighbor).
            """
            ys, xs = np.where(skel > 0)
            endpoints = []
            if ys.size == 0:
                return endpoints

            Hs, Ws = skel.shape
            for y, x in zip(ys, xs):
                y0 = max(0, y - 1)
                y1 = min(Hs, y + 2)
                x0 = max(0, x - 1)
                x1 = min(Ws, x + 2)
                block = skel[y0:y1, x0:x1]
                count = int(block.sum())  # includes (y, x)
                if count == 1:
                    endpoints.append((y, x))
            return endpoints

        # ---------- helper: reachable large-green labels via dilation ----------
        def reachable_large_green_labels_simple(comp_mask, contact_radius=10):
            """
            For a single far-pink component, find all large-green labels that
            come within contact_radius pixels of the component via dilation.
            """
            kernel = cv2.getStructuringElement(
                cv2.MORPH_ELLIPSE,
                (2 * contact_radius + 1, 2 * contact_radius + 1),
            )
            reach_mask = cv2.dilate(comp_mask, kernel)

            ys, xs = np.where(reach_mask > 0)
            labels_hit = set()
            for y, x in zip(ys, xs):
                gl = int(green_labels[y, x])
                if gl != 0 and is_large_green[gl]:
                    labels_hit.add(gl)

            return labels_hit

        # ---------- helper: per-component shape metrics ----------
        def component_metrics(comp_mask):
            skel = skeletonize_bin(comp_mask)
            ys, xs = np.where(skel > 0)
            if ys.size == 0:
                return 0.0, 0.0, 0.0

            y_span = int(ys.max() - ys.min())
            x_span = int(xs.max() - xs.min())
            long_axis = max(y_span, x_span)
            short_axis = max(1, min(y_span, x_span))
            axis_ratio = long_axis / float(short_axis)

            length_px = float(
                np.sum(
                    np.hypot(
                        np.diff(xs.astype(np.float32)),
                        np.diff(ys.astype(np.float32)),
                    )
                )
            )
            if length_px <= 0.0:
                length_px = float(long_axis)

            length_mm = length_px * mm_per_pixel
            return float(axis_ratio), float(length_px), float(length_mm)

        # ---------- helper: farthest distance to reachable green (per component) ----------
        def farthest_distance_to_reachable_green(comp_mask_u8, reachable_labels):
            """
            For a single far-pink component and its set of reachable green labels,
            compute the maximum shortest distance from any pixel in the component
            to the union of those reachable green blobs.
            Returns (max_dist_px, max_dist_mm).
            """
            if not reachable_labels:
                return 0.0, 0.0

            # mask of only the reachable green blobs
            reachable_green_mask = np.zeros_like(green_labels, dtype=np.uint8)
            for gl in reachable_labels:
                reachable_green_mask[green_labels == gl] = 1

            reachable_inv = (reachable_green_mask == 0).astype(np.uint8)
            dist_local = cv2.distanceTransform(reachable_inv, cv2.DIST_L2, 3)

            ys_fp, xs_fp = np.where(comp_mask_u8 > 0)
            if ys_fp.size == 0:
                return 0.0, 0.0

            dists_px = dist_local[ys_fp, xs_fp]
            max_dist_px = float(dists_px.max()) if dists_px.size > 0 else 0.0
            max_dist_mm = max_dist_px * mm_per_pixel
            return max_dist_px, max_dist_mm

        # ---------- connected components over far-pink ----------
        num_far, far_labels = cv2.connectedComponents(far_bin)

        debug_vis = np.zeros((far_bin.shape[0], far_bin.shape[1], 3), dtype=np.uint8)
        rng = np.random.default_rng(7)
        palette = rng.integers(40, 255, size=(num_far, 3), dtype=np.uint8)
        palette[0] = [0, 0, 0]

        green_vis = np.zeros((far_bin.shape[0], far_bin.shape[1], 3), dtype=np.uint8)
        green_vis[green_bin > 0] = (0, 180, 0)

        component_info = []
        n_bridge_regions = 0
        n_projection_regions = 0
        n_protrusion_regions = 0
        score_points = 0

        max_far_px_global = 0.0
        max_far_mm_global = 0.0

        min_bridge_area_px = 200  # tune or derive from physical_to_pixels

        for lbl in range(1, num_far):
            comp = (far_labels == lbl).astype(np.uint8)
            comp_area = int(comp.sum())
            if comp_area == 0:
                continue

            ys, xs = np.where(comp > 0)
            debug_vis[ys, xs] = palette[lbl]

            axis_ratio, length_px, length_mm = component_metrics(comp)
            reachable = reachable_large_green_labels_simple(comp)
            n_reachable = len(reachable)

            comp_max_dist_px, comp_max_dist_mm = farthest_distance_to_reachable_green(comp, reachable)

            # bridge = touches at least two distinct large blobs AND exceeds area threshold
            is_bridge = (n_reachable >= 2) and (comp_area >= min_bridge_area_px)

            if is_bridge and comp_max_dist_mm > 2:
                bridge_mask[ys, xs] = 255
                n_bridge_regions += 1
                score_points += 10 * (n_reachable-1) # award for bridge        
                score_points += (comp_max_dist_mm/2) * (n_reachable-1) # award for length of bridge
                component_type = "bridge"
            elif comp_max_dist_mm > 20:
                projection_mask[ys, xs] = 255
                n_projection_regions += 1 
                score_points += 0.5 * (comp_max_dist_mm/5)
                component_type = "projection"
            else:
                protrusion_mask[ys, xs] = 255
                n_protrusion_regions += 1
                component_type = "protrusion"
                score_points += 0.05

            print(
                component_type,
                "area", comp_area,
                "n_reachable", n_reachable,
                "length_mm", length_mm,
                "max_dist_mm", comp_max_dist_mm,
                "axis_ratio", axis_ratio,
                "reachable", sorted(reachable),
            )

            component_info.append(
                (
                    lbl,
                    comp_area,
                    component_type,
                    float(axis_ratio),
                    float(length_px),
                    float(length_mm),
                    sorted(reachable),
                    float(comp_max_dist_px),
                    float(comp_max_dist_mm),
                    n_is_large_green,
                )
            )

        # combined_debug = cv2.addWeighted(green_vis, 0.35, debug_vis, 0.65, 0)
        combined_debug = cv2.addWeighted(green_vis, 1, debug_vis, 1, 0)

        self.farpink_components_debug = debug_vis
        self.farpink_green_debug = combined_debug
        self.component_info = component_info
        self.bridge_projection_score_points = score_points
        self.max_far_pink_distance_px = max_far_px_global
        self.max_far_pink_distance_mm = max_far_mm_global
        self.log_status(
            f"Max far-pink distance to reachable green: {max_far_px_global:.2f} px, {max_far_mm_global:.3f} mm"
        )


        if hasattr(self, "show_cv2_image"):
            # self.show_cv2_image(debug_vis, title="Far-pink connected components")
            self.show_cv2_image(combined_debug, title="Components plus green mask")

            # save the same debug images
            # fp_path = os.path.join(output_dir, f"{short_tag}_far_pink_components.png")
            cg_path = os.path.join(output_dir, f"4-cell_and_bpp_mask.png")
            # cv2.imwrite(fp_path, debug_vis)
            cv2.imwrite(cg_path, combined_debug)
            # self.log_status(f"{short_tag}: far-pink components saved: {fp_path}")
            self.log_status(f"{short_tag}: components+green saved: {cg_path}")

        return (
            fraction_far_pink_interface,
            n_bridge_regions,
            n_projection_regions,
            n_protrusion_regions,
            score_points,
            score_points/n_is_large_green,
            bridge_mask,
            projection_mask,
            protrusion_mask,
        )

    def export_run_summary_excel(
        self,
        output_excel_path: str,
        fraction_far_pink_interface: float,
        n_bridge_regions: int,
        n_projection_regions: int,
        n_protrusion_regions: int,
        score_points: float,
        score_points_per_cell: float,
    ):
        """
        Export run parameters and far-pink metrics to an Excel file with two sheets.

        Sheet 1: parameters — settings and major run variables
        Sheet 2: bpp_metrics — bridge/projection/protrusion fractions and point score
        """
        import os
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        params_data = [
            {"parameter": "threshold_min", "value": self.threshold_min},
            {"parameter": "threshold_max", "value": self.threshold_max},
            {"parameter": "edge_margin", "value": self.edge_margin},
            {"parameter": "min_hole_area_holes", "value": self.min_hole_area_holes},
            {"parameter": "hard_min_large_green_area", "value": self.hard_min_large_green_area},
            {"parameter": "min_large_green_area", "value": getattr(self, "min_large_green_area", None)},
            {"parameter": "max_r", "value": getattr(self, "max_r", None)},
            {
                "parameter": "image_height",
                "value": getattr(self, "img_color", None).shape[0]
                if getattr(self, "img_color", None) is not None
                else None,
            },
            {
                "parameter": "image_width",
                "value": getattr(self, "img_color", None).shape[1]
                if getattr(self, "img_color", None) is not None
                else None,
            },
        ]
        params_df = pd.DataFrame(params_data)

        metrics_data = [
            #{"metric": "bpp_fraction_interface", "value": fraction_far_pink_interface},
            {"metric": "bpp_bridge_region_count", "value": n_bridge_regions},
            {"metric": "bpp_projection_region_count", "value": n_projection_regions},
            {"metric": "bpp_protrusion_region_count", "value": n_protrusion_regions},
            {"metric": "bpp_score_points", "value": score_points},
            {"metric": "bpp_score_points_per_cell", "value": score_points_per_cell},
        ]
        metrics_df = pd.DataFrame(metrics_data)

        def adjust_column_widths(worksheet):
            """Auto-adjust column widths based on content."""
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

        if os.path.exists(output_excel_path):
            book = load_workbook(output_excel_path)
            with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
                writer.book = book
                if "parameters" in book.sheetnames:
                    del book["parameters"]
                if "bpp_metrics" in book.sheetnames:
                    del book["bpp_metrics"]
                params_df.to_excel(writer, sheet_name="parameters", index=False)
                metrics_df.to_excel(writer, sheet_name="bpp_metrics", index=False)
                
                # Adjust column widths
                adjust_column_widths(writer.sheets["parameters"])
                adjust_column_widths(writer.sheets["bpp_metrics"])
                
                writer.save()
        else:
            with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
                params_df.to_excel(writer, sheet_name="parameters", index=False)
                metrics_df.to_excel(writer, sheet_name="bpp_metrics", index=False)
                
                # Adjust column widths
                adjust_column_widths(writer.sheets["parameters"])
                adjust_column_widths(writer.sheets["bpp_metrics"])

    # ------------------ Convert physical to pixel for calculation fairness ------------------

    # Convert a real-world measurement into pixels using the image DPI.
    # This is necessary because the image-processing code below works in pixels,
    # but we want the analysis to mean the same thing in the real world for every image.
    #
    # For example, a 5 mm neighborhood should represent the same physical distance
    # whether the image is small or large, or whether the image has many pixels or few.
    # Without this conversion, the same pixel radius could accidentally mean a much
    # larger or smaller real-world distance from one image to the next.
    #
    # So the workflow is:
    # 1) choose the physical size we want to use,
    # 2) read the image DPI,
    # 3) convert that physical size into pixels,
    # 4) use the pixel result in the contour, radius, and adjacency calculations.
    #
    # This keeps the processing fair and consistent across images with different
    # resolutions or dimensions.

    def get_image_dpi(self, path):
        try:
            with Image.open(path) as im:
                dpi = im.info.get("dpi", None)
                if dpi and len(dpi) >= 2:
                    return float(dpi[0]), float(dpi[1])
        except Exception:
            pass
        return None, None

    def physical_to_pixels(self, value, dpi, unit="mm"):

        if dpi is None:
            return None
        if unit == "in":
            return int(round(value * dpi))
        if unit == "mm":
            return int(round((value / 25.4) * dpi))
        raise ValueError("unit must be 'in' or 'mm'")

    def threshold_and_contours(
        self,
        path,
        output_dir,
        title="threshold",
        short_tag="overlay",
        min_hole_area_holes=500,
        edge_margin=20,
        min_max_r=14,
    ):
        """
        1) Threshold image
        2) Positive mask + fill small holes
        3) Negative mask as complement, cleaned and smoothed
        4) Extract interface band from NEGATIVE ONLY
        5) Convert band to one or more globally-numbered, clockwise paths
           with a single global index sequence used for CO (|i-j|).
        6) Extract thicker portions of the interface band into a separate overlay.
        """

        import time
        start_time = time.time()

        # ensure attribute exists even if something reads it early
        self.far_pink_mask = None

        # ---------- helper: fill small holes in a binary mask ----------
        def fill_small_holes(mask_bin, min_hole_area_local):
            mask_bin = (mask_bin > 0).astype(np.uint8) * 255
            ys, xs = np.where(mask_bin > 0)
            if len(xs) == 0:
                return mask_bin

            y_min, y_max = ys.min(), ys.max()
            x_min, x_max = xs.min(), xs.max()
            roi = mask_bin[y_min:y_max + 1, x_min:x_max + 1]

            inv_roi = cv2.bitwise_not(roi)
            num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(
                inv_roi, connectivity=8
            )

            for lbl in range(1, num_labels):
                area = stats[lbl, cv2.CC_STAT_AREA]
                if area < min_hole_area_local:
                    roi[labels == lbl] = 255
                else:
                    roi[labels == lbl] = 0

            mask_bin[y_min:y_max + 1, x_min:x_max + 1] = roi
            return mask_bin

        # ---------- I/O, DPI, grayscale and color load ----------
        self.log_status(f"Thresholding {os.path.basename(path)} as {short_tag}...")

        dpi_x, dpi_y = self.get_image_dpi(path)
        self.dpi_x = dpi_x
        self.dpi_y = dpi_y
        print(f"{short_tag}: DPI = {dpi_x} x {dpi_y}")
        self.log_status(f"{short_tag}: DPI = {dpi_x} x {dpi_y}")

        dpi = dpi_x if dpi_x is not None else dpi_y

        img_gray = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        self.img_color = cv2.imread(path)
        if img_gray is None or self.img_color is None:
            raise ValueError(f"Failed to read image: {path}")

        # ---------- positive/negative masks + basic cleanup ----------
        _, thresh = cv2.threshold(
            img_gray,
            self.threshold_min,
            self.threshold_max,
            cv2.THRESH_BINARY,
        )

        min_positive_area = 50
        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(thresh, connectivity=4)
        areas = stats[:, cv2.CC_STAT_AREA]
        keep = areas >= min_positive_area
        keep[0] = False
        clean = np.where(keep[labels], 255, 0).astype(np.uint8)

        self.positive_mask = fill_small_holes(clean, min_hole_area_holes)

        negative_mask = (self.positive_mask == 0).astype(np.uint8) * 255
        negative_mask = fill_small_holes(negative_mask, min_hole_area_holes)

        blur = cv2.GaussianBlur(negative_mask, (5, 5), 0)
        _, negative_mask = cv2.threshold(blur, 127, 255, cv2.THRESH_BINARY)
        negative_mask = fill_small_holes(negative_mask, min_hole_area_holes)

        est_radius = self.estimate_closing_radius(self.positive_mask, frac=0.6)

        # ---------- interface baseline from NEGATIVE mask ----------
        baseline_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))

        neg_core = (negative_mask > 0).astype(np.uint8)
        neg_edge = cv2.morphologyEx(neg_core, cv2.MORPH_GRADIENT, baseline_kernel)
        interface_bin = neg_edge.astype(np.uint8)

        interface_bin = cv2.morphologyEx(
            interface_bin,
            cv2.MORPH_CLOSE,
            baseline_kernel,
            iterations=1,
        )

        pos_bin = (self.positive_mask > 0).astype(np.uint8)
        pos_inv = (pos_bin == 0).astype(np.uint8) * 255
        dist_to_pos = cv2.distanceTransform(pos_inv, cv2.DIST_L2, 3)

        max_offset_px = self.physical_to_pixels(0.5, dpi, unit="mm")
        thin_mask = (dist_to_pos <= max_offset_px).astype(np.uint8)

        baseline = cv2.bitwise_and(interface_bin, thin_mask)
        baseline = (baseline > 0).astype(np.uint8)
        H, W = baseline.shape

        # ---------- green skeleton midline (positive mask skeleton) ----------
        pos_u8 = (pos_bin * 255).astype(np.uint8)

        skel = np.zeros_like(pos_u8)
        element = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))

        done = False
        work = pos_u8.copy()
        while not done:
            eroded = cv2.erode(work, element)
            opened = cv2.morphologyEx(eroded, cv2.MORPH_OPEN, element)
            temp = cv2.subtract(eroded, opened)
            skel = cv2.bitwise_or(skel, temp)
            work = eroded.copy()
            if cv2.countNonZero(work) == 0:
                done = True

        green_midline = (skel > 0).astype(np.uint8)

        ys_mid, xs_mid = np.where(green_midline > 0)
        mid_points = list(zip(ys_mid, xs_mid))

        num_angles = 48
        angles = [2.0 * np.pi * k / num_angles for k in range(num_angles)]

        # ---------- bridge (yellow) search radius configuration ----------
        if self.max_r is None:
            auto_max_r = int(round(1.5 * est_radius))
            auto_max_r = max(min_max_r, min(auto_max_r, 70))
            self.max_r = auto_max_r
            if hasattr(self, "spin_max_r"):
                self.spin_max_r.delete(0, "end")
                self.spin_max_r.insert(0, str(self.max_r))
        else:
            try:
                gui_max_r = int(self.spin_max_r.get())
                gui_max_r = max(min_max_r, gui_max_r)
                self.max_r = gui_max_r
            except Exception:
                pass

        max_r = self.max_r

        # ---------- midline-based bridge drawing (bridges mask) ----------
        bridges = np.zeros_like(baseline, dtype=np.uint8)

        for (y0, x0) in mid_points:
            if baseline[y0, x0] > 0:
                continue

            for theta in angles:
                dy = np.sin(theta)
                dx = np.cos(theta)

                hit1 = None
                hit2 = None

                for step in range(1, max_r + 1):
                    y_f = y0 + dy * step
                    x_f = x0 + dx * step
                    yi = int(round(y_f))
                    xi = int(round(x_f))
                    if yi < 0 or yi >= H or xi < 0 or xi >= W:
                        break
                    if baseline[yi, xi] > 0:
                        hit1 = (yi, xi)
                        break

                for step in range(1, max_r + 1):
                    y_f = y0 - dy * step
                    x_f = x0 - dx * step
                    yi = int(round(y_f))
                    xi = int(round(x_f))
                    if yi < 0 or yi >= H or xi < 0 or xi >= W:
                        break
                    if baseline[yi, xi] > 0:
                        hit2 = (yi, xi)
                        break

                if hit1 is not None and hit2 is not None:
                    y1, x1 = hit1
                    y2, x2 = hit2
                    if pos_bin[y0, x0] == 0:
                        continue
                    cv2.line(bridges, (x1, y1), (x2, y2), 1, 1)

        # ---------- interface band, islands, and final interface mask ----------
        interface_bin_tmp = cv2.bitwise_or(baseline, bridges).astype(np.uint8)

        band_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        interface_band = cv2.dilate(interface_bin_tmp, band_kernel)

        if getattr(self, "enable_interface_islands", True):
            candidates = np.zeros_like(pos_bin, dtype=np.uint8)
            candidates[(pos_bin > 0) & (interface_band > 0) & (interface_bin_tmp == 0)] = 1

            num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(
                candidates, connectivity=8
            )

            max_island_area = self.physical_to_pixels(1.0, dpi, unit="mm") ** 2
            islands = np.zeros_like(candidates, dtype=np.uint8)
            for lbl in range(1, num_labels):
                area = stats[lbl, cv2.CC_STAT_AREA]
                if 0 < area <= max_island_area:
                    islands[labels == lbl] = 1

            bridges = cv2.bitwise_or(bridges, islands)

        # raw interface line (1‑pixel where baseline/bridges are)
        final_interface_bin = cv2.bitwise_or(baseline, bridges)
        interface_mask = (final_interface_bin > 0).astype(np.uint8) * 255
        self.interface_mask = interface_mask.copy()

        # ---------- green-without-interface mask (same as before) ----------
        pos_bin_full = (self.positive_mask > 0).astype(np.uint8)
        int_bin_full = (interface_mask > 0).astype(np.uint8)

        pos_no_if = pos_bin_full.copy()
        pos_no_if[int_bin_full > 0] = 0
        self.positive_no_interface_mask = (pos_no_if > 0).astype(np.uint8) * 255

        # ---------- chips = small positives in positive_no_interface_mask ----------
        pos_no_if_bin = (self.positive_no_interface_mask > 0).astype(np.uint8)

        kernel = np.ones((3, 3), np.uint8)
        pos_no_if_clean = cv2.morphologyEx(pos_no_if_bin, cv2.MORPH_OPEN, kernel)

        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(
            pos_no_if_clean, connectivity=4
        )
        areas = stats[:, cv2.CC_STAT_AREA]
        print("areas:", areas)

        min_chip_area = 5
        max_chip_area = 2000

        chips_filtered = np.zeros_like(pos_no_if_clean, dtype=np.uint8)

        for lbl in range(1, num_labels):
            area = int(areas[lbl])
            if min_chip_area <= area <= max_chip_area:
                chips_filtered[labels == lbl] = 255

        print("chips_filtered sum:", int(chips_filtered.sum()))

        # out_chips = os.path.join(output_dir, f"chips_mask.png")
        # cv2.imwrite(out_chips, chips_filtered)
        # self.log_status(f"{short_tag}: chips mask saved: {out_chips}")

        # ---------- promote chips: add (slightly dilated) chip regions to interface ----------
        chip_band = (chips_filtered > 0).astype(np.uint8)

        # dilate to thicken / fill outlines of chips
        kernel = np.ones((7, 7), np.uint8)          # tune size if needed
        chip_band_dil = cv2.dilate(chip_band, kernel, iterations=2)

        # add dilated chip regions into interface band
        interface_band_filled = interface_band.copy().astype(np.uint8)
        interface_band_filled[chip_band_dil > 0] = 1

        # remove those pixels from positive_no_interface
        pos_no_if_updated = pos_no_if_bin.copy()
        pos_no_if_updated[chip_band_dil > 0] = 0

        # ---------- update masks ----------
        final_interface_bin_extended = interface_band_filled
        interface_mask = (final_interface_bin_extended > 0).astype(np.uint8) * 255
        self.interface_mask = interface_mask.copy()
        self.positive_no_interface_mask = (pos_no_if_updated > 0).astype(np.uint8) * 255

        # ---------- overlay using updated positive_no_interface_mask ----------
        overlay = self.img_color.copy()
        pos_bin_no_interface = (self.positive_no_interface_mask > 0).astype(np.uint8)
        int_bin = (interface_mask > 0).astype(np.uint8)

        overlay[pos_bin_no_interface > 0] = (0, 64, 0)

        ys_if, xs_if = np.where(int_bin > 0)
        for y, x in zip(ys_if, xs_if):
            cv2.circle(
                overlay,
                (int(x), int(y)),
                radius=1,
                color=(0, 255, 255),
                thickness=-1,
                lineType=cv2.LINE_AA,
            )

        base = short_tag
        self._draw_inner_margin_box(overlay)
        out_overlay = os.path.join(output_dir, f"1-cell_and_interface_mask.png")
        cv2.imwrite(out_overlay, overlay)
        self.log_status(f"{short_tag}: overlay saved: {out_overlay}")

        # debug dumps
        # cv2.imwrite(os.path.join(output_dir, f"cell_mask.png"),
        #             self.positive_no_interface_mask)
        # cv2.imwrite(os.path.join(output_dir, f"interface_mask.png"), final_interface_bin * 255)
        # cv2.imwrite(os.path.join(output_dir, f"cell_and_interface_mask.png"), pos_bin_full * 255)
        # cv2.imwrite(os.path.join(output_dir, f"{short_tag}_interface_band.png"), interface_band * 255)

    # ---------- the rest of your quantification / far-pink code goes here ----------
    # (reuse your existing code from "label green blobs (no interface) and basic size stats"
    #  down to the end; it will now see the updated self.positive_no_interface_mask
    #  and self.interface_mask, and far_pink_mask is initialized at the top.)

        # ---------- label green blobs (no interface) and basic size stats ----------
        # H, W = pos_bin_no_interface.shape

        num_glabels, g_labels, g_stats, _ = cv2.connectedComponentsWithStats(
            pos_bin_no_interface.astype(np.uint8), connectivity=8
        )

        green_area = np.zeros(num_glabels, dtype=int)
        for lbl in range(1, num_glabels):
            green_area[lbl] = g_stats[lbl, cv2.CC_STAT_AREA]

        nonzero_areas = green_area[1:][green_area[1:] > 0]

        min_area_for_scale = self.physical_to_pixels(0.15, dpi, unit="mm") ** 2
        scale_areas = nonzero_areas[nonzero_areas >= min_area_for_scale]

        if scale_areas.size > 0:
            median_area = float(np.median(scale_areas))
        else:
            median_area = 0.0

        scale_factor = 20
        if median_area > 0.0:
            scale_based_min = scale_factor * median_area
        else:
            scale_based_min = 0.0

        hard_min_large_green_area = self.min_large_green_area

        if scale_based_min > hard_min_large_green_area:
            min_large_green_area = int(hard_min_large_green_area)
        else:
            min_large_green_area = int(hard_min_large_green_area)

        is_large_green = np.zeros(num_glabels, dtype=bool)
        for lbl in range(1, num_glabels):
            is_large = green_area[lbl] >= min_large_green_area
            is_large_green[lbl] = is_large

        # ---------- label interface (pink) blobs and attach to green ----------
        num_plabels, p_labels, _, _ = cv2.connectedComponentsWithStats(
            int_bin.astype(np.uint8), connectivity=8
        )

        touch_radius = self.physical_to_pixels(0.2, dpi, unit="mm")

        pink_attached_to_large_green = np.zeros(num_plabels, dtype=bool)

        t_offsets = []
        for dy in range(-touch_radius, touch_radius + 1):
            for dx in range(-touch_radius, touch_radius + 1):
                if dy == 0 and dx == 0:
                    continue
                t_offsets.append((dy, dx))

        for pl in range(1, num_plabels):
            ys_p, xs_p = np.where(p_labels == pl)
            if ys_p.size == 0:
                continue

            attached = False
            for y0, x0 in zip(ys_p, xs_p):
                for dy, dx in t_offsets:
                    y = y0 + dy
                    x = x0 + dx
                    if 0 <= y < H and 0 <= x < W:
                        gl = g_labels[y, x]
                        if gl != 0 and is_large_green[gl]:
                            attached = True
                            break
                if attached:
                    break

            pink_attached_to_large_green[pl] = attached

        pink_area = np.zeros(num_plabels, dtype=int)
        for pl in range(1, num_plabels):
            pink_area[pl] = np.sum(p_labels == pl)

        attached_green_area_all = np.zeros(num_plabels, dtype=float)
        attached_green_area_large = np.zeros(num_plabels, dtype=float)
        interface_to_green_ratio_all = np.full(num_plabels, np.nan, dtype=float)
        interface_to_green_ratio_large = np.full(num_plabels, np.nan, dtype=float)

        # ---------- per-interface-blob attached green area and ratios ----------
        for pl in range(1, num_plabels):
            ys_p, xs_p = np.where(p_labels == pl)
            if ys_p.size == 0:
                continue

            attached_green_labels_all = set()
            attached_green_labels_large = set()

            for y0, x0 in zip(ys_p, xs_p):
                for dy, dx in t_offsets:
                    y = y0 + dy
                    x = x0 + dx
                    if 0 <= y < H and 0 <= x < W:
                        gl = g_labels[y, x]
                        if gl != 0:
                            attached_green_labels_all.add(gl)
                            if is_large_green[gl]:
                                attached_green_labels_large.add(gl)

            if not attached_green_labels_all:
                attached_green_area_all[pl] = 0.0
            else:
                total_all = sum(green_area[gl] for gl in attached_green_labels_all)
                attached_green_area_all[pl] = float(total_all)
                if total_all > 0:
                    interface_to_green_ratio_all[pl] = pink_area[pl] / float(total_all)
                else:
                    interface_to_green_ratio_all[pl] = np.inf

            if attached_green_labels_large:
                total_large = sum(green_area[gl] for gl in attached_green_labels_large)
                attached_green_area_large[pl] = float(total_large)
                if total_large > 0:
                    interface_to_green_ratio_large[pl] = pink_area[pl] / float(total_large)
                else:
                    interface_to_green_ratio_large[pl] = np.inf
            else:
                attached_green_area_large[pl] = 0.0
                interface_to_green_ratio_large[pl] = np.nan

        # ---------- filter interface to get int_bin_filtered (scale-/ratio-aware) ----------
        min_attached_green_area_all = self.physical_to_pixels(0.3, dpi, unit="mm") ** 2
        max_interface_to_green_all = 1.0

        int_bin_filtered = int_bin.copy()
        for pl in range(1, num_plabels):
            ga_all = attached_green_area_all[pl]
            ga_large = attached_green_area_large[pl]
            ratio_all = interface_to_green_ratio_all[pl]

            if ga_large > 0.0:
                continue

            if ga_all < min_attached_green_area_all:
                int_bin_filtered[p_labels == pl] = 0
                continue

            if np.isfinite(ratio_all) and ratio_all > max_interface_to_green_all:
                int_bin_filtered[p_labels == pl] = 0

        # ---------- far-pink classification (low local green density) ----------
        radius = self.physical_to_pixels(5.0, dpi, unit="mm")
        max_green_density = 0.25

        rr = int(radius)
        offsets = []
        for dy in range(-rr, rr + 1):
            for dx in range(-rr, rr + 1):
                if dx * dx + dy * dy <= radius * radius:
                    offsets.append((dy, dx))
        circle_area = float(len(offsets))

        far_pink_mask = np.zeros_like(int_bin_filtered, dtype=np.uint8)

        ys_pink, xs_pink = np.where(int_bin_filtered > 0)
        for y0, x0 in zip(ys_pink, xs_pink):
            pl = p_labels[y0, x0]
            if pl == 0:
                continue
            if not pink_attached_to_large_green[pl]:
                continue

            green_count = 0
            for dy, dx in offsets:
                y = y0 + dy
                x = x0 + dx
                if 0 <= y < H and 0 <= x < W:
                    if pos_bin_no_interface[y, x] > 0:
                        green_count += 1

            green_density = green_count / circle_area
            if green_density <= max_green_density:
                far_pink_mask[y0, x0] = 255

        kernel_far = np.ones((3, 3), np.uint8)
        far_pink_mask = cv2.morphologyEx(far_pink_mask, cv2.MORPH_OPEN, kernel_far, iterations=1)
        self.far_pink_mask = far_pink_mask.copy()

        # ---------- combined overlay (interface + far pink) ----------
        combined = self.img_color.copy()

        ys_if, xs_if = np.where(int_bin_filtered > 0)
        for y, x in zip(ys_if, xs_if):
            cv2.circle(
                combined,
                (int(x), int(y)),
                radius=1,
                color=(0, 255, 255),
                thickness=-1,
                lineType=cv2.LINE_AA,
            )

        alpha = 0.6
        ys_far, xs_far = np.where(far_pink_mask > 0)
        for y, x in zip(ys_far, xs_far):
            y_i = int(y)
            x_i = int(x)
            bg = combined[y_i, x_i].astype(np.float32)
            fg = np.array([0, 0, 255], dtype=np.float32)
            blended = alpha * fg + (1.0 - alpha) * bg
            combined[y_i, x_i] = blended.astype(np.uint8)

        self._draw_inner_margin_box(combined)
        combined_path = os.path.join(output_dir, f"2-bpp_overlay.png")
        cv2.imwrite(combined_path, combined)

        self.log_status(f"{short_tag}: full overlay saved: {out_overlay}")
        self.show_cv2_image(overlay, title=f"{title} (positive + interface)")

        self.log_status(f"{short_tag}: far-pink overlay saved: {combined_path}")
        self.show_cv2_image(combined, title=f"{title} (far pink on interface)")

        # ---------- quantification + Excel export ----------

        (
            fraction_far_pink_interface,
            n_bridge_regions,
            n_projection_regions,
            n_protrusion_regions,
            score_points,
            score_points_per_cell,
            bridge_mask,
            projection_mask,
            protrusion_mask,
        ) = self.compute_combined_far_pink_score(output_dir, projection_axis_ratio_cutoff=2, contact_radius=100)

        bridge_mask_u8 = (bridge_mask > 0).astype(np.uint8) * 255
        projection_mask_u8 = (projection_mask > 0).astype(np.uint8) * 255
        protrusion_mask_u8 = (protrusion_mask > 0).astype(np.uint8) * 255

        bridge_path = os.path.join(output_dir, f"5-bridge_mask.png")
        projection_path = os.path.join(output_dir, f"6-projection_mask.png")
        protrusion_path = os.path.join(output_dir, f"7-protrusion_mask.png")

        cv2.imwrite(bridge_path, bridge_mask_u8)
        cv2.imwrite(projection_path, projection_mask_u8)
        cv2.imwrite(protrusion_path, protrusion_mask_u8)

        self.log_status(f"{short_tag} bridge mask saved {bridge_path}")
        self.log_status(f"{short_tag} projection mask saved {projection_path}")
        self.log_status(f"{short_tag} protrusion mask saved {protrusion_path}")

        self.show_cv2_image(
            cv2.cvtColor(bridge_mask_u8, cv2.COLOR_GRAY2BGR),
            title=f"{short_tag} bridge mask",
        )
        self.show_cv2_image(
            cv2.cvtColor(projection_mask_u8, cv2.COLOR_GRAY2BGR),
            title=f"{short_tag} projection mask",
        )
        self.show_cv2_image(
            cv2.cvtColor(protrusion_mask_u8, cv2.COLOR_GRAY2BGR),
            title=f"{short_tag} protrusion mask",
        )

        excel_path = os.path.join(output_dir, "run_summary.xlsx")
        self.export_run_summary_excel(
            output_excel_path=excel_path,
            fraction_far_pink_interface=fraction_far_pink_interface,
            n_bridge_regions=n_bridge_regions,
            n_projection_regions=n_projection_regions,
            n_protrusion_regions=n_protrusion_regions,
            score_points=score_points,
            score_points_per_cell=score_points_per_cell,
        )

        summary_line = (
            f"{short_tag}: "
            f"fp/int={fraction_far_pink_interface:.4f}, "
            f"bridges={n_bridge_regions}, projections={n_projection_regions}, "
            f"protrusions={n_protrusion_regions}, score={score_points}, score_points_per_cell={score_points_per_cell}, "
            f"max_r={self.max_r}"
        )
        self.log_status(summary_line)
        print(summary_line)

        # timing at the very end
        elapsed = time.time() - start_time
        timing_msg = f"{short_tag}: threshold_and_contours completed in {elapsed:.2f} s"
        print(timing_msg)

if __name__ == "__main__":
    app = ImageTriangulatorApp()
    app.mainloop()
