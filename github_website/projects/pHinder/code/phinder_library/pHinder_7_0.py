global zero
zero = 1e-10

# Import dependencies here for enhanced performance.
####################################################
from pdbFile import PseudoAtom
from math import sqrt, exp, pi
from copy import deepcopy
#from tempfile import TemporaryFile
# from xlwt import Workbook # Replaced with openpyxl 2021.03
from os import mkdir, sep, listdir
from os.path import exists
from shutil import rmtree
from convexHull3D_1_0 import convexHull3D 
from convexHull4D_2_22 import convexHull4D, Simplex1
from pHinderSurface import calculateSurface, inLocalSurface, generateVirtualScreeningVertices
from goFo import *
from compGeometry import *
from sphere import Sphere
from minimizeNetworks import *
from writeFunctions import *
import multiprocessing as mp
import gzip
import platform
import pickle
import openpyxl

# Residue Sets (Gly is skipped).
#################################
modifiedResidues = ["MLZ", "MLY", "M3L"]
ionizableSet = ["ASP","GLU","HIS","CYS","LYS","ARG", "MLY", "MLZ", "M3L"]
ionizableSetNoCys = ["ASP","GLU","HIS","LYS","ARG", "MLY", "MLZ", "M3L"]
acidicSet = ["ASP","GLU"]
basicSet = ["HIS","LYS","ARG"]
polarSet = ["ASN","GLN","SER","THR","TYR","TRP"]
apolarSet = ["GLY", "ALA","ILE","LEU","MET","PHE","PRO","VAL"]
polarAndIonizableSet = ionizableSet + polarSet
allSet = ionizableSet + polarSet + apolarSet

# Heteroatom Sets (is K spacing correct?).
##########################################
ions = {" MG":2, " CA":2, " ZN":2, " NA":1, " CL":-1, " MN":2, "  K":1, " CD":2, " FE":3, " SR":2, " CU":2, " IOD":-1, " HG":2, " BR":-1, " CO":2, " NI":2, " FE2":2}


def s2StringOutput(v1,v2):

	format = "%15.6f%15.6f%15.6f "
	string = format % (v1.x,v1.y,v1.z) + \
			format % (v2.x,v2.y,v2.z) + "\n"
	return string

def s3StringOutput(v1,v2,v3):

	format = "%15.6f%15.6f%15.6f "
	string = format % (v1.x,v1.y,v1.z) + \
			format % (v2.x,v2.y,v2.z) + \
			format % (v3.x,v3.y,v3.z) + "\n"
	return string

def s4StringOutput(v1,v2,v3,v4):

	format = "%15.6f%15.6f%15.6f "

	# s31
	string = format % (v1.x,v1.y,v1.z) + \
			format % (v2.x,v2.y,v2.z) + \
			format % (v3.x,v3.y,v3.z) + "\n"
	# s32
	string += format % (v1.x,v1.y,v1.z) + \
			format % (v2.x,v2.y,v2.z) + \
			format % (v4.x,v4.y,v4.z) + "\n"
	# s33
	string += format % (v1.x,v1.y,v1.z) + \
			format % (v3.x,v3.y,v3.z) + \
			format % (v4.x,v4.y,v4.z) + "\n"
	# s34
	string += format % (v2.x,v2.y,v2.z) + \
			format % (v3.x,v3.y,v3.z) + \
			format % (v4.x,v4.y,v4.z) + "\n"

	return string

def proceed(orientedSurface,hull3D):

	edgeDict = makeEdgeDict(orientedSurface,hull3D)
	quadSurface = getQuadSurface(edgeDict)
	artEdges = []
	for e in quadSurface:
		rL = rotateCwOverTheLeftFace(quadSurface[e],quadSurface,check=1)
		rR = rotateCwOverTheRightFace(quadSurface[e],quadSurface,check=1)
		if not rL or not rR:
			print("Rotation Error", e)
			artEdges.append(e)
	
	if artEdges:
		return artEdges
	else:
		return 0

# Class for individual sidechain classifications
class sidechainClassification:

	def __init__(self, v):

		# Vertex information
		self.v = v

		# Sidechain type
		self.ionizable = 0
		self.polar = 0
		self.apolar = 0

		# Set sidechain type
		if self.v.data.residue.name in ionizableSet + ["HOH",]:
			self.ionizable = 1
		elif self.v.data.atom_name in ions:
			self.ionizable = 1
		# Not an Atom() instance; a PDB file line; therefore, the atom_name contains spaces.
		elif "OXT" in self.v.data.atom_name:
			self.ionizable = 1
		elif self.v.data.atom_name in polarSet:
			self.polar = 1
		elif self.v.data.atom_name in apolarSet:
			self.apolar = 1

		# Sidechain depth reported by TSC vertex depth 
		self.depth = 0.

		# Vertex location classification
		self.core = 0
		self.margin = 0
		self.exposed = 0

		# String for associating a label with the classified sidechain
		self.classificationString = ""

		# The other vertices in the microenvironment of self.v
		self.neighbors = [] 

		# Classify sidechains in domain cores versus those at protein-protein interfaces 
		self.domain = 0
		self.interfacial = 0

	def setClassificationString(self):

		self.classificationString = ""
		if not self.core and not self.margin and not self.exposed:
			locationKey = "N"
		elif self.core:
			locationKey = "C"
		elif self.margin:
			locationKey = "M"
		elif self.exposed:
			locationKey = "E"
		dep = self.depth
		num = self.v.data.residue.num
		res = self.v.data.residue.name
		chn = self.v.data.residue.chn
		self.classificationString = "%4s %5i%4s %s %5.1f" % (locationKey, num, res, chn, dep)


	# def __repr__(self):
		
	# 	if self.exposed:
	# 		return self.v.data.residue.name + "-" + \
	# 		    `self.v.data.residue.num` + \
	# 		    " Exposed residue at depth " + \
	# 		    "%4.2f" % (self.depth) + "\n"

	# 	elif self.margin:
	# 		return self.v.data.residue.name + "-" + \
	# 		    `self.v.data.residue.num` + \
	# 		    " Margin residue at depth " + \
	# 		    "%4.2f" % (self.depth) + "\n"
	# 	else:
	# 		return self.v.data.residue.name + "-" + \
	# 		    `self.v.data.residue.num` + \
	# 		    " Buried residue at depth " + \
	# 		    "%4.2f" % (self.depth) + "\n"

def orientS2s(triangulation):
	# @triangulation really is an s2Dict of a triangulation
	for node in triangulation:
		orientedS2s = []
		for s2 in triangulation[node].s2s:
			if node == s2[0]:
				orientedS2s.append(s2)
			else:
				orientedS2s.append((s2[1],s2[0]))
		# Update the triangulation.
		triangulation[node].s2s = orientedS2s


def printResNums(nodeList,t):
	nums = ""
	nodeList = list(nodeList)
	nodeList.sort()
	for node in nodeList:
		if t[node].s1.data.residue.name == "HOH": # HOH here
			nums += str(t[node].s1.data.residue.num) + "-" + t[node].s1.data.residue.chn + "w "
		else:
			nums += str(t[node].s1.data.residue.num) + "-" + t[node].s1.data.residue.chn + " "
			n = t[node].s1.data.residue.num

	tuple(nodeList)
	return nums

def findMinSidechainDistance(a, b):

	# Added in v_100.
	# The parameters a and b are atom distances.
	# This function finds the shortest distance between the atoms of two sidechains.
	# The calculation does not include the backbone atoms of the sidechain. 
	#################################################################################
	aAtoms = None
	if a.residue.name in ["HOH", "ARG", "LYS"]:
		aAtoms = [a,]
	else:
		aAtoms = a.residue.get_sidechain_atoms()
	bAtoms = None
	if b.residue.name in ["HOH", "ARG", "LYS"]:
		bAtoms = [b,]
	else:
		bAtoms = b.residue.get_sidechain_atoms()
	minDistance = 10000000.
	for aAtom in aAtoms:
		for bAtom in bAtoms:
			d = distance(aAtom, bAtom)
			# Make fuzzy sidechains due to reduced representations. This "softens" the network calculation.
			###############################################################################################
			if aAtom.residue.name in ["HOH", "ARG", "LYS", "CYS"]:
				d -= 0.5
			if bAtom.residue.name in ["HOH", "ARG", "LYS", "CYS"]:
				d -= 0.5
			if d < minDistance:
				minDistance = d

	return minDistance

# 2016.07.22 Updating, refining, simplifying.
#
# This version of parity should be entitled parityElectrostatics.
#################################################################
# v_105 Added
# v_106 Modified.
#################

def backboneHydrogenParity(networks, triangulation):

	# Write parity degree.
	######################
	minusResidues = ["ASP", "GLU", "CYS"]
	plusResidues  = ["HIS", "LYS", "ARG"]
	
	t = triangulation
	
	parityScores = {}
	for network in networks:
		for node1 in network:
			
			parityScore = 0
			
			r1 = t[node1].s1.data.residue
			resNum = r1.num
			resChain = r1.chn
			resName = r1.name
			
			if "H" in t[node1].s1.data.atom_name:
				continue

			nNeighbors = len(t[node1].s2s)
			for s2 in t[node1].s2s:
				for node2 in s2:
					if node1 != node2:
						r2 = t[node2].s1.data.residue
						d = distance(r1.get_terminal_sidechain_atom(), r2.get_terminal_sidechain_atom())
						
						# Set specific selection rules.
						###############################
						if t[node2].s1.data.atom_name != " H":
							continue
						if d > 6.0:
							continue
						if t[node2].s1.data.residue_sequence_number == resNum:
							continue
						
						increment = 1

						parityScore += increment

			if r1.name in minusResidues:
				parityScore = -1*parityScore
	
			parityScores.update({(r1.key):parityScore})

	return parityScores

def sidechainHydrogenParity(networks, triangulation):
	
	# Write parity degree.
	######################
	minusResidues = ["ASP", "GLU", "CYS"]
	plusResidues  = ["HIS", "LYS", "ARG"]
	
	t = triangulation
	
	parityScores = {}
	for network in networks:
		for node1 in network:
			
			parityScore = 0
			
			r1 = t[node1].s1.data.residue
			resNum = r1.num
			resChain = r1.chn
			resName = r1.name
			
			if "H" in t[node1].s1.data.atom_name:
				continue
		
			nNeighbors = len(t[node1].s2s)
			dontDoubleCountSidechains = {}
			for s2 in t[node1].s2s:
				for node2 in s2:
					if node1 != node2:
						r2 = t[node2].s1.data.residue
						d = distance(r1.get_terminal_sidechain_atom(), r2.get_terminal_sidechain_atom())
						
						# Set specific selection rules.
						###############################
						if "H" not in t[node2].s1.data.atom_name:
							continue
						if t[node2].s1.data.atom_name == " H":
							continue
						if d > 6.0:
							continue
						if t[node2].s1.data.residue_sequence_number == resNum:
							continue

						increment = 1
						
						if t[node2].s1.data.residue_sequence_number not in dontDoubleCountSidechains:
							parityScore += increment
							dontDoubleCountSidechains.update({t[node2].s1.data.residue_sequence_number:None})
			
			if r1.name in minusResidues:
				parityScore = -1*parityScore
					
			parityScores.update({(r1.key):parityScore})
	
	return parityScores

def ionParity(networks, triangulation):
	
	# Ion parity residues.
	######################
	minusResidues = ["ASP", "GLU", "CYS"]
	plusResidues  = ["HIS", "LYS", "ARG"]
	skipResidues = ["PSA"]
	t, parityScores = triangulation, {}
	for network in networks:
		for node1 in network:
			r1Key = t[node1].s1.data.residue.key
			r1 = t[node1].s1.data.residue.name
			# In cases where there are less than 5 vertices to initiate a 4D convex hull calculation,
			# PSAs are added to make up the difference, and should be removed from this caluclation.
			if r1 in skipResidues:
				continue
			# Set charge of r1
			c1 = 0
			if r1 in minusResidues:
				c1 = -1
			else:
				c1 = 1
			parityScore = 0
			for s2 in t[node1].s2s:
				for node2 in s2:
					if node1 != node2:
						r2 = t[node2].s1.data.residue.name
						# In cases where there are less than 5 vertices to initiate a 4D convex hull calculation,
						# PSAs are added to make up the difference, and should be removed from this caluclation.
						if r2 in skipResidues:
							continue
						# Set charge of r2 
						c2 = 0
						if r2 in minusResidues:
							c2 = -1
						else:
							c2 = 1
						# Summed "Coulomb parity"
						# Negative parity score is favorable.
						# Positive parity score is unfavorable.
						parityScore += c1*c2
			nodeDegree = len(t[node1].s2s)
			parityScores.update({(parityScore, r1Key):(parityScore, nodeDegree)})

	return parityScores

def resParity(networks, triangulation):
	
	# Residue groups based on core probablity: Isom et al., Biochemistry 2016.
	##########################################################################
	resGroup1 = ["ILE", "PHE", "LEU", "VAL", "CYS", "TRP", "MET", "ALA"]
	resGroup2 = ["TYR", "THR", "SER", "HIS", "PRO"]
	resGroup3 = ["ASN", "GLN", "ASP", "ARG", "GLU", "LYS"]
	
	t = triangulation
	
	parityScores = {}
	for network in networks:
		for node1 in network:
			
			r1Key = t[node1].s1.data.residue.key
			r1 = t[node1].s1.data.residue.name
			parityScore = 0
			for s2 in t[node1].s2s:
				for node2 in s2:
					if node1 != node2:
						r2 = t[node2].s1.data.residue.name
						c2 = 0
						if r1 in resGroup1 and r2 in resGroup1:
							c2 = 1
						elif r1 in resGroup2 and r2 in resGroup2:
							c2 = 1
						elif r1 in resGroup3 and r2 in resGroup3:
							c2 = 1
						else:
							c2 = -1
						parityScore += c2
			nodeDegree = len(t[node1].s2s)
			parityScores.update({(parityScore, r1Key):(parityScore, nodeDegree)})

	return parityScores

def classifySidechains(tscVertices, tscVertexDict, allAtomVertexDict, quadSurface, displayDirectory, pdbCode, coreCutoff, marginCutoff):

	# Using the global protein surface to classify all sidechains as being core, marginal, or exposed.
	##################################################################################################
	vNonEvaluated, vIn, vMarg, vOut = {}, {}, {}, {}
	classifications = inLocalSurface(tscVertices, quadSurface, coreCutoff, marginCutoff)
	
	for classification in classifications:
		
		v, status, depth = classification[0], classification[1], classification[2]
		if status == -1:
			v.data.core = 1
			vIn.update({v:[depth, depth, None]})
		elif status == 0:
			v.data.margin = 1
			vMarg.update({v:[depth, depth, None]})
		elif status == 1:
			v.data.exposed = 1
			vOut.update({v:[depth, depth, None]})
	
	# Convert all sidechain vertices to sidechainClassification objects.
	vCore, vMargin, vExposed = {}, {}, {}

	# Convert vOut to vExposed classified using the pHinder surface.
	for v in vOut:
		#v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name)] # 2016.01.29
		v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name, v.data.atom_name)]
		classification = sidechainClassification(v4D)
		classification.depth = vOut[v][1]
		classification.exposed = 1
		classification.setClassificationString()
		vExposed.update({classification.v:classification})
	
	# Convert vMarg to vMargin classified using the pHinder surface.
	for v in vMarg:
		#v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name)] # 2016.01.29
		v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name, v.data.atom_name)]
		classification = sidechainClassification(v4D)
		classification.depth = vMarg[v][1]
		classification.margin = 1
		classification.setClassificationString()
		vMargin.update({classification.v:classification})

	# Covert v to vCore classified using the pHinder surface.
	for v in vIn:
		#v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name)] # 2016.01.29
		v4D = tscVertexDict[(v.data.residue.num, v.data.residue.chn, v.data.residue.name, v.data.atom_name)]
		classification = sidechainClassification(v4D)
		classification.depth = vIn[v][1]
		classification.core = 1
		classification.setClassificationString()
		vCore.update({classification.v:classification})

	return (vExposed, vMargin, vCore)

def inHull(vertices, hull3D):

	# This function can't be part of the pHinder class because instance methods cannot be pickled, 
	# which is a necessary operation for functions used in parallel calculations using the mp module.

	# Identify the grid vertices inside of the protein's convex hull.
	#################################################################
	vKeep = []
	for v in vertices:
		if hull3D.test_if_this_vertex_is_in_the_hull(v):
			vKeep.append(v)
	return vKeep

def removeClashes(gridVertices, atoms, clashCutoff):

	# This function can't be part of the pHinder class because instance methods cannot be pickled, 
	# which is a necessary operation for functions used in parallel calculations using the mp module.

	# Identify the grid vertices that do not clash with any protein atoms.
	######################################################################
	vNoClash = []
	for v in gridVertices:
		clash = 0
		for a in atoms:
			d = distance(v.data, a)
			if d < clashCutoff:
				clash = 1
				break
		if not clash:
			vNoClash.append(v)

	return vNoClash

# def calculateVoidSurfaceSamplingPoints(surfacePointSet, radius, cut):

# 	# Using a sphere, generate surface sampling points for the void network.
# 	########################################################################
# 	lastGpAtomSerial = 0
# 	lastGp = surfacePointSet[-1]
# 	lastGpAtomSerial = lastGp.atom_serial + 1
# 	nSurfacePoints = len(surfacePointSet)
# 	surfaceVertexSet, surfaceVertexDict = [], {}

# 	for gp in surfacePointSet:
		
# 		# Set the u-coordinate, vertex, and atom identifiers.
# 		#####################################################
# 		x, y, z = gp.x, gp.y, gp.z
# 		u = x**2 + y**2 + z**2
# 		vertex = Vertex4D((x,y,z,u),setC=1)
# 		vertex.id = gp.atom_serial
# 		gp.v = vertex # Overwrite the original Vertex() object with new Vertex4D() object.
# 		vertex.data = gp
# 		surfaceVertexDict.update({(gp.residue_sequence_number, gp.chain_identifier, gp.atom_name): vertex})
# 		surfaceVertexSet.append(vertex)
		
# 		# The surface point set will be expanded.
# 		#########################################
# 		if lastGpAtomSerial:
# 			s = Sphere()
# 			s.radius = 1.0
# 			s.origin_x = x
# 			s.origin_y = y
# 			s.origin_z = z
# 			s.generate_surface_points(pi/cut, pi/cut, general_position=1)
# 			psas = s.get_pseudoatoms()
# 			for psa in psas:
# 				x, y, z, u = psa.x, psa.y, psa.z, (psa.x**2 + psa.y**2 + psa.z**2)
# 				pv = Vertex4D((x, y, z, u), setC=1)
# 				psa.atom_serial = lastGpAtomSerial
# 				psa.residue_sequence_number = lastGpAtomSerial
# 				psa.chain_identifier = gp.chain_identifier
# 				pv.id = psa.atom_serial
# 				psa.v = pv # Overwrite the original Vertex() object with new Vertex4D() object.
# 				pv.data = psa
# 				psa.reinitialize()
				
# 				surfaceVertexDict.update({(psa.residue_sequence_number, psa.chain_identifier, psa.atom_name): pv})
# 				surfaceVertexSet.append(pv)
# 				lastGpAtomSerial += 1

# 	# Return the set of 4D surface vertices.
# 	######################################
# 	return surfaceVertexSet

def calculateVoidSurfaces(clashFreeSurfaceVertexSet):

	surfaceVertices = clashFreeSurfaceVertexSet[0]
	surfaceVertexDict = clashFreeSurfaceVertexSet[1]

	print("Calculating surface of void volume...", len(surfaceVertices), "...vertices.")

	voidSurface = calculateSurface(surfaceVertices, 3.0, longEdgeCutLimit=2.0, allowSmallSurfaces=self.allowSmallSurfaces)

	# Return the quad surface.
	##########################
	if voidSurface:
		return voidSurface[1]
	else:
		return 0

# def calculateVoidSurfaceSamplingShells(rankedVoidSurface, allAtomVertexDict, outputDirectory, pdbCode):

# 	i = rankedVoidSurface[0]
# 	voidSurface = rankedVoidSurface[1]

# 	print("in calculateVoidSurfaceSamplingShells()...", len(voidSurface))

# 	shellEdgeStringColors, shellResidues, voidSurfaceEdges = "", {}, ""
# 	for e in voidSurface:

# 		# The edge object.
# 		##################
# 		e = voidSurface[e]

# 		# Set the midPoint attribute of the Edge object.
# 		################################################
# 		eMidpoint = e.midPoint

# 		# Identify the closest atom to the Edge object.
# 		###############################################
# 		minD, surfaceAtom, closestAtom = 10000000., eMidpoint, None
# 		for v in allAtomVertexDict.values():

# 			d = distance(eMidpoint, v)
# 			if d < minD:
# 				minD, closestAtom = d, v.data

# 		# Set the closestAtom attribute of the Edge object.
# 		###################################################
# 		e.closestAtom = closestAtom

# 		if minD < 6.0: # Should I provide as an paramater argument

# 			# The atom vertex must be on the outside the surface.
# 			#####################################################
# 			if 1 not in [e.lf.beneath_beyond(closestAtom.v), e.rf.beneath_beyond(closestAtom.v)]:

# 				a1, a2 = surfaceAtom, closestAtom
# 				shellResidues.update({a2.residue.key:a2.residue})


# 				if 'C' in closestAtom.atom_name:

# 					shellEdgeStringColors += 'C:' + s2StringOutput(a1, a2)
# 					voidSurfaceEdges += 'C:' + s2StringOutput(e.origin_vertex.data, e.destination_vertex.data)

# 				elif 'S' in closestAtom.atom_name:

# 					shellEdgeStringColors += 'S:' + s2StringOutput(a1, a2)
# 					voidSurfaceEdges += 'S:' + s2StringOutput(e.origin_vertex.data, e.destination_vertex.data)

# 				elif 'N' in closestAtom.atom_name:

# 					shellEdgeStringColors += 'N:' + s2StringOutput(a1, a2)
# 					voidSurfaceEdges += 'N:' + s2StringOutput(e.origin_vertex.data, e.destination_vertex.data)

# 				elif 'O' in closestAtom.atom_name:

# 					shellEdgeStringColors += 'O:' + s2StringOutput(a1, a2)
# 					voidSurfaceEdges += 'O:' + s2StringOutput(e.origin_vertex.data, e.destination_vertex.data)

# 			else:
# 				voidSurfaceEdges += 'X:' + s2StringOutput(e.origin_vertex, e.destination_vertex)

# 		else:
# 			voidSurfaceEdges += 'X:' + s2StringOutput(e.origin_vertex, e.destination_vertex)

# 	# Write shell to file.
# 	######################
# 	if shellResidues:

# 		outFile = outputDirectory + pdbCode + ".shellResidues." + str(i) + ".pdb.gz"
# 		outputSurfEdgesResidues = gzip.open(outFile,"wt")
# 		for shellResidue in shellResidues:
# 			outputSurfEdgesResidues.write(str(shellResidues[shellResidue]))
# 		outputSurfEdgesResidues.close()

# 		outFile = outputDirectory + pdbCode + ".shellContacts." + str(i) + ".txt.gz"
# 		outputSurfEdges = gzip.open(outFile,"wt")
# 		outputSurfEdges.write(shellEdgeStringColors)
# 		outputSurfEdges.close()

# 		outFile = outputDirectory + pdbCode + ".shellEdges." + str(i) + ".txt.gz"
# 		outputSurfEdges = gzip.open(outFile,"wt")
# 		outputSurfEdges.write(voidSurfaceEdges)
# 		outputSurfEdges.close()

# 		# Pickle the virtual void surface for future access.
# 		####################################################
# 		pickledOutput = open(outputDirectory + pdbCode + ".pickled.shellEdges." + str(i) + ".pkl", 'wb')
# 		pickle.dump(voidSurface, pickledOutput)
# 		pickledOutput.close()


def create_timestamped_directory(base_dir, dir_name):

	import os
	import datetime

	"""
	Create a directory with a human-readable timestamp added to the name.
	
	Parameters:
	- base_dir (str): The base directory where the new folder will be created.
	- dir_name (str): The name of the directory to create.
	
	Returns:
	- str: The full path of the created directory.
	"""
	# Generate a human-readable timestamp in the format YYYY-MM-DD at HH.MM.SS
	timestamp = datetime.datetime.now().strftime("%Y-%m-%d at %I.%M.%S %p")
	# Combine the directory name with the timestamp
	timestamped_dir_name = f"{dir_name}_{timestamp}"
	# Full path of the new directory
	full_path = os.path.join(base_dir, timestamped_dir_name)
	
	# Create the directory
	os.makedirs(full_path, exist_ok=True)
	return full_path
class pHinder:

	def __init__(self):

		# Indicate if input is originating from the GUI instead of command line
		self.gui = False

		# CPU count.
		############
		self.cpuCount = mp.cpu_count()

		# Default pHinder argument variables.
		#####################################
		self.pdbFormat = "pdb" # "pdb" or "mmCIF"
		self.pdbFilePath = ""
		self.pdbFileName = "" 
		# self.pdbFilePathList = [] 
		self.chains = []
		self.group_chains = 0
		self.chainString = ""
		self.maxNetworkEdgeLength = 10.0
		self.minNetworkSize=1
		self.includeWater = 0
		self.includeIons = 0
		self.zip = 0
		self.reducedNetworkRepresentation = 1
		self.highResolutionSurface = 1
		self.saveSurface = 1
		self.allowSmallSurfaces = 0
		self.saveLigandSurfaces = 1
		self.saveNetworkTriangulation = 1
		self.residueSet = "allSet"
		self.allowCysCoreSeeding = 0
		self.writeSurfaceCreationAnimation = 0
		self.coreCutoff = -3.0
		self.marginCutoff = 1.05
		self.marginCutoffCoreNetwork = -2.0
		self.includeHydrogens = 0
		self.processes = 1
		self.gridIncrement=2.0
		self.maxVoidNetworkEdgeLength=2.0
		self.minVoidNetworkSize=1

		# pHinder name and path variables.
		##################################
		self.pdbCode = ""
		self.queryResidueSet = ""
		self.protein = None
		self.outPath = ""
		self.outPaths_bychain = {}
		#self.outputDirectory = ""#self.outPath

		# Residue collections.
		######################
		self.residues = None
		# self.nucleicAcids = None
		self.het_residues = None
		# self.waters = None

		# Atom collections.
		###################
		self.hydrogenAtoms = []
		self.allAtoms = []
		self.bbAtoms = [] 
		self.caAtoms = [] 
		self.tscAtoms = [] 
		self.residuesMissingTscAtom = []

		# 4D vertex sets.
		#################
		self.hetVertices = []
		self.hetVertexDict = {}
		self.waterVertices = []
		self.waterVertexDict = {}
		self.caVertices = [] 
		self.caVertexDict = {}
		self.cbVertices = [] 
		self.cbVertexDict = {}
		self.tscVertices = []
		self.tscVertexDict = {}
		self.terminiVertices = [] 
		self.terminiVertexDict = {}

		# Tsc set selection variables.
		##############################
		self.tscVerticesSelection = []
		self.tscVertexDictSelection = {}

		# Tsc triangulation variables.
		##############################
		self.tscTriangulation = None
		self.originalTscTriangulation = None
		self.mutatedTscTriangulation = None 
		self.triangulations_bychain = {}
		self.networks = None

		# Tsc triangulation variables that aide side chain classification.
		###################################################################
		self.tscTriangulationClassification = None
		self.triangulations_bychain_classification = {}

		# Open a global workbook for writing network report information.
		# At the appropriate point in the code, a sheet for each network type is added to the global workbook.
		######################################################################################################
		#self.networkBook = Workbook()
		self.networkBook = openpyxl.Workbook()
		sheet = self.networkBook["Sheet"]
		self.networkBook.remove(sheet)

		# Network analysis parameters and variables.
		############################################
		self.disulfideBonds = {}
		self.disulfideBonds_bychain = {}
		self.ionPairCutoff = 4.0

		# Networks by grouped chains, if calculation is done with grouped PDB chains...
		###############################################################################
		self.topology_networks = None
		self.ionpair_networks = None
		self.disulfide_networks = None
		self.parity_networks = None
		self.parity_negative_networks = None
		self.parity_positive_networks = None
		self.exposed_networks = None
		self.margin_networks = None
		self.core_networks = None

		# Networks by chain, if calculation is done in individual PDB chains...
		#######################################################################
		self.topology_networks_bychain = {}
		self.ionpair_networks_bychain = {}
		self.disulfide_networks_bychain = {}
		self.parity_networks_bychain = {}
		self.parity_negative_networks_bychain = {}
		self.parity_positive_networks_bychain = {}
		self.exposed_networks_bychain = {}
		self.margin_networks_bychain = {}
		self.core_networks_bychain = {}

		# Surface variables for grouped chains, if calculation is done with grouped PDB chains...
		#########################################################################################
		self.hull3D = {}
		self.hull4D = {}
		self.quadSurface = {}
		self.ligandSurfaces = {}

		# Surface variables for chains, if calculation is done in individual PDB chains...
		##################################################################################
		self.hull3D_bychain = {}
		self.hull4D_bychain = {}
		self.quadSurface_bychain = {}
		self.ligandSurfaces_bychain = {}

		# Sidechain classification variables for grouped chains, if calculation is done with grouped PDB chains...
		##########################################################################################################
		self.vExposed = {}
		self.vMargin = {}
		self.vCore = {}

		# Sidechain classification variables for chains, if calculation is done in individual PDB chains...
		###################################################################################################
		self.vExposed_bychain = {}
		self.vMargin_bychain = {}
		self.vCore_bychain = {}

		# Virtual screening variables
		#############################
		self.virtualClashCutoff=3.0
		self.inIterations=1
		self.inIterationsStepSize=1
		self.outIterations=1
		self.outIterationsStepSize=1
		self.gridVertices = []
		self.gridVertices_bychain = {}
		self.gridAtoms = []
		self.gridAtoms_bychain = {}
		self.gridVerticesNoClash = {}
		self.gridVerticesNoClash_bychain = {}
		self.gridPointMesh = {}
		self.gridPointMesh_bychain = {}
		self.gridPointTriangulation = None
		self.gridPointTriangulation_bychain = {}
		self.voidSurfaceAtomSets = []
		self.voidSurfaceVertexSets = []
		self.prunedVoidNetworks = {}
		self.prunedVoidNetworks_bychain = {}
		self.prunedVoidNetworkTriangulation = {}
		self.prunedVoidNetworkTriangulation_bychain = {}
		self.voidNetworks = {}
		self.voidNetworks_bychain = {}
		self.rankedVoidNetworks = {}
		self.rankedVoidNetworks_bychain = {}

	def setQuerySet(self):

		# Define the set of query atoms
		if "customSet" in self.residueSet:
			resNameCsvList = self.residueSet.split(":")[1]
			aa_selection = []
			for resName in resNameCsvList.split(","):
				aa_selection.append(resName)
			self.queryResidueSet = aa_selection
			self.residueSet = "customSet"
		elif self.residueSet == "ionizableSet":
			self.queryResidueSet = ionizableSet
		elif self.residueSet == "ionizableSetNoCys":
			self.queryResidueSet = ionizableSetNoCys
		elif self.residueSet == "acidicSet":
			self.queryResidueSet = acidicSet
		elif self.residueSet == "basicSet":
			self.queryResidueSet = basicSet
		elif self.residueSet == "polarSet":
			self.queryResidueSet = polarSet
		elif self.residueSet == "apolarSet":
			self.queryResidueSet = apolarSet
		elif self.residueSet == "allSet":
			self.queryResidueSet = allSet
		else:
			self.queryResidueSet = allSet

		print("Done defining query atom set...")

	def openPDBs(self, pdbFilePath, pdbFileName, zip_status=0, make_output_directory=1):

		# File location information
		self.pdbFilePath = pdbFilePath
		self.pdbFileName = pdbFileName
		self.zip = zip_status

		# Make a directory for output. Delete pre-existing ouput
		if make_output_directory:
			if self.outPath: 
				if not exists(self.outPath):
					mkdir(self.outPath)

		# # Get the PDB code.
		# pdbFileName = self.pdbFilePathList[0].split("/")[-1]
		# tempPdbCode = ""
		# for w1 in pdbFileName.split("."):
		# 	if w1 not in ["pdb","gz"]:
		# 		tempPdbCode += w1 + "."
		# tempPdbCode = tempPdbCode[0:-1]
		# self.pdbCode = tempPdbCode

		# # Create an instance of a PDBfile (really protein) object.
		# self.protein = PDBfile(self.pdbFilePathList,zip=self.zip)

		# Get the PDB code...
		tempPdbCode = ""
		for w1 in self.pdbFileName.split("."):
			if w1 not in ["pdb", "cif", "gz"]:
				tempPdbCode += w1 + "."
		tempPdbCode = tempPdbCode[0:-1]
		self.pdbCode = tempPdbCode

		# Create an instance of a PDBfile (really protein) object
		if self.pdbFormat == "mmCIF":
			import pdbFile_cif
			self.protein = pdbFile_cif.PDBfile(self.pdbFilePath, self.pdbFileName, zip_status=self.zip)
		else:
			import pdbFile
			self.protein = pdbFile.PDBfile(self.pdbFilePath, self.pdbFileName, zip_status=self.zip)	

		# Make PDB code directory...
		self.outPath = create_timestamped_directory(self.outPath, self.pdbCode) + sep
		print(self.outPath)
		# outPath = self.outPath + self.pdbCode + sep
		# if make_output_directory:
		# 	if not exists(outPath):
		# 		mkdir(outPath)

		# Automatically set to properties of group_chains in there is only one chain in the PDB file...
		# if len(self.protein.chains) == 1:
		# 	self.group_chains = 1

		# One or more specific PDB chains was set by the user prior to this function call...
		if self.gui and self.group_chains:
			self.chains = self.chains[:-1]

		if self.chains:

			self.chainString = ""
			for chain in self.chains:
				self.chainString += chain + "."
			self.chainString = self.chainString[:-1]

			# Make PDB chain directories for output...
			if self.group_chains:
				if self.outPath:
					# self.outPath = self.outPath + self.pdbCode + sep + self.chainString + sep 
					# self.outPath = outPath + sep + self.chainString + sep 
					self.outPath += self.chainString + sep 
					if not exists(self.outPath):
						mkdir(self.outPath)

			else:
				self.outPaths_bychain = {}
				for chain in self.chains:
					if chain in self.protein.chains:
						# outPath = self.outPath + self.pdbCode + sep + chain + sep
						# outPath_chain = outPath + sep + chain + sep
						outPath_chain = self.outPath + chain + sep
						self.outPaths_bychain[chain] = outPath_chain 
						if not exists(outPath_chain):
							mkdir(outPath_chain)	
					else:
						print("Chain", chain, "does not exist in the PDB file.")
			
		# The PDB file name or self.group_chains parameter will determine the chains used...
		else:

			# Process PDB chain choices (or lack thereof)...
			if "-chns" in pdbFileName:
				self.chainString = pdbFileName.split("-chns.")[1].split(".pdb")[0]
				self.chains = []
				for chain in self.chainString.split("."):
					if chain:
						self.chains.append(chain)
			else:
				self.chains = sorted(self.protein.chains)
				self.chainString = ""
				for chain in self.chains:
					self.chainString += chain + "."
				self.chainString = self.chainString[:-1]

			# Make PDB chain directories for output...
			if self.group_chains:
				if self.outPath:
					# self.outPath = self.outPath + self.pdbCode + sep + self.chainString + sep 
					# self.outPath = outPath + sep + self.chainString + sep
					self.outPath += self.chainString + sep
					if not exists(self.outPath):
						mkdir(self.outPath)
			else:
				self.outPaths_bychain = {}
				for chain in self.chains:
					# outPath = self.outPath + self.pdbCode + sep + chain + sep 
					# outPath_chain = outPath + sep + chain + sep
					outPath_chain = self.outPath + chain + sep
					self.outPaths_bychain[chain] = outPath_chain
					if not exists(outPath_chain):
						mkdir(outPath_chain)		

		# self.chains = self.protein.chains
		self.residues = self.protein.residues	
		self.het_residues = self.protein.het_residues

		print("Done opening PDB file(s)...")

	# def makeResidueCollections(self):

	# 	# Collect the desired protein chains and their associated ligands and water molecules.
	# 	self.residues, self.nucleicAcids, self.het_residues, self.waters = {}, {}, {}, {}
	# 	for chain in self.chains:

	# 		# Get the residues of the protein chain.
	# 		try:
	# 			self.residues.update(self.protein.res_chains[chain].residues)
	# 		except:
	# 			pass

	# 		# Get the ligands of the protein chain.
	# 		try:
	# 			# Get the hetero-residue regardless of its composition.
	# 			self.het_residues.update(self.protein.het_bychain[chain])

	# 			# If the hetero-residue belongs to DNA or RNA, save it by chain in nucleicAcids.
	# 			nucleicAcidResidue = 0
	# 			for hetResidue in self.protein.het_bychain[chain]:
	# 				if self.protein.het_bychain[chain][hetResidue].name in ["G", "C", "A", "T", "U"]:
	# 					nucleicAcidResidue = 1
	# 					break
	# 			if nucleicAcidResidue:
	# 				if chain not in self.nucleicAcids:
	# 					self.nucleicAcids.update({chain:{}})
	# 				self.nucleicAcids.update(self.protein.het_bychain[chain])
	# 		except:
	# 			pass

	# 		# Get water molecules associated with the protein chain.
	# 		try:
	# 			self.waters.update(self.protein.wat_bychain[chain])
	# 		except:
	# 			pass

	# 	print("Done gathering residue collections...")

	def hetLigand4D(self):

		# Convert the ligand residues to 4-dimensional vertices compatible with triangulation.
		self.hetVertices, self.hetVertexDict = [], {}
		for hetResidue in self.het_residues:
			for atom in self.het_residues[hetResidue].atoms.values():
				x, y, z= atom.x, atom.y, atom.z
				u = atom.x**2 + atom.y**2 + atom.z**2
				vertex = Vertex4D((x,y,z,u),setC=1)
				# Have to explicitly assign the residue instance.
				atom.residue = self.het_residues[hetResidue]
				vertex.data = atom
				vertex.id = atom.atom_serial
				self.hetVertexDict.update({(atom.residue.num, atom.residue.chn, atom.residue.name, atom.atom_name): vertex})
				self.hetVertices.append(vertex)

			# If the PDB file contains modified residues, add them to the residues dictionary.
			if hetResidue[1] in modifiedResidues:
				self.residues.update({hetResidue:self.het_residues[hetResidue]})  

			print("Done creating 4D ligand atoms....")

	# def hetWater4D(self):

	# 	# Convert the water molecules to 4-dimensional vertices compatible with triangulation.
	# 	self.waterVertices, self.waterVertexDict = [], {}
	# 	for water in self.waters:
	# 		for atom in self.waters[water].atoms.values():
	# 			x, y, z= atom.x, atom.y, atom.z
	# 			u = atom.x**2 + atom.y**2 + atom.z**2
	# 			vertex = Vertex4D((x,y,z,u),setC=1)
	# 			# Have to explicitly assign the residue instance.
	# 			atom.residue = self.waters[water]
	# 			vertex.data = atom
	# 			vertex.id = atom.atom_serial
	# 			self.waterVertexDict.update({(atom.residue.num, atom.residue.chn, atom.residue.name, atom.atom_name):vertex})
	# 			self.waterVertices.append(vertex)

	# 	print("Done creating 4D water atoms....")

	def hydrogens(self):

		# Collect all hydrogen atoms.
		if self.includeHydrogens:
			residue_keys = sorted(self.residues)
			self.hydrogenAtoms = []
			for key in residue_keys:
				residueAtoms = self.residues[key].atoms.values()
				for residueAtom in residueAtoms:
					if "H" in residueAtom.atom_name:
						self.hydrogenAtoms.append(residueAtom)

		print("Done gathering hydrogen atoms....")

	def makeAtomCollections(self):

		residue_keys = sorted(self.residues)
		self.allAtoms, self.bbAtoms, self.caAtoms, self.cbAtoms, self.tscAtoms, self.residuesMissingTscAtoms = [], [], [], [], [], []
		for key in residue_keys:
			self.allAtoms += self.residues[key].atoms.values()
			self.bbAtoms += self.residues[key].get_backbone_atoms()
			# Residues missing TSC atoms are not included in the tscAtom set.
			if not self.residues[key].get_terminal_sidechain_atom():  
				# Log that the residue is missing side chain atoms.
				self.residues[key].missingSidechainAtoms = 1
				self.residuesMissingTscAtoms.append(self.residues[key])
				try:
					# Collect the C alpha atom of the side chain.
					self.caAtoms.append(self.residues[key].atoms["CA"])
					try:
						self.cbAtoms.append(self.residues[key].atoms["CB"])
					except:
						self.cbAtoms.append(self.residues[key].atoms["CA"])
				except:
					# No C alpha atom.
					pass
			else:
				# Use the TSC atom to represent the side chain.
				self.tscAtoms.append(self.residues[key].get_terminal_sidechain_atom())
				try:
					# Collect the C alpha atom of the side chain.
					self.caAtoms.append(self.residues[key].atoms["CA"])
					try:
						self.cbAtoms.append(self.residues[key].atoms["CB"])
					except:
						self.cbAtoms.append(self.residues[key].atoms["CA"])
				except:
					# No C alpha atom.
					pass	

		# If caAtom or tscAtoms is empty because of oddities on the given PDB file, terminate the program.
		if not self.caAtoms or not self.tscAtoms:
			print("Cannot load C-alpha and/or terminal sidechain atoms.")
			print("No result.")
			return 0

		print("Done gathering atoms collections....")

	def calculateConvexHull3D(self):

		print("Calculating convex hull of protein in 3D start...")
		self.hull3D = convexHull3D(self.caVertices)
		print("Calculating convex hull of protein in 3D end...")

	def makeVertices4D(self):

		# Convert atoms to 4-dimensional vertices compatible with triangulation.
		########################################################################

		# Covert all C-alpha atoms of the protein to Vertex4D objects for triangulating.
		################################################################################
		self.caVertices, self.caVertexDict = [], {}
		for caAtom in self.caAtoms:
			x, y, z = caAtom.x, caAtom.y, caAtom.z
			u = caAtom.x**2 + caAtom.y**2 + caAtom.z**2
			vertex = Vertex4D((x,y,z,u),setC=1)
			vertex.data = caAtom
			vertex.id = caAtom.atom_serial
			self.caVertexDict.update({(caAtom.residue.num, caAtom.residue.chn, caAtom.residue.name, caAtom.atom_name):vertex})
			self.caVertices.append(vertex)

		# Covert all C-beta atoms of the protein to Vertex4D objects for triangulating.
		################################################################################
		self.cbVertices, self.cbVertexDict = [], {}
		for cbAtom in self.cbAtoms:
			x, y, z = cbAtom.x, cbAtom.y, cbAtom.z
			u = cbAtom.x**2 + cbAtom.y**2 + cbAtom.z**2
			vertex = Vertex4D((x,y,z,u),setC=1)
			vertex.data = cbAtom
			vertex.id = cbAtom.atom_serial
			self.cbVertexDict.update({(cbAtom.residue.num, cbAtom.residue.chn, cbAtom.residue.name, cbAtom.atom_name):vertex})
			self.cbVertices.append(vertex)

		# Covert all terminal side chain atoms, TSC atoms, to Vertex4D objects for triangulating.
		#########################################################################################
		self.tscVertices, self.tscVertexDict, i  = [], {}, 1
		for tscAtom in self.tscAtoms:
			x, y, z= tscAtom.x, tscAtom.y, tscAtom.z
			u = tscAtom.x**2 + tscAtom.y**2 + tscAtom.z**2
			vertex = Vertex4D((x,y,z,u),setC=1)
			vertex.data = tscAtom
			vertex.id = tscAtom.atom_serial 
			# Residue selection is made here.
			################################
			if tscAtom.residue_name != "GLY":
				self.tscVertexDict.update({(tscAtom.residue.num, tscAtom.residue.chn, tscAtom.residue.name, tscAtom.atom_name):vertex})
				self.tscVertices.append(vertex)
			i += 1
		self.tscVerticesSelection, self.tscVertexDictSelection = self.tscVertices, self.tscVertexDict

		# Convert all protein atoms to Vertex4D objects for triangulating.
		##################################################################
		self.allAtomVertices, self.allAtomVertexDict = [], {}
		for allAtom in self.allAtoms:
			x, y, z = allAtom.x, allAtom.y, allAtom.z
			u = allAtom.x**2 + allAtom.y**2 + allAtom.z**2
			vertex = Vertex4D((x,y,z,u),setC=1)
			vertex.data = allAtom
			vertex.id = allAtom.atom_serial
			self.allAtomVertexDict.update({(allAtom.residue.num, allAtom.residue.chn, allAtom.residue.name, allAtom.atom_name):vertex})
			self.allAtomVertices.append(vertex)

		# Identify the terminal carboxyl atom, OXT, and convert it to a Vertex4D object for triangulating.
		##################################################################################################
		self.terminiVertices, self.terminiVertexDict = [], {}
		if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:
			for caAtom in self.caAtoms:	
				for atom in caAtom.v.data.residue.atoms:
					if atom == "OXT":
						oxtAtom = caAtom.v.data.residue.atoms[atom]
						x, y, z= oxtAtom.x, oxtAtom.y, oxtAtom.z
						u = oxtAtom.x**2 + oxtAtom.y**2 + oxtAtom.z**2
						vertex = Vertex4D((x,y,z,u),setC=1)
						vertex.data = oxtAtom
						vertex.id = oxtAtom.atom_serial
						oxtAtom.v = vertex
						self.terminiVertexDict.update({(oxtAtom.residue.num, oxtAtom.residue.chn, oxtAtom.residue.name, oxtAtom.atom_name):vertex})
						self.terminiVertices.append(vertex)
						self.tscVertexDict.update({(oxtAtom.residue.num, oxtAtom.residue.chn, oxtAtom.residue.name, oxtAtom.atom_name):vertex})
						self.tscVertices.append(vertex)
						break

		# If indicated, include waters as ionizable groups in the triangulation.
		########################################################################
		if self.includeWater:
			self.tscVertexDict.update(self.waterVertexDict)
			self.allAtomVertexDict.update(self.waterVertexDict)
			self.tscVertices += self.waterVertices
			self.allAtomVertices += self.waterVertices

		# If indicated, include ions as ionizable groups in the triangulation.
		######################################################################
		if self.includeIons:
			ionizableSet.extend(ions)
			for vKey in self.hetVertexDict:
				if vKey[3] in ions:
					self.tscVertexDict.update({vKey:self.hetVertexDict[vKey]})
					self.allAtomVertexDict.update({vKey:self.hetVertexDict[vKey]})
					self.tscVertices.append(self.hetVertexDict[vKey])
					self.allAtomVertices.append(self.hetVertexDict[vKey])

	def selectTscTriangulationAtoms(self, returnCa=0, returnSidechain=0):

		self.tscVerticesSelection, self.tscVertexDictSelection = [], {}
		for tscVertexKey in self.tscVertexDict:
			if self.tscVertexDict[tscVertexKey].data.chain_identifier not in self.chains:
				continue
			if self.tscVertexDict[tscVertexKey].data.residue.name in self.queryResidueSet:
				if returnCa:
					caAtom = self.tscVertexDict[tscVertexKey].data.residue.atoms["CA"]
					caAtomKey = (caAtom.residue.num, caAtom.residue.chn, caAtom.residue.name, caAtom.atom_name)
					self.tscVerticesSelection.append(self.caVertexDict[caAtomKey])
					self.tscVertexDictSelection.update({caAtomKey:self.caVertexDict[caAtomKey]})
				elif returnSidechain:
					scAtoms = self.tscVertexDict[tscVertexKey].data.residue.get_sidechain_atoms()
					for scAtom in scAtoms:
						scAtomKey = (scAtom.residue.num, scAtom.residue.chn, scAtom.residue.name, scAtom.atom_name)
						self.tscVerticesSelection.append(self.allAtomVertexDict[scAtomKey])
						self.tscVertexDictSelection.update({scAtomKey:self.allAtomVertexDict[scAtomKey]})
				else:
					self.tscVerticesSelection.append(self.tscVertexDict[tscVertexKey])
					self.tscVertexDictSelection.update({tscVertexKey:self.tscVertexDict[tscVertexKey]})

		# If there are no vertices in the selection, exit pHinder
		if not self.tscVerticesSelection:
			print("FAIL: No TSC atoms in this PDB file...")
			return 0

		print("Done selecting TSC atoms to triangulates....")

	def selectTscClassificationAtoms(self, returnCa=0):

		self.tscVerticesSelection, self.tscVertexDictSelection = [], {}
		for tscVertexKey in self.tscVertexDict:
			if self.tscVertexDict[tscVertexKey].data.chain_identifier not in self.chains:
				continue
			if self.tscVertexDict[tscVertexKey].data.residue.name in self.queryResidueSet:
				if not returnCa:
					self.tscVerticesSelection.append(self.tscVertexDict[tscVertexKey])
					self.tscVertexDictSelection.update({tscVertexKey:self.tscVertexDict[tscVertexKey]})
				else:
					caAtom = self.tscVertexDict[tscVertexKey].data.residue.atoms["CA"]
					caAtomKey = (caAtom.residue.num, caAtom.residue.chn, caAtom.residue.name, caAtom.atom_name)
					self.tscVerticesSelection.append(self.caVertexDict[caAtomKey])
					self.tscVertexDictSelection.update({caAtomKey:self.caVertexDict[caAtomKey]})

		# If there are no vertices in the selection, exit pHinder
		if not self.tscVerticesSelection:
			print("FAIL: No ionizable side chains in this PDB chain...")
			return 0

		print("Done selecting TSC atoms to triangulates....")

	def triangulateTscAtoms(self):

		# Grouped PDB chains...
		if self.group_chains:

			# The global side chain topology network
			########################################
			# Triangulate the selected set of protein vertices
			print("\n\nTriangulating query vertex set....chain group", self.chainString)
			print(80*"-")
			self.tscTriangulation = convexHull4D(self.tscVertexDictSelection.values())
			print("Done triangulating query vertex set....chain group", self.chainString)

			# CRITICAL FUNCTION CALL !!!!
			# The orientation of the networks edges needs to be updated after the maxNetworkEdgeLength condition is applied.
			# Also, rename the original triangulation, which will not be mutated, for clarity.
			self.tscTriangulation.orientS2s()
			self.originalTscTriangulation = self.tscTriangulation.s1Dict
			# Make a copy of the triangulation that can be mutated.
			self.mutatedTscTriangulation = deepcopy(self.tscTriangulation.s1Dict)

			print("Done triangulating TSC atoms....chain group", self.chainString)

		# Individual PDB chains...
		else:

			# The global side chain topology network
			########################################
			# Triangulate the selected set of protein vertices
			print("\n\nTriangulating query vertex set....as ungrouped chains", self.chainString)
			print(80*"-")
			self.tscTriangulation = convexHull4D(self.tscVertexDictSelection.values())
			print("Done triangulating query vertex set....as ungrouped chains", self.chainString)

			# CRITICAL FUNCTION CALL !!!!
			# The orientation of the networks edges needs to be updated after the maxNetworkEdgeLength condition is applied.
			# Also, rename the original triangulation, which will not be mutated, for clarity.
			self.tscTriangulation.orientS2s()
			self.originalTscTriangulation = self.tscTriangulation.s1Dict
			# Make a copy of the triangulation that can be mutated.
			self.mutatedTscTriangulation = deepcopy(self.tscTriangulation.s1Dict)


			# Side chain topology network per chain
			#######################################
			for chain in self.chains:

				tscVertexDictSelection = {}
				for key in self.tscVertexDictSelection:
					if key[1] == chain:
						tscVertexDictSelection[key] = self.tscVertexDictSelection[key]

				# Triangulate the selected set of protein vertices
				print("\n\nTriangulating query vertex set....chain", chain)
				print(80*"-")
				tscTriangulation = convexHull4D(tscVertexDictSelection.values())
				print("Done triangulating query vertex set....chain", chain)

				# CRITICAL FUNCTION CALL !!!!
				# The orientation of the networks edges needs to be updated after the maxNetworkEdgeLength condition is applied.
				# Also, rename the original triangulation, which will not be mutated, for clarity.
				tscTriangulation.orientS2s()
				originalTscTriangulation = tscTriangulation.s1Dict
				# Make a copy of the triangulation that can be mutated.
				mutatedTscTriangulation = deepcopy(tscTriangulation.s1Dict)

				self.triangulations_bychain[chain] = [tscTriangulation, mutatedTscTriangulation, originalTscTriangulation]
				print("Done triangulating TSC atoms....chain", chain)

	def writeTriangulation(self):

		# Write the triangulation of the side chains used for the calculation.
		if self.saveNetworkTriangulation and self.tscTriangulation:

			# Grouped PDB chains...
			if self.group_chains:

				coordinateDoubles = []
				for s2 in self.tscTriangulation.getS2s():
					t1 = self.tscTriangulation.s1Dict[s2[0]].s1.coordinate_tuple[0:-1]
					t2 = self.tscTriangulation.s1Dict[s2[1]].s1.coordinate_tuple[0:-1]
					coordinateDoubles.append((t1, t2))

				script_body = "triangulation_name = " + "\"triangulation." + self.pdbCode + "-" + self.chainString + "\"\n"
				script_body += "coordinateDoubles = " + str(coordinateDoubles) + "\n"
				script_body += "from pymol.cgo import *\n"
				script_body += "from pymol import cmd\n"
				script_body += "background_color = \"white\"\n"
				script_body += "cmd.bg_color(color=background_color)\n"
				script_body += "cgo_line_width = 3.0\n"
				script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
				script_body += "obj = [ BEGIN, LINES,\n"
				script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
				script_body += "for coordinateDouble in coordinateDoubles:\n"
				script_body += "	c = coordinateDouble\n"
				script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
				script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
				script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
				script_body += "obj.extend([END,])\n"
				script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
				script_body += "cmd.load_cgo(obj, triangulation_name)\n"

				outFile = open(self.outPath + self.pdbCode + ".triangulation-" + self.chainString + ".py", "w")
				outFile.write(script_body)
				outFile.close()

				print("Done writing amino acid side chain triangulation to file....")

			# Individual PDB chains...
			else:

				for chain in self.chains:

					tscTriangulation = self.triangulations_bychain[chain][0]

					coordinateDoubles = []
					for s2 in tscTriangulation.getS2s():
						t1 = tscTriangulation.s1Dict[s2[0]].s1.coordinate_tuple[0:-1]
						t2 = tscTriangulation.s1Dict[s2[1]].s1.coordinate_tuple[0:-1]
						coordinateDoubles.append((t1, t2))

					script_body = "triangulation_name = " + "\"triangulation." + self.pdbCode + "-" + chain + "\"\n"
					script_body += "coordinateDoubles = " + str(coordinateDoubles) + "\n"
					script_body += "from pymol.cgo import *\n"
					script_body += "from pymol import cmd\n"
					script_body += "background_color = \"white\"\n"
					script_body += "cmd.bg_color(color=background_color)\n"
					script_body += "cgo_line_width = 3.0\n"
					script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
					script_body += "obj = [ BEGIN, LINES,\n"
					script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
					script_body += "for coordinateDouble in coordinateDoubles:\n"
					script_body += "	c = coordinateDouble\n"
					script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
					script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
					script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
					script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x3, y3, z3,])\n"
					script_body += "obj.extend([END,])\n"
					script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
					script_body += "cmd.load_cgo(obj, triangulation_name)\n"

					outPath = self.outPaths_bychain[chain]
					outFile = open(outPath + self.pdbCode + ".triangulation-" + chain + ".py", "w")
					outFile.write(script_body)
					outFile.close()

					print("Done writing amino acid side chain triangulation to file....chain", chain)

	def pruneTriangulation(self):

		# Grouped PDB chains...
		if self.group_chains:

			# The global side chain topology network
			########################################
			pruneResult = pruneTriangulationGoFo(self.mutatedTscTriangulation, self.maxNetworkEdgeLength)
			self.networks = pruneResult[1]

			print("Done pruning triangulation....")

		# Individual PDB chains...
		else:

			# The global side chain topology network
			########################################
			pruneResult = pruneTriangulationGoFo(self.mutatedTscTriangulation, self.maxNetworkEdgeLength)
			self.networks = pruneResult[1]

			print("Done pruning triangulation....")

			# Side chain topology network per chain
			#######################################
			for chain in self.chains:

				mutatedTscTriangulation = self.triangulations_bychain[chain][1]

				pruneResult = pruneTriangulationGoFo(mutatedTscTriangulation, self.maxNetworkEdgeLength)
				networks = pruneResult[1]
				self.topology_networks_bychain[chain] = networks

				print("Done pruning triangulation....chain", chain)

	def minimizePrunedTriangulation(self):

		# Minimize the pruned triangulation.
		# This algorithm removes redundant network connections by biasing shorter network edges.
		# Note: getFinalMinimizedNetworks() must be called to update self.network with the results of the minimization.

		# Grouped PDB chains...
		if self.group_chains:

			print("2, here---------------------------->>>>>>>>>>>")

			# The global side chain topology network
			#########################################
			minimizeNetworks(self.networks, self.mutatedTscTriangulation)
			print("Done minimizing pruned triangulation....")

			self.networks = getFinalMinimizedNetworks(self.networks)
			print("Done updating minimized networks....")

			# Write networks to PDB format with CONECT data...
			if self.networks:
				self.topology_networks = self.networks
				writeNetworkToFile(self.networks, self.mutatedTscTriangulation, "topo", self.outPath, self.pdbCode, chain=self.chainString)
			print("Done writing network reports...")

		# Individual PDB chains...
		else:

			# Side chain topology network per chain
			#######################################
			for chain in self.chains:

				networks = self.topology_networks_bychain[chain]
				mutatedTscTriangulation = self.triangulations_bychain[chain][1]
				minimizeNetworks(networks, mutatedTscTriangulation)
				print("Done minimizing pruned triangulation....chain", chain)

				networks = getFinalMinimizedNetworks(networks)
				print("Done updating minimized networks....chain", chain)

				# Write networks to PDB format with CONECT data...
				if networks:
					outPath = self.outPaths_bychain[chain]
					self.topology_networks_bychain[chain] = networks
					writeNetworkToFile(networks, mutatedTscTriangulation, "topo", outPath, self.pdbCode, chain=chain)
				print("Done writing network reports...chain", chain)

	# def triangulateAllTscAtomsForClassfication(self):

	# 	# Grouped PDB chains...
	# 	if self.group_chains:

	# 		# Triangulate all the protein vertices
	# 		print("\n\nTriangulating all vertices....chain group", self.chainString)
	# 		print(80*"-")
	# 		self.tscTriangulationClassification = convexHull4D(self.tscVertexDictSelection.values())
	# 		print("Done triangulating all vertices....chain group", self.chainString)

	# 		# CRITICAL FUNCTION CALL !!!!
	# 		# The orientation of the networks edges needs to be updated after the maxNetworkEdgeLength condition is applied.
	# 		# Also, rename the original triangulation, which will not be mutated, for clarity.
	# 		self.tscTriangulationClassification.orientS2s()
	# 		self.originalTscTriangulationClassification = self.tscTriangulationClassification.s1Dict
	# 		print("Done triangulating all TSC atoms....chain group", self.chainString)

	# 	# Individual PDB chains...
	# 	else:

	# 		for chain in self.chains:

	# 			tscVertexDictSelection = {}
	# 			for key in self.tscVertexDict:
	# 				if key[1] == chain:
	# 					tscVertexDictSelection[key] = self.tscVertexDict[key]

	# 			# Triangulate all the protein vertices
	# 			print("\n\nriangulating all vertices....chain", chain)
	# 			print(80*"-")
	# 			tscTriangulation = convexHull4D(tscVertexDictSelection.values())
	# 			print("Done triangulating all vertices....chain", chain)

	# 			# CRITICAL FUNCTION CALL !!!!
	# 			# The orientation of the networks edges needs to be updated after the maxNetworkEdgeLength condition is applied.
	# 			# Also, rename the original triangulation, which will not be mutated, for clarity.
	# 			tscTriangulation.orientS2s()
	# 			originalTscTriangulation = tscTriangulation.s1Dict

	# 			self.triangulations_bychain_classfication[chain] = [tscTriangulation, mutatedTscTriangulation, originalTscTriangulation]
	# 			print("Done triangulating TSC atoms....chain", chain)

	def identifyTightBonds(self):

		# Set the list of negatively charged (minus) and postively charged (plus) residues...
		minusList, plusList = ["ASP", "GLU", "CYS"], ["HIS", "LYS", "ARG"]
		if self.residueSet == "ionizableSetNoCys":
			minusList = ["ASP", "GLU"]

		# Grouped PDB chains...
		if self.group_chains:

			# Identify ion pairs...
			#####################################################################################
			ionPairs = {}
			for network in self.networks:
				for node in network:
					node_atom = self.mutatedTscTriangulation[node].s1.data
					for s2 in self.mutatedTscTriangulation[node].s2s:
						for node_neighbor in s2:

							if node_neighbor == node:
								continue
								
							node_neighbor_atom = self.mutatedTscTriangulation[node_neighbor].s1.data
							rn1 = node_atom.residue_name 
							rn2 = node_neighbor_atom.residue_name

							is_ion_pair = 0
							if rn1 in minusList and rn2 in plusList:
								is_ion_pair = 1
							elif rn2 in minusList and rn1 in plusList:
								is_ion_pair = 1
							if is_ion_pair:

								d = distance(node_atom, node_neighbor_atom)
								if d < self.ionPairCutoff:
									key = sorted([node_atom.residue_sequence_number, node_neighbor_atom.residue_sequence_number])
									ionPairs[tuple(key)] = {node:None, node_neighbor:None}
										
			networks = {}
			for ionPair in ionPairs:
				network = sorted(ionPairs[ionPair])
				networks[tuple(network)] = ionPairs[ionPair]

			# Copy the minimized topologial network of amino acid side chains and trim...
			ionPairTriangulation = deepcopy(self.mutatedTscTriangulation)

			for network in networks:
				for node in network:
					networks[network][node] = ionPairTriangulation[node].s1
					s2Keep = {}
					for s2_key in ionPairTriangulation[node].s2s:
						if s2_key[0] in network and s2_key[1] in network:
							# Residue names (could include ion names if included)...
							rn1 = ionPairTriangulation[s2_key[0]].s1.data.residue_name 
							rn2 = ionPairTriangulation[s2_key[1]].s1.data.residue_name
							if rn1 in minusList and rn2 in plusList:
								s2Keep[s2_key] = None
							elif rn2 in minusList and rn1 in plusList:
								s2Keep[s2_key] = None
					s2Keep = list(s2Keep)
					ionPairTriangulation[node].s2s = s2Keep

			# Format and write the PDB file...
			if networks:
				getFinalMinimizedNetworks(networks)
				self.ionpair_networks = networks
				writeNetworkToFile(networks, ionPairTriangulation, "ionPair", self.outPath, self.pdbCode, chain=self.chainString)

			# Identify disulfide bonds...
			#####################################################################################
			disulfideBonds = {}
			for network in self.networks:
				for node in network:
					node_atom = self.mutatedTscTriangulation[node].s1.data
					if node_atom.residue_name != "CYS":
						continue
					for s2 in self.mutatedTscTriangulation[node].s2s:
						for node_neighbor in s2:
							if node_neighbor == node:
								continue
							node_neighbor_atom = self.mutatedTscTriangulation[node_neighbor].s1.data
							if node_neighbor_atom.residue_name == "CYS":
								d = distance(node_atom, node_neighbor_atom)
								if d < self.ionPairCutoff:
									key = sorted([node_atom.residue_sequence_number, node_neighbor_atom.residue_sequence_number])
									disulfideBonds[tuple(key)] = {node:None, node_neighbor:None}

			networks = {}
			for disulfideBond in disulfideBonds:
				network = sorted(disulfideBonds[disulfideBond])
				networks[tuple(network)] = disulfideBonds[disulfideBond]

			# Copy the minimized topologial network of amino acid side chains and trim...
			disulfideBondTriangulation = deepcopy(self.mutatedTscTriangulation)

			for network in networks:
				for node in network:
					networks[network][node] = disulfideBondTriangulation[node].s1
					s2Keep = {}
					for s2_key in disulfideBondTriangulation[node].s2s:
						if s2_key[0] in network and s2_key[1] in network:
							# Residue names (could include ion names if included)...
							rn1 = disulfideBondTriangulation[s2_key[0]].s1.data.residue_name 
							rn2 = disulfideBondTriangulation[s2_key[1]].s1.data.residue_name
							if rn1 == "CYS" and rn2 == "CYS":
								s2Keep[s2_key] = None
					s2Keep = list(s2Keep)
					disulfideBondTriangulation[node].s2s = s2Keep

			# Format and write the PDB file...
			if networks:
				getFinalMinimizedNetworks(networks)
				self.disulfide_networks = networks
				writeNetworkToFile(networks, disulfideBondTriangulation, "disulfideBond", self.outPath, self.pdbCode, chain=self.chainString)

			print("Done identifying ion pairs and disulfide bonds...")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				networks = self.topology_networks_bychain[chain]
				mutatedTscTriangulation = self.triangulations_bychain[chain][1]

				# Identify ion pairs...
				#####################################################################################
				ionPairs = {}
				for network in networks:
					for node in network:
						node_atom = mutatedTscTriangulation[node].s1.data
						for s2 in mutatedTscTriangulation[node].s2s:
							for node_neighbor in s2:
								if node_neighbor == node:
									continue
								node_neighbor_atom = mutatedTscTriangulation[node_neighbor].s1.data
								if node_atom.residue_name in minusList:
									if node_neighbor_atom.residue_name in plusList:
										d = distance(node_atom, node_neighbor_atom)
										if d < self.ionPairCutoff:
											key = sorted([node_atom.residue_sequence_number, node_neighbor_atom.residue_sequence_number])
											ionPairs[tuple(key)] = {node:None, node_neighbor:None}
								else:
									if node_neighbor_atom.residue_name in minusList:
										d = distance(node_atom, node_neighbor_atom)
										if d < self.ionPairCutoff:
											key = sorted([node_atom.residue_sequence_number, node_neighbor_atom.residue_sequence_number])
											ionPairs[tuple(key)] = {node:None, node_neighbor:None}	
											
				ion_networks = {}
				for ionPair in ionPairs:
					network = sorted(ionPairs[ionPair])
					ion_networks[tuple(network)] = ionPairs[ionPair]

				# Copy the minimized topologial network of amino acid side chains and trim...
				ionPairTriangulation = deepcopy(mutatedTscTriangulation)

				for network in ion_networks:
					for node in network:
						ion_networks[network][node] = ionPairTriangulation[node].s1
						s2Keep = {}
						for s2_key in ionPairTriangulation[node].s2s:
							if s2_key[0] in network and s2_key[1] in network:
								# Residue names (could include ion names if included)...
								rn1 = ionPairTriangulation[s2_key[0]].s1.data.residue_name 
								rn2 = ionPairTriangulation[s2_key[1]].s1.data.residue_name
								if rn1 in minusList and rn2 in plusList:
									s2Keep[s2_key] = None
								elif rn2 in minusList and rn1 in plusList:
									s2Keep[s2_key] = None
						s2Keep = list(s2Keep)
						ionPairTriangulation[node].s2s = s2Keep

				# Format and write the PDB file...
				if ion_networks:
					outPath = self.outPaths_bychain[chain]
					getFinalMinimizedNetworks(ion_networks)
					self.ionpair_networks_bychain[chain] = ion_networks
					writeNetworkToFile(ion_networks, ionPairTriangulation, "ionPair", outPath, self.pdbCode, chain=chain)

				# Identify disulfide bonds...
				#####################################################################################
				disulfideBonds = {}
				for network in networks:
					for node in network:
						node_atom = mutatedTscTriangulation[node].s1.data
						if node_atom.residue_name != "CYS":
							continue
						for s2 in mutatedTscTriangulation[node].s2s:
							for node_neighbor in s2:
								if node_neighbor == node:
									continue
								node_neighbor_atom = mutatedTscTriangulation[node_neighbor].s1.data
								if node_neighbor_atom.residue_name == "CYS":
									d = distance(node_atom, node_neighbor_atom)
									if d < self.ionPairCutoff:
										disulfideBonds[node_atom.residue.key] = {node:None, node_neighbor:None}
										disulfideBonds[node_neighbor_atom.residue.key] = {node:None, node_neighbor:None}
				self.disulfideBonds_bychain[chain] = disulfideBonds

				disulfide_networks = {}
				for disulfideBond in disulfideBonds:
					network = sorted(disulfideBonds[disulfideBond])
					disulfide_networks[tuple(network)] = disulfideBonds[disulfideBond]

				# Copy the minimized topologial network of amino acid side chains and trim...
				disulfideBondTriangulation = deepcopy(mutatedTscTriangulation)

				for network in disulfide_networks:
					for node in network:
						disulfide_networks[network][node] = disulfideBondTriangulation[node].s1
						s2Keep = {}
						for s2_key in disulfideBondTriangulation[node].s2s:
							if s2_key[0] in network and s2_key[1] in network:
								# Residue names (could include ion names if included)...
								rn1 = disulfideBondTriangulation[s2_key[0]].s1.data.residue_name 
								rn2 = disulfideBondTriangulation[s2_key[1]].s1.data.residue_name
								if rn1 == "CYS" and rn2 == "CYS":
									s2Keep[s2_key] = None
						s2Keep = list(s2Keep)
						disulfideBondTriangulation[node].s2s = s2Keep

				# Format and write the PDB file...
				if disulfide_networks:
					outPath = self.outPaths_bychain[chain]
					getFinalMinimizedNetworks(disulfide_networks)
					self.disulfide_networks_bychain[chain] = disulfide_networks
					writeNetworkToFile(disulfide_networks, disulfideBondTriangulation, "disulfideBond", outPath, self.pdbCode, chain=chain)

				print("Done identifying ion pairs and disulfide bonds...chain", chain)

	def calculateNetworkParity(self):

		# Calculate parity
		if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:

			# Set the list of negatively charged (minus) and postively charged (plus) residues...
			minusList, plusList = ["ASP", "GLU", "CYS"], ["HIS", "LYS", "ARG"]
			if self.residueSet == "ionizableSetNoCys":
				minusList = ["ASP", "GLU"]

			# Set the list of negatively and postively charged ions, if they are included in the calculation...
			if self.includeIons:
				minusList.extend(["CL", "IOD", "BR"])
				plusList.extend(["MG", "CA", "ZN", "NA", "MN", "K", "CD", "FE", "SR", "CU", "HG", "CO", "NI", "FE2"])

			# Grouped PDB chains...
			if self.group_chains:

				parityNetworks = {}
				for network in self.networks:
					# Use goFo to calculate sub-networks with charge parity.
					for node in network:
						parityNetwork = {}
						goFoParity(node, parityNetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, self.disulfideBonds)
						if len(parityNetwork) > 1:
							parityNetworks[tuple(sorted(parityNetwork))] = parityNetwork
				
				# Copy the minimized topologial network of amino acid side chains and trim for parity...
				parityTriangulation = deepcopy(self.mutatedTscTriangulation)

				for network in parityNetworks:
					for node in network:
						parityNetworks[network][node] = parityTriangulation[node].s1
						s2Keep = {}
						for s2_key in parityTriangulation[node].s2s:
							if s2_key[0] in network and s2_key[1] in network:
								# Residue names (could include ion names if included)...
								rn1 = parityTriangulation[s2_key[0]].s1.data.residue_name 
								rn2 = parityTriangulation[s2_key[1]].s1.data.residue_name
								# Trim for parity
								if rn1 in minusList and rn2 in plusList:
									s2Keep[s2_key] = None
								elif rn2 in minusList and rn1 in plusList:
									s2Keep[s2_key] = None
						s2Keep = list(s2Keep)
						parityTriangulation[node].s2s = s2Keep

				# The most efficient network topology may hide some parity...
				# This loop recovers parity nodes lost to topological efficiency...
				if parityNetworks:
					for network in parityNetworks:
						for node in network:
							for s2_key in self.originalTscTriangulation[node].s2s:
								if s2_key not in parityTriangulation[node].s2s:
									# Residue names (could include ion names if included)...
									rn1 = parityTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = parityTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in minusList and rn2 in plusList:
										d = distance(parityTriangulation[s2_key[0]].s1.data, parityTriangulation[s2_key[1]].s1.data)
										if d <= self.maxNetworkEdgeLength:
											parityTriangulation[node].s2s.append(s2_key)	
									elif rn1 in plusList and rn2 in minusList:
										d = distance(parityTriangulation[s2_key[0]].s1.data, parityTriangulation[s2_key[1]].s1.data)
										if d <= self.maxNetworkEdgeLength:
											parityTriangulation[node].s2s.append(s2_key)	

				# Format and write the finalized charge parity network to PDB file...
				if parityNetworks:
					getFinalMinimizedNetworks(parityNetworks)
					self.parity_networks = parityNetworks
					writeNetworkToFile(parityNetworks, parityTriangulation, "parity", self.outPath, self.pdbCode, chain=self.chainString)

				#
				# Identify sub-networks with anti-parity
				#
				#######################################################################################################
				positiveNetworks, negativeNetworks = {}, {}
				for network in self.networks:

					# Identify anti-parity sub-networks.
					for node in network:

						# Collect postively charged parity networks.
						if self.mutatedTscTriangulation[node].s1.data.residue.name in plusList:

							positiveSubNetwork = {}
							goFoAntiParity(node, positiveSubNetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, "plus", self.disulfideBonds)
							if len(positiveSubNetwork) > 1:
								positiveNetworks[tuple(sorted(positiveSubNetwork))] = positiveSubNetwork

						# Collect negatively charged parity networks.
						if self.mutatedTscTriangulation[node].s1.data.residue.name in minusList:

							negativeSubNetwork = {}
							goFoAntiParity(node, negativeSubNetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, "minus", self.disulfideBonds)
							if len(negativeSubNetwork) > 1:
								negativeNetworks[tuple(sorted(negativeSubNetwork))] = negativeSubNetwork

				# Copy the minimized topologial network of amino acid side chains and trim for postive anti-parity...
				######################################################################################################
				positiveTriangulation = deepcopy(self.mutatedTscTriangulation)
				for network in positiveNetworks:
					for node in network:
						positiveNetworks[network][node] = positiveTriangulation[node].s1
						s2Keep = {}
						for s2_key in positiveTriangulation[node].s2s:
							if s2_key[0] in network and s2_key[1] in network:
								# Residue names (could include ion names if included)...
								rn1 = positiveTriangulation[s2_key[0]].s1.data.residue_name 
								rn2 = positiveTriangulation[s2_key[1]].s1.data.residue_name
								# Trim for parity
								if rn1 in plusList and rn2 in plusList:
									s2Keep[s2_key] = None
						s2Keep = list(s2Keep)
						positiveTriangulation[node].s2s = s2Keep

				# The most efficient network topology may hide some antiparity...
				# This loop recovers antiparity nodes lost to topological efficiency...
				if positiveNetworks:
					for network in positiveNetworks:
						for node in network:
							for s2_key in self.originalTscTriangulation[node].s2s:
								if s2_key not in positiveTriangulation[node].s2s:
									# Residue names (could include ion names if included)...
									rn1 = positiveTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = positiveTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in minusList and rn2 in minusList:
										d = distance(positiveTriangulation[s2_key[0]].s1.data, positiveTriangulation[s2_key[1]].s1.data)
										if d <= self.maxNetworkEdgeLength:
											positiveTriangulation[node].s2s.append(s2_key)	

				# Format and write the finalized positively charged anti-parity network to PDB file...
				if positiveNetworks:
					getFinalMinimizedNetworks(positiveNetworks)
					self.parity_positive_networks = positiveNetworks
					writeNetworkToFile(positiveNetworks, positiveTriangulation, "parityPositive", self.outPath, self.pdbCode, chain=self.chainString)

				# Copy the minimized topologial network of amino acid side chains and trim for negative anti-parity...
				######################################################################################################
				negativeTriangulation = deepcopy(self.mutatedTscTriangulation)
				for network in negativeNetworks:
					for node in network:
						negativeNetworks[network][node] = negativeTriangulation[node].s1
						s2Keep = {}
						for s2_key in negativeTriangulation[node].s2s:
							if s2_key[0] in network and s2_key[1] in network:
								# Residue names (could include ion names if included)...
								rn1 = negativeTriangulation[s2_key[0]].s1.data.residue_name 
								rn2 = negativeTriangulation[s2_key[1]].s1.data.residue_name
								# Trim for parity
								if rn1 in minusList and rn2 in minusList:
									s2Keep[s2_key] = None
						s2Keep = list(s2Keep)
						negativeTriangulation[node].s2s = s2Keep

				# The most efficient network topology may hide some antiparity...
				# This loop recovers antiparity nodes lost to topological efficiency...
				if negativeNetworks:
					for network in negativeNetworks:
						for node in network:
							for s2_key in self.originalTscTriangulation[node].s2s:
								if s2_key not in negativeTriangulation[node].s2s:
									# Residue names (could include ion names if included)...
									rn1 = negativeTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = negativeTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in minusList and rn2 in minusList:
										d = distance(negativeTriangulation[s2_key[0]].s1.data, negativeTriangulation[s2_key[1]].s1.data)
										if d <= self.maxNetworkEdgeLength:
											negativeTriangulation[node].s2s.append(s2_key)				

				# Format and write the finalized negatively charged anti-parity network to PDB file...
				if negativeNetworks:
					getFinalMinimizedNetworks(negativeNetworks)
					self.parity_negative_networks = negativeNetworks
					writeNetworkToFile(negativeNetworks, negativeTriangulation, "parityNegative", self.outPath, self.pdbCode, chain=self.chainString)

				print("Done calculating network parity...")

			# Individual PDB chains...
			else:

				for chain in self.chains:

					networks = self.topology_networks_bychain[chain]
					mutatedTscTriangulation = self.triangulations_bychain[chain][1]
					disulfideBonds = self.disulfideBonds_bychain[chain]

					parityNetworks = {}
					for network in networks:
						# Use goFo to calculate sub-networks with charge parity.
						for node in network:
							parityNetwork = {}
							goFoParity(node, parityNetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, disulfideBonds)
							if len(parityNetwork) > 1:
								parityNetworks[tuple(sorted(parityNetwork))] = parityNetwork
					
					# Copy the minimized topologial network of amino acid side chains and trim for parity...
					parityTriangulation = deepcopy(mutatedTscTriangulation)

					for network in parityNetworks:
						for node in network:
							parityNetworks[network][node] = parityTriangulation[node].s1
							s2Keep = {}
							for s2_key in parityTriangulation[node].s2s:
								if s2_key[0] in network and s2_key[1] in network:
									# Residue names (could include ion names if included)...
									rn1 = parityTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = parityTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in minusList and rn2 in plusList:
										s2Keep[s2_key] = None
									elif rn2 in minusList and rn1 in plusList:
										s2Keep[s2_key] = None
							s2Keep = list(s2Keep)
							parityTriangulation[node].s2s = s2Keep

					# Format and write the finalized charge parity network to PDB file...
					if parityNetworks:
						outPath = self.outPaths_bychain[chain]
						getFinalMinimizedNetworks(parityNetworks)
						self.parity_networks_bychain[chain] = parityNetworks
						writeNetworkToFile(parityNetworks, parityTriangulation, "parity", outPath, self.pdbCode, chain=chain)

					#
					# Identify sub-networks with anti-parity
					#
					#######################################################################################################
					positiveNetworks, negativeNetworks = {}, {}
					for network in networks:

						# Identify anti-parity sub-networks.
						for node in network:

							# Collect postively charged parity networks.
							if mutatedTscTriangulation[node].s1.data.residue.name in plusList:

								positiveSubNetwork = {}
								goFoAntiParity(node, positiveSubNetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, "plus", disulfideBonds)
								if len(positiveSubNetwork) > 1:
									positiveNetworks[tuple(sorted(positiveSubNetwork))] = positiveSubNetwork

							# Collect negatively charged parity networks.
							if mutatedTscTriangulation[node].s1.data.residue.name in minusList:

								negativeSubNetwork = {}
								goFoAntiParity(node, negativeSubNetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, "minus", disulfideBonds)
								if len(negativeSubNetwork) > 1:
									negativeNetworks[tuple(sorted(negativeSubNetwork))] = negativeSubNetwork

					# Copy the minimized topologial network of amino acid side chains and trim for postive anti-parity...
					######################################################################################################
					positiveTriangulation = deepcopy(mutatedTscTriangulation)
					for network in positiveNetworks:
						for node in network:
							positiveNetworks[network][node] = positiveTriangulation[node].s1
							s2Keep = {}
							for s2_key in positiveTriangulation[node].s2s:
								if s2_key[0] in network and s2_key[1] in network:
									# Residue names (could include ion names if included)...
									rn1 = positiveTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = positiveTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in plusList and rn2 in plusList:
										s2Keep[s2_key] = None
							s2Keep = list(s2Keep)
							positiveTriangulation[node].s2s = s2Keep

					# Format and write the finalized positively charged anti-parity network to PDB file...
					if positiveNetworks:
						outPath = self.outPaths_bychain[chain]
						getFinalMinimizedNetworks(positiveNetworks)
						self.parity_positive_networks_bychain[chain] = positiveNetworks
						writeNetworkToFile(positiveNetworks, positiveTriangulation, "parityPositive", outPath, self.pdbCode, chain=chain)

					# Copy the minimized topologial network of amino acid side chains and trim for negative anti-parity...
					######################################################################################################
					negativeTriangulation = deepcopy(mutatedTscTriangulation)
					for network in negativeNetworks:
						for node in network:
							negativeNetworks[network][node] = negativeTriangulation[node].s1
							s2Keep = {}
							for s2_key in negativeTriangulation[node].s2s:
								if s2_key[0] in network and s2_key[1] in network:
									# Residue names (could include ion names if included)...
									rn1 = negativeTriangulation[s2_key[0]].s1.data.residue_name 
									rn2 = negativeTriangulation[s2_key[1]].s1.data.residue_name
									# Trim for parity
									if rn1 in minusList and rn2 in minusList:
										s2Keep[s2_key] = None
							s2Keep = list(s2Keep)
							negativeTriangulation[node].s2s = s2Keep

					# Format and write the finalized negatively charged anti-parity network to PDB file...
					if negativeNetworks:
						outPath = self.outPaths_bychain[chain]
						getFinalMinimizedNetworks(negativeNetworks)
						self.parity_negative_networks_bychain[chain] = negativeNetworks
						writeNetworkToFile(negativeNetworks, negativeTriangulation, "parityNegative", outPath, self.pdbCode, chain=chain)

					print("Done calculating network parity...chain", chain)

	def surface(self, circumSphereRadiusLimit=6.5, minArea=3.0):

		# Default values for parameters for protein surfaces.
		# circumSphereRadiusLimit=6.5
		# minArea=3.0
		#####################################################

		# Grouped PDB chains...
		if self.group_chains:

			# Calculate the protein surface.
			if self.writeSurfaceCreationAnimation:
				self.writeSurfaceCreationAnimation = self.outPath

			# The grouped PDB chains may not be every chain in the PDB file
			ca_vertices = []
			for ca_v in self.caVertices:
				if ca_v.data.chain_identifier in self.chains:
					ca_vertices.append(ca_v)

			surfaceResult = calculateSurface(ca_vertices, circumSphereRadiusLimit, minArea=minArea, writeSurfaceCreationAnimation=self.writeSurfaceCreationAnimation, highResolutionSurface=self.highResolutionSurface, allowSmallSurfaces=self.allowSmallSurfaces)
			# Only proceed if there is a surface to work with.
			if not surfaceResult:
				print("Using the current set of parameters, no surface could be calculated...")
				return 0
			else:
				self.hull4D = surfaceResult[0]
				self.quadSurface = surfaceResult[1]

			print("Done calculating surface...")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				# Select the C-alpha vertices in the chain.
				caVertices = []
				for caVertex in self.caVertices:
					if caVertex.data.residue.chn == chain:
						caVertices.append(caVertex)

				# Calculate the protein surface.
				writeSurfaceCreationAnimation = None
				if self.writeSurfaceCreationAnimation:
					writeSurfaceCreationAnimation = self.outPaths_bychain[chain]
				surfaceResult = calculateSurface(caVertices, circumSphereRadiusLimit, minArea=minArea, writeSurfaceCreationAnimation=writeSurfaceCreationAnimation, highResolutionSurface=self.highResolutionSurface, allowSmallSurfaces=self.allowSmallSurfaces)
				
				# 2023.04.04 Revising to maintain chain, even if there is no surface for it
				# # Only proceed if there is a surface to work with.
				# if not surfaceResult:
				# 	print("Using the current set of parameters, no surface could be calculated...chain", chain)
				# 	return 0
				# else:
				self.hull4D_bychain[chain] = surfaceResult[0]
				self.quadSurface_bychain[chain] = surfaceResult[1]

				print("Done calculating surface...chain", chain)	

	def writeSurface(self):

		# Grouped PDB chains...
		if self.group_chains:

			coordinateTriples = []
			if self.saveSurface and self.quadSurface:
				surfaceFacets = {}
				for e in self.quadSurface:
					surfaceFacets.update({self.quadSurface[e].lf.id:self.quadSurface[e].lf})
					surfaceFacets.update({self.quadSurface[e].rf.id:self.quadSurface[e].rf})
				for fKey in surfaceFacets:
					f = surfaceFacets[fKey]
					t1 = f.v1.coordinate_tuple[0:-1]
					t2 = f.v2.coordinate_tuple[0:-1]
					t3 = f.v3.coordinate_tuple[0:-1]
					coordinateTriples.append((t1, t2, t3))

			script_body = "surfaceName = " + "\"surface." + self.pdbCode + "-" + self.chainString + "\"\n"
			script_body += "coordinateTriples = " + str(coordinateTriples) + "\n"
			script_body += "from pymol.cgo import *\n"
			script_body += "from pymol import cmd\n"
			script_body += "background_color = \"white\"\n"
			script_body += "cmd.bg_color(color=background_color)\n"
			script_body += "cgo_line_width = 2.0\n"
			script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
			script_body += "obj = [ BEGIN, LINES,\n"
			script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
			script_body += "for coordinateTriple in coordinateTriples:\n"
			script_body += "	c = coordinateTriple\n"
			script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
			script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
			script_body += "	x3, y3, z3 = float(c[2][0]), float(c[2][1]), float(c[2][2])\n"
			script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
			script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x3, y3, z3,])\n"
			script_body += "	obj.extend([VERTEX, x2, y2, z2,VERTEX, x3, y3, z3,])\n"
			script_body += "obj.extend([END,])\n"
			script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
			script_body += "cmd.load_cgo(obj, surfaceName)\n"

			outFile = open(self.outPath + self.pdbCode + ".surface-" + self.chainString + ".py", "w")
			outFile.write(script_body)
			outFile.close()

			print("Done writing surface Pymol script...")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				# The chain has no pHinder surface
				if chain not in self.quadSurface_bychain:
					continue

				# The chain has a pHinder surface
				coordinateTriples = []
				if self.saveSurface and self.quadSurface_bychain[chain]:
					surfaceFacets = {}
					for e in self.quadSurface_bychain[chain]:
						surfaceFacets.update({self.quadSurface_bychain[chain][e].lf.id:self.quadSurface_bychain[chain][e].lf})
						surfaceFacets.update({self.quadSurface_bychain[chain][e].rf.id:self.quadSurface_bychain[chain][e].rf})
					for fKey in surfaceFacets:
						f = surfaceFacets[fKey]
						t1 = f.v1.coordinate_tuple[0:-1]
						t2 = f.v2.coordinate_tuple[0:-1]
						t3 = f.v3.coordinate_tuple[0:-1]
						coordinateTriples.append((t1, t2, t3))

				script_body = "surfaceName = " + "\"surface." + self.pdbCode + "-" + chain + "\"\n"
				script_body += "coordinateTriples = " + str(coordinateTriples) + "\n"
				script_body += "from pymol.cgo import *\n"
				script_body += "from pymol import cmd\n"
				script_body += "background_color = \"white\"\n"
				script_body += "cmd.bg_color(color=background_color)\n"
				script_body += "cgo_line_width = 2.0\n"
				script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
				script_body += "obj = [ BEGIN, LINES,\n"
				script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
				script_body += "for coordinateTriple in coordinateTriples:\n"
				script_body += "	c = coordinateTriple\n"
				script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
				script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
				script_body += "	x3, y3, z3 = float(c[2][0]), float(c[2][1]), float(c[2][2])\n"
				script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
				script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x3, y3, z3,])\n"
				script_body += "	obj.extend([VERTEX, x2, y2, z2,VERTEX, x3, y3, z3,])\n"
				script_body += "obj.extend([END,])\n"
				script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
				script_body += "cmd.load_cgo(obj, surfaceName)\n"

				outFile = open(self.outPaths_bychain[chain] + self.pdbCode + ".surface-" + chain + ".py", "w")
				outFile.write(script_body)
				outFile.close()

				print("Done writing surface Pymol script...chain", chain)

	def surfaceLigands(self):

		# Could make parallel version: per ligand.

		# Grouped PDB chains...
		if self.group_chains:

			if self.saveLigandSurfaces and self.het_residues:
				
				# Calculate the surface of each hetero-residue individually.
				for residue in self.het_residues:
					
					atoms = self.het_residues[residue].atoms.values()
					iHetVertices = []
					for atom in atoms:
						atomKey = (atom.residue.num, atom.residue.chn, atom.residue.name, atom.atom_name)
						iHetVertices.append(self.hetVertexDict[atomKey])
						atomVertex = self.hetVertexDict[atomKey]
						
						# Expand the volume of the hetero-atom
						from sphere import Sphere
						print("hellow")
						s = Sphere()
						s.radius = 1.5 # Average van Der Waals radius
						s.origin_x = atomVertex.x
						s.origin_y = atomVertex.y
						s.origin_z = atomVertex.z
						# Dividing by odd numbers helps achieve general position.
						s.generate_surface_points(pi/3., pi/3., general_position=1)
						psas = s.get_pseudoatoms()
						for psa in psas:
							x, y, z, u = psa.x, psa.y, psa.z, (psa.x**2 + psa.y**2 + psa.z**2)
							pv = Vertex4D((x, y, z, u), setC=1)
							pv.data = psa
							pv.id = psa.atom_serial
							iHetVertices.append(pv)

					if iHetVertices:
				
						# Calculate the ligand surface.
						heteroSurfaceResult = calculateSurface(iHetVertices, 3.0, minArea=2.0, writeSurfaceCreationAnimation="", highResolutionSurface=self.highResolutionSurface, allowSmallSurfaces=50)
						if not heteroSurfaceResult:
							continue
						heteroHull4D = heteroSurfaceResult[0]
						heteroQuadSurface = heteroSurfaceResult[1]
						self.ligandSurfaces[residue] = heteroQuadSurface

			print("Done calculating ligand surfaces...")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				if self.saveLigandSurfaces and self.het_residues and chain in self.ligandSurfaces_bychain:

					# Calculate the surface of each hetero-residue individually.
					self.ligandSurfaces_bychain[chain] = {}
					for residue_key in self.het_residues:

						residue = self.het_residues[residue_key]
						if residue.chn == chain:

							atoms = residue.atoms.values()
							iHetVertices = []
							for atom in atoms:
								atomKey = (atom.residue.num, atom.residue.chn, atom.residue.name, atom.atom_name)
								iHetVertices.append(self.hetVertexDict[atomKey])
								atomVertex = self.hetVertexDict[atomKey]
								
								# Expand the volume of the hetero-atom
								from sphere import Sphere
								s = Sphere()
								s.radius = 1.5 # Average van Der Waals radius
								s.origin_x = atomVertex.x
								s.origin_y = atomVertex.y
								s.origin_z = atomVertex.z
								# Dividing by odd numbers helps achieve general position.
								s.generate_surface_points(pi/3., pi/3., general_position=1)
								psas = s.get_pseudoatoms()
								for psa in psas:
									x, y, z, u = psa.x, psa.y, psa.z, (psa.x**2 + psa.y**2 + psa.z**2)
									pv = Vertex4D((x, y, z, u), setC=1)
									pv.data = psa
									pv.id = psa.atom_serial
									iHetVertices.append(pv)

							if iHetVertices:
						
								# Calculate the ligand surface.
								heteroSurfaceResult = calculateSurface(iHetVertices, 3.0, minArea=2.0, writeSurfaceCreationAnimation="", highResolutionSurface=self.highResolutionSurface, allowSmallSurfaces=50)
								if not heteroSurfaceResult:
									continue
								heteroHull4D = heteroSurfaceResult[0]
								heteroQuadSurface = heteroSurfaceResult[1]
								self.ligandSurfaces_bychain[chain][residue.key] = heteroQuadSurface

							print("Done calculating ligand surfaces...chain", chain)

	def writeLigandSurfaces(self):

		# Grouped PDB chains...
		if self.group_chains:

			if self.saveLigandSurfaces and self.ligandSurfaces:

				for ligand_key in self.ligandSurfaces:

					ligandQuadSurface = self.ligandSurfaces[ligand_key]

					surfaceFacets, coordinateTriples = {}, []
					for e in ligandQuadSurface:
						surfaceFacets.update({ligandQuadSurface[e].lf.id:ligandQuadSurface[e].lf})
						surfaceFacets.update({ligandQuadSurface[e].rf.id:ligandQuadSurface[e].rf})
					for fKey in surfaceFacets:
						f = surfaceFacets[fKey]
						t1 = f.v1.coordinate_tuple[0:-1]
						t2 = f.v2.coordinate_tuple[0:-1]
						t3 = f.v3.coordinate_tuple[0:-1]
						coordinateTriples.append((t1, t2, t3))

					script_body = "surfaceName = " + "\"surface." + self.pdbCode + "-" + ligand_key[1] + "\"\n"
					script_body += "coordinateTriples = " + str(coordinateTriples) + "\n"
					script_body += "from pymol.cgo import *\n"
					script_body += "from pymol import cmd\n"
					script_body += "background_color = \"white\"\n"
					script_body += "cmd.bg_color(color=background_color)\n"
					script_body += "cgo_line_width = 2.0\n"
					script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
					script_body += "obj = [ BEGIN, LINES,\n"
					script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
					script_body += "for coordinateTriple in coordinateTriples:\n"
					script_body += "	c = coordinateTriple\n"
					script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
					script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
					script_body += "	x3, y3, z3 = float(c[2][0]), float(c[2][1]), float(c[2][2])\n"
					script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
					script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x3, y3, z3,])\n"
					script_body += "	obj.extend([VERTEX, x2, y2, z2,VERTEX, x3, y3, z3,])\n"
					script_body += "obj.extend([END,])\n"
					script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
					script_body += "cmd.load_cgo(obj, surfaceName)\n"

					outFile = open(self.outPath + self.pdbCode + ".surface-" + ligand_key[1] + ".py", "w")
					outFile.write(script_body)
					outFile.close()

					print("Done writing surface Pymol script...")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				if self.saveLigandSurfaces and chain in self.ligandSurfaces_bychain:

					for ligand_key in self.ligandSurfaces_bychain[chain]:

						ligandQuadSurface = self.ligandSurfaces_bychain[chain][ligand_key]

						surfaceFacets, coordinateTriples = {}, []
						for e in ligandQuadSurface:
							surfaceFacets.update({ligandQuadSurface[e].lf.id:ligandQuadSurface[e].lf})
							surfaceFacets.update({ligandQuadSurface[e].rf.id:ligandQuadSurface[e].rf})
						for fKey in surfaceFacets:
							f = surfaceFacets[fKey]
							t1 = f.v1.coordinate_tuple[0:-1]
							t2 = f.v2.coordinate_tuple[0:-1]
							t3 = f.v3.coordinate_tuple[0:-1]
							coordinateTriples.append((t1, t2, t3))

						script_body = "surfaceName = " + "\"surface." + self.pdbCode + "-" + ligand_key[1] + "-" + chain + "\"\n"
						script_body += "coordinateTriples = " + str(coordinateTriples) + "\n"
						script_body += "from pymol.cgo import *\n"
						script_body += "from pymol import cmd\n"
						script_body += "background_color = \"white\"\n"
						script_body += "cmd.bg_color(color=background_color)\n"
						script_body += "cgo_line_width = 2.0\n"
						script_body += "colorR, colorG, colorB = 0.55, 0.57, 0.67\n"
						script_body += "obj = [ BEGIN, LINES,\n"
						script_body += "		COLOR, float(colorR), float(colorG), float(colorB),]\n"
						script_body += "for coordinateTriple in coordinateTriples:\n"
						script_body += "	c = coordinateTriple\n"
						script_body += "	x1, y1, z1 = float(c[0][0]), float(c[0][1]), float(c[0][2])\n"
						script_body += "	x2, y2, z2 = float(c[1][0]), float(c[1][1]), float(c[1][2])\n"
						script_body += "	x3, y3, z3 = float(c[2][0]), float(c[2][1]), float(c[2][2])\n"
						script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x2, y2, z2,])\n"
						script_body += "	obj.extend([VERTEX, x1, y1, z1,VERTEX, x3, y3, z3,])\n"
						script_body += "	obj.extend([VERTEX, x2, y2, z2,VERTEX, x3, y3, z3,])\n"
						script_body += "obj.extend([END,])\n"
						script_body += "cmd.set('cgo_line_width', cgo_line_width)\n"
						script_body += "cmd.load_cgo(obj, surfaceName)\n"

						outFile = open(self.outPaths_bychain[chain] + self.pdbCode + ".surface-" + ligand_key[1] + "-" + chain + ".py", "w")
						outFile.write(script_body)
						outFile.close()

						print("Done writing surface Pymol script...chain", chain)

	def classifySidechainLocation(self):

		# Grouped PDB chains...
		if self.group_chains:	

			# Can't classify side chains if there is no surface 
			if not self.quadSurface:
				print("No calculable surface. Cannot classify side chains as core, margin, or exposed...")
			
			else:

				print("Classifying side chains as core, margin, or exposed...")
				print(80*"-")
				result = classifySidechains(self.tscVerticesSelection, self.tscVertexDictSelection, self.allAtomVertexDict, self.quadSurface, self.outPath, self.pdbCode, self.coreCutoff, self.marginCutoff)
				self.vExposed, self.vMargin, self.vCore = result[0], result[1], result[2]

				#
				# Exposed amino acid sub-networks...
				############################################################################################
				if self.vExposed:
					classification_subnetworks  = {}
					for network in self.networks:
						
						# Filter out self.vExposed vertices in the network
						verticesInThisNetwork = []
						for v in self.vExposed:
							if v.id in network:
								verticesInThisNetwork.append(v.id)

						touched_nodes = {}
						for node in verticesInThisNetwork:

							if node in touched_nodes:
								continue

							classification_subnetwork = {}
							goFoClassification(node, classification_subnetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, verticesInThisNetwork)
							classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
							touched_nodes.update(classification_subnetwork)

					# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
					triangulation = deepcopy(self.mutatedTscTriangulation)

					networks = classification_subnetworks
					for network in networks:
						for node in network:
							networks[network][node] = triangulation[node].s1
							s2Keep = {}
							# Network of more than one node
							if triangulation[node].s2s:
								for s2_key in triangulation[node].s2s:
									if s2_key[0] in network and s2_key[1] in network:
										# Residue names (could include ion names if included)...
										rn1 = triangulation[s2_key[0]].s1.data.residue_name 
										rn2 = triangulation[s2_key[1]].s1.data.residue_name
										s2Keep[s2_key] = None
								s2Keep = list(s2Keep)
							# Network of only one node
							else:
								s2Keep = [(node, node)]
							triangulation[node].s2s = s2Keep

					# Format and write the finalized exposed network nodes to PDB file...
					if networks:
						getFinalMinimizedNetworks(networks)
						writeNetworkToFile(networks, triangulation, "exposed", self.outPath, self.pdbCode, chain=self.chainString)

				#
				# Margin amino acid sub-networks...
				############################################################################################
				if self.vMargin:
					classification_subnetworks  = {}
					for network in self.networks:
						
						# Filter out self.vMargin vertices in the network
						verticesInThisNetwork = []
						for v in self.vMargin:
							if v.id in network:
								verticesInThisNetwork.append(v.id)

						touched_nodes = {}
						for node in verticesInThisNetwork:

							if node in touched_nodes:
								continue

							classification_subnetwork = {}
							goFoClassification(node, classification_subnetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, verticesInThisNetwork)
							classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
							touched_nodes.update(classification_subnetwork)

					# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
					triangulation = deepcopy(self.mutatedTscTriangulation)

					# networks = classification_subnetworks
					# for network in networks:
					# 	for node in network:
					# 		networks[network][node] = triangulation[node].s1
					# 		s2Keep = {}
					# 		for s2_key in triangulation[node].s2s:
					# 			if s2_key[0] in network and s2_key[1] in network:
					# 				# Residue names (could include ion names if included)...
					# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
					# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
					# 				s2Keep[s2_key] = None
					# 		s2Keep = list(s2Keep)
					# 		triangulation[node].s2s = s2Keep

					networks = classification_subnetworks
					for network in networks:
						for node in network:
							networks[network][node] = triangulation[node].s1
							s2Keep = {}
							# Network of more than one node
							if triangulation[node].s2s:
								for s2_key in triangulation[node].s2s:
									if s2_key[0] in network and s2_key[1] in network:
										# Residue names (could include ion names if included)...
										rn1 = triangulation[s2_key[0]].s1.data.residue_name 
										rn2 = triangulation[s2_key[1]].s1.data.residue_name
										s2Keep[s2_key] = None
								s2Keep = list(s2Keep)
							# Network of only one node
							else:
								s2Keep = [(node, node)]
							triangulation[node].s2s = s2Keep

					# Format and write the finalized margin network nodes to PDB file...
					if networks:
						getFinalMinimizedNetworks(networks)
						writeNetworkToFile(networks, triangulation, "margin", self.outPath, self.pdbCode, chain=self.chainString)

				#
				# Core amino acid sub-networks...
				############################################################################################
				if self.vCore:
					classification_subnetworks  = {}
					for network in self.networks:
						
						# Filter out self.vCore vertices in the network
						verticesInThisNetwork = []
						for v in self.vCore:
							if v.id in network:
								verticesInThisNetwork.append(v.id)

						touched_nodes = {}
						for node in verticesInThisNetwork:

							if node in touched_nodes:
								continue

							classification_subnetwork = {}
							goFoClassification(node, classification_subnetwork, self.mutatedTscTriangulation[node].s2s, self.mutatedTscTriangulation, verticesInThisNetwork)
							classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
							touched_nodes.update(classification_subnetwork)

					# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
					triangulation = deepcopy(self.mutatedTscTriangulation)

					# networks = classification_subnetworks
					# for network in networks:
					# 	for node in network:
					# 		networks[network][node] = triangulation[node].s1
					# 		s2Keep = {}
					# 		for s2_key in triangulation[node].s2s:
					# 			if s2_key[0] in network and s2_key[1] in network:
					# 				# Residue names (could include ion names if included)...
					# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
					# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
					# 				s2Keep[s2_key] = None
					# 		s2Keep = list(s2Keep)
					# 		triangulation[node].s2s = s2Keep

					networks = classification_subnetworks
					for network in networks:
						for node in network:
							networks[network][node] = triangulation[node].s1
							s2Keep = {}
							# Network of more than one node
							if triangulation[node].s2s:
								for s2_key in triangulation[node].s2s:
									if s2_key[0] in network and s2_key[1] in network:
										# Residue names (could include ion names if included)...
										rn1 = triangulation[s2_key[0]].s1.data.residue_name 
										rn2 = triangulation[s2_key[1]].s1.data.residue_name
										s2Keep[s2_key] = None
								s2Keep = list(s2Keep)
							# Network of only one node
							else:
								s2Keep = [(node, node)]
							triangulation[node].s2s = s2Keep

					# Format and write the finalized core network nodes to PDB file...
					if networks:
						getFinalMinimizedNetworks(networks)
						writeNetworkToFile(networks, triangulation, "core", self.outPath, self.pdbCode, chain=self.chainString)

		# Individual PDB chains...
		else:

			for chain in self.chains:

				tscVertexDictSelection = {}
				for key in self.tscVertexDictSelection:
					if key[1] == chain:
						tscVertexDictSelection[key] = self.tscVertexDictSelection[key]
				tscVerticesSelection = list(tscVertexDictSelection.values())

				# Can't classify side chains if there is no surface 
				# Therefore, classify all side chains as exposed
				################################################################################################
				if not self.quadSurface_bychain[chain]:
					print("No calculable surface. Classifying all side chains as exposed...chain", chain)
					self.vExposed_bychain[chain] = {}
					for v in tscVerticesSelection:
						v4D = tscVertexDictSelection[(v.data.residue.num, v.data.residue.chn, v.data.residue.name, v.data.atom_name)]
						classification = sidechainClassification(v4D)
						classification.depth = 0.0
						classification.exposed = 1
						classification.setClassificationString()
						# print(list(self.vExposed_bychain))
						self.vExposed_bychain[chain][classification.v] = classification
						# vExposed.update({classification.v:classification})

					# Exposed amino acid sub-networks...
					############################################################################################
					vExposed = self.vExposed_bychain[chain]
					if vExposed:
						classification_subnetworks  = {}
						for network in self.topology_networks_bychain[chain]:
							
							# Filter out self.vExposed vertices in the network
							verticesInThisNetwork = []
							for v in vExposed:
								if v.id in network:
									verticesInThisNetwork.append(v.id)

							touched_nodes = {}
							for node in verticesInThisNetwork:

								if node in touched_nodes:
									continue

								classification_subnetwork = {}
								goFoClassification(node, classification_subnetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, verticesInThisNetwork)
								classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
								touched_nodes.update(classification_subnetwork)

						# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
						triangulation = deepcopy(mutatedTscTriangulation)

						# networks = classification_subnetworks
						# for network in networks:
						# 	for node in network:
						# 		networks[network][node] = triangulation[node].s1
						# 		s2Keep = {}
						# 		for s2_key in triangulation[node].s2s:
						# 			if s2_key[0] in network and s2_key[1] in network:
						# 				# Residue names (could include ion names if included)...
						# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
						# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
						# 				s2Keep[s2_key] = None
						# 		s2Keep = list(s2Keep)
						# 		triangulation[node].s2s = s2Keep

						networks = classification_subnetworks
						for network in networks:
							for node in network:
								networks[network][node] = triangulation[node].s1
								s2Keep = {}
								# Network of more than one node
								if triangulation[node].s2s:
									for s2_key in triangulation[node].s2s:
										if s2_key[0] in network and s2_key[1] in network:
											# Residue names (could include ion names if included)...
											rn1 = triangulation[s2_key[0]].s1.data.residue_name 
											rn2 = triangulation[s2_key[1]].s1.data.residue_name
											s2Keep[s2_key] = None
									s2Keep = list(s2Keep)
								# Network of only one node
								else:
									s2Keep = [(node, node)]
								triangulation[node].s2s = s2Keep

						# Format and write the finalized exposed network nodes to PDB file...
						if networks:
							getFinalMinimizedNetworks(networks)
							outPath = self.outPaths_bychain[chain]
							writeNetworkToFile(networks, triangulation, "exposed", outPath, self.pdbCode, chain=chain)

				else:

					print("Classifying side chains as core, margin, or exposed...chain", chain)
					print(80*"-")
					mutatedTscTriangulation = self.triangulations_bychain[chain][1]
					quadSurface = self.quadSurface_bychain[chain]
					outPath = self.outPaths_bychain[chain]
					result = classifySidechains(tscVerticesSelection, tscVertexDictSelection, self.allAtomVertexDict, quadSurface, outPath, self.pdbCode, self.coreCutoff, self.marginCutoff)
					vExposed, vMargin, vCore = result[0], result[1], result[2]
					self.vExposed_bychain[chain], self.vMargin_bychain[chain], self.vCore_bychain[chain] = vExposed, vMargin, vCore

					#
					# Exposed amino acid sub-networks...
					############################################################################################
					if vExposed:
						classification_subnetworks  = {}
						for network in self.topology_networks_bychain[chain]:
							
							# Filter out self.vExposed vertices in the network
							verticesInThisNetwork = []
							for v in vExposed:
								if v.id in network:
									verticesInThisNetwork.append(v.id)

							touched_nodes = {}
							for node in verticesInThisNetwork:

								if node in touched_nodes:
									continue

								classification_subnetwork = {}
								goFoClassification(node, classification_subnetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, verticesInThisNetwork)
								classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
								touched_nodes.update(classification_subnetwork)

						# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
						triangulation = deepcopy(mutatedTscTriangulation)

						# networks = classification_subnetworks
						# for network in networks:
						# 	for node in network:
						# 		networks[network][node] = triangulation[node].s1
						# 		s2Keep = {}
						# 		for s2_key in triangulation[node].s2s:
						# 			if s2_key[0] in network and s2_key[1] in network:
						# 				# Residue names (could include ion names if included)...
						# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
						# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
						# 				s2Keep[s2_key] = None
						# 		s2Keep = list(s2Keep)
						# 		triangulation[node].s2s = s2Keep

						networks = classification_subnetworks
						for network in networks:
							for node in network:
								networks[network][node] = triangulation[node].s1
								s2Keep = {}
								# Network of more than one node
								if triangulation[node].s2s:
									for s2_key in triangulation[node].s2s:
										if s2_key[0] in network and s2_key[1] in network:
											# Residue names (could include ion names if included)...
											rn1 = triangulation[s2_key[0]].s1.data.residue_name 
											rn2 = triangulation[s2_key[1]].s1.data.residue_name
											s2Keep[s2_key] = None
									s2Keep = list(s2Keep)
								# Network of only one node
								else:
									s2Keep = [(node, node)]
								triangulation[node].s2s = s2Keep

						# Format and write the finalized exposed network nodes to PDB file...
						if networks:
							getFinalMinimizedNetworks(networks)
							outPath = self.outPaths_bychain[chain]
							writeNetworkToFile(networks, triangulation, "exposed", outPath, self.pdbCode, chain=chain)

					#
					# Margin amino acid sub-networks...
					############################################################################################
					if vMargin:
						classification_subnetworks  = {}
						for network in self.topology_networks_bychain[chain]:
							
							# Filter out self.vMargin vertices in the network
							verticesInThisNetwork = []
							for v in vMargin:
								if v.id in network:
									verticesInThisNetwork.append(v.id)

							touched_nodes = {}
							for node in verticesInThisNetwork:

								if node in touched_nodes:
									continue

								classification_subnetwork = {}
								goFoClassification(node, classification_subnetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, verticesInThisNetwork)
								classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
								touched_nodes.update(classification_subnetwork)

						# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
						triangulation = deepcopy(mutatedTscTriangulation)

						# networks = classification_subnetworks
						# for network in networks:
						# 	for node in network:
						# 		networks[network][node] = triangulation[node].s1
						# 		s2Keep = {}
						# 		for s2_key in triangulation[node].s2s:
						# 			if s2_key[0] in network and s2_key[1] in network:
						# 				# Residue names (could include ion names if included)...
						# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
						# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
						# 				s2Keep[s2_key] = None
						# 		s2Keep = list(s2Keep)
						# 		triangulation[node].s2s = s2Keep

						networks = classification_subnetworks
						for network in networks:
							for node in network:
								networks[network][node] = triangulation[node].s1
								s2Keep = {}
								# Network of more than one node
								if triangulation[node].s2s:
									for s2_key in triangulation[node].s2s:
										if s2_key[0] in network and s2_key[1] in network:
											# Residue names (could include ion names if included)...
											rn1 = triangulation[s2_key[0]].s1.data.residue_name 
											rn2 = triangulation[s2_key[1]].s1.data.residue_name
											s2Keep[s2_key] = None
									s2Keep = list(s2Keep)
								# Network of only one node
								else:
									s2Keep = [(node, node)]
								triangulation[node].s2s = s2Keep

						# Format and write the finalized margin network nodes to PDB file...
						if networks:
							getFinalMinimizedNetworks(networks)
							outPath = self.outPaths_bychain[chain]
							writeNetworkToFile(networks, triangulation, "margin", outPath, self.pdbCode, chain=chain)

					#
					# Core amino acid sub-networks...
					############################################################################################
					if vCore:
						classification_subnetworks  = {}
						for network in self.topology_networks_bychain[chain]:
							
							# Filter out self.vCore vertices in the network
							verticesInThisNetwork = []
							for v in vCore:
								if v.id in network:
									verticesInThisNetwork.append(v.id)

							touched_nodes = {}
							for node in verticesInThisNetwork:

								if node in touched_nodes:
									continue

								classification_subnetwork = {}
								goFoClassification(node, classification_subnetwork, mutatedTscTriangulation[node].s2s, mutatedTscTriangulation, verticesInThisNetwork)
								classification_subnetworks[tuple(sorted(classification_subnetwork))] = classification_subnetwork
								touched_nodes.update(classification_subnetwork)

						# Copy the minimized topologial network of amino acid side chains and trim for exposed network nodes...
						triangulation = deepcopy(mutatedTscTriangulation)

						# networks = classification_subnetworks
						# for network in networks:
						# 	for node in network:
						# 		networks[network][node] = triangulation[node].s1
						# 		s2Keep = {}
						# 		for s2_key in triangulation[node].s2s:
						# 			if s2_key[0] in network and s2_key[1] in network:
						# 				# Residue names (could include ion names if included)...
						# 				rn1 = triangulation[s2_key[0]].s1.data.residue_name 
						# 				rn2 = triangulation[s2_key[1]].s1.data.residue_name
						# 				s2Keep[s2_key] = None
						# 		s2Keep = list(s2Keep)
						# 		triangulation[node].s2s = s2Keep

						networks = classification_subnetworks
						for network in networks:
							for node in network:
								networks[network][node] = triangulation[node].s1
								s2Keep = {}
								# Network of more than one node
								if triangulation[node].s2s:
									for s2_key in triangulation[node].s2s:
										if s2_key[0] in network and s2_key[1] in network:
											# Residue names (could include ion names if included)...
											rn1 = triangulation[s2_key[0]].s1.data.residue_name 
											rn2 = triangulation[s2_key[1]].s1.data.residue_name
											s2Keep[s2_key] = None
									s2Keep = list(s2Keep)
								# Network of only one node
								else:
									s2Keep = [(node, node)]
								triangulation[node].s2s = s2Keep

						# Format and write the finalized core network nodes to PDB file...
						if networks:
							getFinalMinimizedNetworks(networks)
							outPath = self.outPaths_bychain[chain]
							writeNetworkToFile(networks, triangulation, "core", outPath, self.pdbCode, chain=chain)

	def identifyMissingTscAtoms(self):

		# Quality control: write residues that have missing side chains
		# atoms that could not be included in the exposed, margin, core calculation.
		if self.residuesMissingTscAtoms:

			# Grouped PDB chains...
			if self.group_chains:

				# Write residues with missing side chain atoms.
				###############################################
				if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:
					output = open(self.outPath + self.pdbCode + ".missingIonizable-" + self.chainString + ".pdb", "w")
				elif self.residueSet == "polarSet":
					output = open(self.outPath + self.pdbCode + ".missingPolar-" + self.chainString + ".pdb", "w")
				elif self.residueSet == "apolarSet":
					output = open(self.outPath + self.pdbCode + ".missingApolar-" + self.chainString + ".pdb", "w")
				else:
					output = open(self.outPath + self.pdbCode + ".missing-" + self.chainString + ".pdb", "w")

				for r in self.residuesMissingTscAtoms:
					output.write(str(r))
				output.close()

				print("Done writing residues with missing TSC atoms...")

			# Individual PDB chains...
			else:

				for chain in self.chains:
				
					# Write residues with missing side chain atoms.
					###############################################
					outPath = self.outPaths_bychain[chain]
					if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:
						output = open(outPath + self.pdbCode + ".missingIonizable-" + chain + ".pdb", "w")
					elif self.residueSet == "polarSet":
						output = open(outPath + self.pdbCode + ".missingPolar-" + chain + ".pdb", "w")
					elif self.residueSet == "apolarSet":
						output = open(outPath + self.pdbCode + ".missingApolar-" + chain + ".pdb", "w")
					else:
						output = open(outPath + self.pdbCode + ".missing-" + chain + ".pdb", "w")

					for r in self.residuesMissingTscAtoms:
						output.write(str(r))
					output.close()

				print("Done writing residues with missing TSC atoms for chain...", chain)

		else:
			print("There are no residues with missing TSC atoms...")

	def writeSidechainClassificationResults(self):

		# Grouped PDB chains...
		if self.group_chains:

			# Collect all the classified vertices into one dictionary keyed by atom_key.
			vAll = {}
			for classificationInstance in self.vCore.values():
				vAll[classificationInstance.v.data.atom_key] = classificationInstance
			for classificationInstance in self.vMargin.values():
				vAll[classificationInstance.v.data.atom_key] = classificationInstance
			for classificationInstance in self.vExposed.values():
				vAll[classificationInstance.v.data.atom_key] = classificationInstance

			# Collect the classification strings for writing to file.
			classificationStrings = ""
			for v in vAll:
				classificationStrings +=  vAll[v].classificationString + "\n" 

			# Covert to Microsoft Excel format and write to file.
			classBook = openpyxl.Workbook()
			sheet = classBook["Sheet"]
			classBook.remove(sheet)
			sheet1 = classBook.create_sheet("Sidechain Classification")
			cell = sheet1.cell(row=1, column=1)
			cell.value = "Classification"
			cell = sheet1.cell(row=1, column=2)
			cell.value = "Residue Number"
			cell = sheet1.cell(row=1, column=3)
			cell.value = "Residue Name"
			cell = sheet1.cell(row=1, column=4)
			cell.value = "Residue Chain"
			cell = sheet1.cell(row=1, column=5)
			cell.value = "pHinder Depth"

			# Rank the side chain classifications by side chain depth.
			allStrings = classificationStrings.split("\n")
			rankedClassification = {}
			for x in allStrings:
				splitX = x.split(":")
				fullSplit = []
				for y in splitX:
					fullSplit += y.split()
				if fullSplit:
					key = (float(fullSplit[-1]), fullSplit[0], fullSplit[1], fullSplit[2])
					rankedClassification.update({key:fullSplit})

			row = 2
			rankedClassificationKeys = sorted(rankedClassification)
			for rankedClassificationKey in rankedClassificationKeys:
				fullSplit = rankedClassification[rankedClassificationKey]
				for col in range(len(fullSplit)):
					if col in (1,):
						cell = sheet1.cell(row=row, column=col+1)
						cell.value = int(fullSplit[col])
					elif col in (4,):
						cell = sheet1.cell(row=row, column=col+1)
						cell.value = float(fullSplit[col])
					else:
						cell = sheet1.cell(row=row, column=col+1)
						cell.value = fullSplit[col]
				row += 1

			if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:
				classBook.save(self.outPath + self.pdbCode + ".IonizableSidechainClassification.xlsx")
			elif self.residueSet == "polarSet":
				classBook.save(self.outPath + self.pdbCode + ".PolarSidechainClassification.xlsx")
			elif self.residueSet == "apolarSet":
				classBook.save(self.outPath + self.pdbCode + ".ApolarSidechainClassification.xlsx")
			else:
				classBook.save(self.outPath + self.pdbCode + ".SidechainClassification.xlsx")

		# Individual PDB chains...
		else:

			for chain in self.chains:

				# Collect all the classified vertices into one dictionary keyed by atom_key.
				vAll = {}
				if chain in self.vCore_bychain:
					for classificationInstance in self.vCore_bychain[chain].values():
						vAll[classificationInstance.v.data.atom_key] = classificationInstance
				if chain in self.vMargin_bychain:
					for classificationInstance in self.vMargin_bychain[chain].values():
						vAll[classificationInstance.v.data.atom_key] = classificationInstance
				if chain in self.vExposed_bychain:
					for classificationInstance in self.vExposed_bychain[chain].values():
						vAll[classificationInstance.v.data.atom_key] = classificationInstance

				# Surface and classification could not be calculated
				if not vAll:
					continue

				# Collect the classification strings for writing to file.
				classificationStrings = ""
				for v in vAll:
					classificationStrings +=  vAll[v].classificationString + "\n" 

				# Covert to Microsoft Excel format and write to file.
				classBook = openpyxl.Workbook()
				sheet = classBook["Sheet"]
				classBook.remove(sheet)
				sheet1 = classBook.create_sheet("Sidechain Classification")
				cell = sheet1.cell(row=1, column=1)
				cell.value = "Classification"
				cell = sheet1.cell(row=1, column=2)
				cell.value = "Residue Number"
				cell = sheet1.cell(row=1, column=3)
				cell.value = "Residue Name"
				cell = sheet1.cell(row=1, column=4)
				cell.value = "Residue Chain"
				cell = sheet1.cell(row=1, column=5)
				cell.value = "pHinder Depth"

				# Rank the side chain classifications by side chain depth.
				allStrings = classificationStrings.split("\n")
				rankedClassification = {}
				for x in allStrings:
					splitX = x.split(":")
					fullSplit = []
					for y in splitX:
						fullSplit += y.split()
					if fullSplit:
						key = (float(fullSplit[-1]), fullSplit[0], fullSplit[1], fullSplit[2])
						rankedClassification.update({key:fullSplit})

				row = 2
				rankedClassificationKeys = sorted(rankedClassification)
				for rankedClassificationKey in rankedClassificationKeys:
					fullSplit = rankedClassification[rankedClassificationKey]
					for col in range(len(fullSplit)):
						if col in (1,):
							cell = sheet1.cell(row=row, column=col+1)
							cell.value = int(fullSplit[col])
						elif col in (4,):
							cell = sheet1.cell(row=row, column=col+1)
							cell.value = float(fullSplit[col])
						else:
							cell = sheet1.cell(row=row, column=col+1)
							cell.value = fullSplit[col]
					row += 1

				outPath = self.outPaths_bychain[chain]
				if self.residueSet in ["ionizableSet", "ionizableSetNoCys"]:
					classBook.save(outPath + self.pdbCode + ".IonizableSidechainClassification-" + chain + ".xlsx")
				elif self.residueSet == "polarSet":
					classBook.save(outPath + self.pdbCode + ".PolarSidechainClassification-" + chain + ".xlsx")
				elif self.residueSet == "apolarSet":
					classBook.save(outPath + self.pdbCode + ".ApolarSidechainClassification-" + chain + ".xlsx")
				else:
					classBook.save(outPath + self.pdbCode + ".SidechainClassification-" + chain + ".xlsx")

	def classifyInterfaceSidechains(self, distance_filter=8.0):

		interface = {}
		if not self.group_chains:

			if len(self.chains) > 1:

				i = 0
				while i < len(self.chains) - 1:

					chain_1, v_1s = sorted(self.chains)[i], {}
					for classificationInstance in self.vMargin_bychain[chain_1].values():
						v_1s[classificationInstance.v.data.atom_key] = classificationInstance
					for classificationInstance in self.vExposed_bychain[chain_1].values():
						v_1s[classificationInstance.v.data.atom_key] = classificationInstance

					j = i + 1
					while j < len(self.chains):

						chain_2, v_2s = sorted(self.chains)[j], {}
						for classificationInstance in self.vMargin_bychain[chain_2].values():
							v_2s[classificationInstance.v.data.atom_key] = classificationInstance
						for classificationInstance in self.vExposed_bychain[chain_2].values():
							v_2s[classificationInstance.v.data.atom_key] = classificationInstance

						for v_1_key in v_1s:
							classificationInstance_1 = v_1s[v_1_key]
							for v_2_key in v_2s:
								classificationInstance_2 = v_2s[v_2_key]
								d = distance(classificationInstance_1.v, classificationInstance_2.v)
								if d < distance_filter:
									chain_key = tuple(sorted([chain_1, chain_2]))
									if chain_key not in interface:
										interface[chain_key] = set()
										interface[chain_key].add(classificationInstance_1.v.data.residue)
										interface[chain_key].add(classificationInstance_2.v.data.residue)
									else:
										interface[chain_key].add(classificationInstance_1.v.data.residue)
										interface[chain_key].add(classificationInstance_2.v.data.residue)

								# print(chain_1, chain_2, i, j, len(v_1s), len(v_2s), d)

						j += 1

					i += 1

			else:

				print("\nWarning:")
				print("There must be at least two independent protein change to calculation interface residues.")
				print("No calculation was done.")
				print()

		else:

				print("\nWarning:")
				print("There must be at least two independent protein change to calculation interface residues.")
				print("No calculation was done.")
				print()

		# Write interface files
		for chain_combo in interface:
			f = open(self.outPath + self.pdbCode + ".interface-" + "".join(chain_combo) + ".pdb", "w")
			for contacting_residue in interface[chain_combo]:
				f.write(str(contacting_residue))
			f.close()

	def virtualScreen(self):

		# Report the calculation parameters.
		####################################
		print("\n\n")
		print(80*'-')
		print('pHinder virtual screening calculation parameters.')
		print('Grid increment                    :', self.gridIncrement)
		print('Maximum void network edge distance:', self.maxVoidNetworkEdgeLength)
		print('Minimium void network size        :', self.minVoidNetworkSize)
		print('Clash cutoff (Angstroms)          :', self.virtualClashCutoff)

	def multiprocessInHull(self):

		# Ensure that the number of processes matches the number of available CPUs.
		if self.processes > self.cpuCount:
			self.processes = self.cpuCount

		# Python and System information.
		################################
		print('\nPython version  :', platform.python_version())
		print('Compiler          :', platform.python_compiler())
		print(80*'-')
		print('System            :', platform.system())
		print('Release           :', platform.release())
		print('Machine           :', platform.machine())
		print('Processor         :', platform.processor())
		print('CPU count         :', self.cpuCount)
		print('Requested CPUs    :', self.processes)
		print('Interpreter       :', platform.architecture()[0])
		print('\n')

		# Divide the set of grid vertices by the set of available CPUs.
		###############################################################
		distributedGridVertices = []
		i, j, iDistributionStep, iDistributionCount, m = 0, 0, len(self.gridVertices)/self.processes, len(self.gridVertices)/self.processes, len(self.gridVertices) % self.processes
		if m:
			iDistributionCount = iDistributionStep + m
		while i < len(self.gridVertices) - 1:
			iDistributedGridVertices = []
			while j < iDistributionCount:
				iDistributedGridVertices.append(self.gridVertices[i])
				j += 1
				i += 1
			distributedGridVertices.append(iDistributedGridVertices)
			iDistributionCount += iDistributionStep

		# Parallelize the calculation.
		##############################
		pool = mp.Pool(processes=self.processes)
		results = [pool.apply_async(inHull, args=(vSet, self.hull3D)) for vSet in distributedGridVertices]
		results = [p.get() for p in results]
		pool.close()

		# Merge the results from the parallel calculations.
		###################################################
		result = []
		for r in results:
			result.extend(r)

		# Set the result.
		#################
		self.gridVertices = result

	def multiprocessRemoveClashes(self, gridVertices):

		# Ensure that the number of processes matches the number of available CPUs.
		if self.processes > self.cpuCount:
			self.processes = self.cpuCount

		# Python and System information.
		################################
		print('\nPython version  :', platform.python_version())
		print('Compiler          :', platform.python_compiler())
		print(80*'-')
		print('System            :', platform.system())
		print('Release           :', platform.release())
		print('Machine           :', platform.machine())
		print('Processor         :', platform.processor())
		print('CPU count         :', self.cpuCount)
		print('Requested CPUs    :', self.processes)
		print('Interpreter       :', platform.architecture()[0])
		print('\n')

		# Divide the set of grid vertices by the set of available CPUs.
		###############################################################
		distributedGridVertices = []
		i, j, iDistributionStep, iDistributionCount, m = 0, 0, int(len(gridVertices)/self.processes), len(gridVertices)/self.processes, len(gridVertices) % self.processes
		if m:
			iDistributionCount = iDistributionStep + m
		while i < len(gridVertices) - 1:
			v = gridVertices[i]
			iDistributedGridVertices = []
			while j < iDistributionCount:
				iDistributedGridVertices.append(gridVertices[i])
				j += 1
				i += 1
			distributedGridVertices.append(iDistributedGridVertices)
			iDistributionCount += iDistributionStep

		# Parallelize the calculation.
		##############################
		pool = mp.Pool(processes=self.processes)
		results = [pool.apply_async(removeClashes, args=(vSet, self.allAtoms, self.virtualClashCutoff)) for vSet in distributedGridVertices]
		results = [p.get() for p in results]
		pool.close()

		# Merge the results from the parallel calculations.
		###################################################
		result = []
		for r in results:
			result.extend(r)

		# Because this function is used for several different class-based calculations, 
		# it needs to return a result instead of set a class instance.
		#################
		return result

	# def multiprocessCalculateVoidSurfaceSamplingPoints(self, radius, cut):

	# 	# Ensure that the number of processes matches the number of available CPUs.
	# 	if self.processes > self.cpuCount:
	# 		self.processes = self.cpuCount

	# 	# Python and System information.
	# 	################################
	# 	print('\nPython version  :', platform.python_version())
	# 	print('Compiler          :', platform.python_compiler())
	# 	print(80*'-')
	# 	print('System            :', platform.system())
	# 	print('Release           :', platform.release())
	# 	print('Machine           :', platform.machine())
	# 	print('Processor         :', platform.processor())
	# 	print('CPU count         :', self.cpuCount)
	# 	print('Requested CPUs    :', self.processes)
	# 	print('Interpreter       :', platform.architecture()[0])
	# 	print('\n')

	# 	# Parallelize the calculation.
	# 	##############################
	# 	pool = mp.Pool(processes=self.processes)
	# 	results = [pool.apply_async(calculateVoidSurfaceSamplingPoints, args=(voidSurfaceAtomSet, radius, cut)) for voidSurfaceAtomSet in self.voidSurfaceAtomSets]
	# 	results = [p.get() for p in results]
	# 	pool.close()

	# 	# Merge the results from the parallel calculations.
	# 	###################################################
	# 	result = []
	# 	for r in results:
	# 		result.append(r)

	# 	# Return the result.
	# 	####################
	# 	self.voidSurfaceVertexSets = result

	def multiprocessCalculateVoidSurfaces(self, clashFreeSurfaceVertexSets):

		# Python and System information.
		################################
		print('\nPython version  :', platform.python_version())
		print('Compiler          :', platform.python_compiler())
		print(80*'-')
		print('System            :', platform.system())
		print('Release           :', platform.release())
		print('Machine           :', platform.machine())
		print('Processor         :', platform.processor())
		print('CPU count         :', self.cpuCount)
		print('Requested CPUs    :', self.processes)
		print('Interpreter       :', platform.architecture()[0])
		print('\n')

		# Parallelize the calculation.
		##############################
		pool = mp.Pool(processes=self.processes)
		results = [pool.apply_async(calculateVoidSurfaces, args=(vSet,)) for vSet in clashFreeSurfaceVertexSets]
		results = [p.get() for p in results]
		pool.close()

		# Merge the results from the parallel calculations.
		###################################################
		result = []
		for r in results:
			result.append(r)

		# Return the result.
		####################
		return result

	def multiprocessCalculateVoidSurfaceSamplingShells(self, voidSurfaces):

		# Python and System information.
		################################
		print('\nPython version  :', platform.python_version())
		print('Compiler          :', platform.python_compiler())
		print(80*'-')
		print('System            :', platform.system())
		print('Release           :', platform.release())
		print('Machine           :', platform.machine())
		print('Processor         :', platform.processor())
		print('CPU count         :', self.cpuCount)
		print('Requested CPUs    :', self.processes)
		print('Interpreter       :', platform.architecture()[0])
		print('\n')

		# Parallelize the calculation.
		##############################
		pool = mp.Pool(processes=self.processes)
		i, rankedVoidSurfaces = 1, []
		for voidSurface in voidSurfaces:
			rankedVoidSurfaces.append((i,voidSurface))
			i += 1

		results = [pool.apply_async(calculateVoidSurfaceSamplingShells, args=(rankedVoidSurface, self.allAtomVertexDict, self.outputDirectory, self.pdbCode)) for rankedVoidSurface in rankedVoidSurfaces]
		results = [p.get() for p in results]
		pool.close()

	def makeSamplingGridCubic(self):

		# Sampling option #1 (my preferred method)
		# Use the protein coordinates to generate a grid of sampling points. 
		
		# Get the maximum and minimum X, Y, and Z coordinates.
		######################################################
		minX, minY, minZ, maxX, maxY, maxZ = 100000., 100000., 100000., -100000., -100000., -100000.
		for atom in self.allAtoms:
			
			# Current x, y, and z coordinates.
			##################################
			x, y, z = atom.x, atom.y, atom.z
			
			# X
			####
			if x < minX:
				minX = x
			if x > maxX:
				maxX = x

			# Y
			####
			if y < minY:
				minY = y
			if y > maxY:
				maxY = y
					
			# Z
			####
			if z < minZ:
				minZ = z
			if z > maxZ:
				maxZ = z

		# Calculate the boundaries of the grid cube.
		############################################
		self.gridVertices, self.gridAtoms = [], []
		x, y, z, i = minX, minY, minZ, 1
		while x <= maxX + self.gridIncrement:
			while y <= maxY + self.gridIncrement:
				while z <= maxZ + self.gridIncrement:
					newSphere = Sphere()
					newSphere.origin_x = x
					newSphere.origin_y = y
					newSphere.origin_z = z
					u = x**2 + y**2 + z**2
					vertex = Vertex4D((x,y,z,u),setC=1)
					vertex.id = i
					vertex.data = newSphere.get_origin_psa()
					vertex.data.v = vertex # Overwrite the original Vertex() object with new Vertex4D() object.
					self.gridVertices.append(vertex)
					self.gridAtoms.append(newSphere)
					i += 1
					z += self.gridIncrement
				z = minZ
				y += self.gridIncrement
			y = minY
			x += self.gridIncrement

		# Write the grid to file.
		#########################
		output = gzip.open(self.outputDirectory + self.pdbCode + ".cubeGrid.pdb.gz","wt")
		for gridAtom in self.gridAtoms:
			output.write(str(gridAtom.get_origin_psa()))
		output.close()

		print("Done calculating virtual sampling grid...")

	def makeSamplingGridUsingProteinSurface(self):

		# Sampling option #2 (my preferred method)
		# Use the protein surface to generate the virtual sampling grid.

		# XXXX Options for swelling?

		if self.group_chains:

			self.gridVertices = generateVirtualScreeningVertices(self.quadSurface, inIterations=self.inIterations, inIterationsStepSize=self.inIterationsStepSize, outIterations=self.outIterations, outIterationsStepSize=self.outIterationsStepSize)
			self.gridAtoms = []
			for gridVertex in self.gridVertices:
				self.gridAtoms.append(gridVertex.data)

			output = gzip.open(self.outPath + self.pdbCode + ".grid-" + self.chainString + ".pdb.gz", "wt")
			for gridAtom in self.gridAtoms:
				output.write(str(gridAtom))
			output.close()

		else:

			for chain in self.quadSurface_bychain:

				quadSurface_chain = self.quadSurface_bychain[chain]
				self.gridVertices = generateVirtualScreeningVertices(quadSurface_chain, inIterations=self.inIterations, inIterationsStepSize=self.inIterationsStepSize, outIterations=self.outIterations, outIterationsStepSize=self.outIterationsStepSize)
				self.gridVertices_bychain[chain] = self.gridVertices
				self.gridAtoms_bychain[chain] = []
				for gridVertex in self.gridVertices:
					self.gridAtoms_bychain[chain].append(gridVertex.data)

				output = gzip.open(self.outPath + chain + sep + self.pdbCode + ".grid-" + chain + ".pdb.gz", "wt")
				for gridAtom in self.gridAtoms_bychain[chain]:
					output.write(str(gridAtom))
				output.close()

		# # Grouped PDB chains...
		# if self.group_chains:

		# 	# Write the grid to file.
		# 	#########################
		# 	output = gzip.open(self.outPath + self.pdbCode + ".grid.pdb.gz", "wt")
		# 	for gridAtom in self.gridAtoms:
		# 		output.write(str(gridAtom))
		# 	output.close()

		# # Individual PDB chains...
		# else:

		# 	for chain in self.chains:

		# 		outPath = self.outPaths_bychain[chain]


		# 		outFile = open(outPath + self.pdbCode + ".triangulation-" + chain + ".py", "w")
		# 		outFile.write(script_body)
		# 		outFile.close()

		# # Write the grid to file.
		# #########################
		# # output = gzip.open(self.outputDirectory + self.pdbCode + ".grid.pdb.gz", "wt")
		# output = gzip.open(self.outPath + self.pdbCode + ".grid.pdb.gz", "wt")
		# for gridAtom in self.gridAtoms:
		# 	output.write(str(gridAtom))
		# output.close()

	# def makeSamplingGridUsingProteinSurfaceCreviceFacets(self):

	# 	# Sampling option #3 
	# 	# Use the protein surface to generate the virtual sampling grid from only crevice facets.

	# 	# XXXX Options for swelling?

	# 	self.gridVertices = generateVirtualScreeningVertices(self.quadSurfaceCrevice, inIterations=self.inIterations, inIterationsStepSize=self.inIterationsStepSize, outIterations=self.outIterations, outIterationsStepSize=self.outIterationsStepSize)
	# 	self.gridAtoms = []
	# 	for gridVertex in self.gridVertices:
	# 		self.gridAtoms.append(gridVertex.data)

	# 	# Write the grid to file.
	# 	#########################
	# 	output = gzip.open(self.outputDirectory + self.pdbCode + ".grid.crevices.pdb.gz", "wt")
	# 	for gridAtom in self.gridAtoms:
	# 		output.write(str(gridAtom))
	# 	output.close()

	def filterSamplingPointsUsingConvexHull3D(self):

		print("Calculating the 3D convex hull of the protein...")
		self.hull3D = convexHull3D(self.caVertices)

		print("\nStart -- Parallel call to calculate multiprocessInHull.")
		self.multiprocessInHull()
		print("\nEnd -- Parallel call to calculate multiprocessInHull."	)

	def filterSamplingPointsUsingClashes(self):

		# Round 1 of clash determination.
		# Although costly, this speeds things up by avoiding many unecessary volume calculations.
		#########################################################################################
		print("\nStart -- Parallel call to calculate multiprocessRemoveClashes.")

		if self.group_chains:

			self.gridVerticesNoClash = self.multiprocessRemoveClashes(self.gridVertices)

			# Provide residue numbers and atom serial numbers.
			##################################################
			i = 1
			for v in self.gridVerticesNoClash:
				v.data.atom_serial = i
				v.data.residue_sequence_number = i
				v.data.reinitialize()
				i += 1

			# Write the remaining grid points to file.
			##########################################
			# output = gzip.open(self.outputDirectory + self.pdbCode + ".gridFinal.pdb.gz", "wt")
			output = gzip.open(self.outPath + self.pdbCode + ".gridFinal-" + self.chainString + ".pdb.gz", "wt")
			for v in self.gridVerticesNoClash:
				output.write(str(v.data))
			output.close()

		else:

			for chain in self.gridVertices_bychain:

				gridVertices = self.gridVertices_bychain[chain]

				self.gridVerticesNoClash_bychain[chain] = self.multiprocessRemoveClashes(gridVertices)

				# Provide residue numbers and atom serial numbers.
				##################################################
				i = 1
				for v in self.gridVerticesNoClash_bychain[chain]:
					v.data.atom_serial = i
					v.data.residue_sequence_number = i
					v.data.reinitialize()
					i += 1

				# Write the remaining grid points to file.
				##########################################
				# output = gzip.open(self.outputDirectory + self.pdbCode + ".gridFinal.pdb.gz", "wt")
				output = gzip.open(self.outPath + chain + sep + self.pdbCode + ".gridFinal-" + chain + ".pdb.gz", "wt")
				for v in self.gridVerticesNoClash_bychain[chain]:
					output.write(str(v.data))
				output.close()

		print("\nDone -- Parallel call to calculate multiprocessRemoveClashes.")

	# def filterSamplingPointsUsingProximity(self, proximityLimit=1.0):

	# 	print("\nStart -- Call to calculate filterSamplingPointsUsingProximity.")

	# 	gridVertices = {}
	# 	for v in self.gridVerticesNoClash:
	# 		gridVertices[(v.x, v.y, v.z)] = v

	# 	proximityViolators = 1
	# 	while proximityViolators:

	# 		proximityViolators, proximityCondensed = {}, {}
	# 		for v1 in gridVertices:

	# 			# Convert from key to vertex.
	# 			#############################
	# 			v1 = gridVertices[v1]

	# 			breakFlag = 0
	# 			for v2 in gridVertices:

	# 				# Convert from key to vertex.
	# 				#############################
	# 				v2 = gridVertices[v2]

	# 				if v1 == v2:
	# 					continue

	# 				# Condense two grid points to one using the proximity limit.
	# 				############################################################
	# 				if distance(v1, v2) < proximityLimit:
	# 					comVertex = centroid4D((v1, v2), PseudoAtom())
	# 					proximityCondensed[(comVertex.x, comVertex.y, comVertex.z)] = comVertex
	# 					proximityViolators[v1], proximityViolators[v2] = None, None
	# 					breakFlag = 1
	# 					break

	# 			if breakFlag:
	# 				break

	# 		for proximityViolator in proximityViolators:
	# 			del gridVertices[(proximityViolator.x, proximityViolator.y, proximityViolator.z)]
	# 		gridVertices.update(proximityCondensed)

	# 	self.gridVerticesNoClash = gridVertices.values()

	# 	print("\nEnd -- Call to calculate filterSamplingPointsUsingProximity.")

	def triangulateRemainingGridPoints(self):

		print("triangulateRemainingGridPoints.........")

		if self.group_chains:

			# Triangulate the remaining grid points that do not clash with the protein.
			###########################################################################
			print("Start -- Calling serial triangulateRemainingGridPoints...", len(self.gridVerticesNoClash))
			gridPointMesh = convexHull4D(self.gridVerticesNoClash, walkDistanceLimit=0.25)
			gridPointTriangulation = gridPointMesh.s1Dict
			print("End -- Calling serial triangulateRemainingGridPoints...")

			self.gridPointTriangulation = gridPointTriangulation
			self.gridPointMesh = gridPointMesh

		else:

			for chain in self.gridVerticesNoClash_bychain:

				gridVerticesNoClash = self.gridVerticesNoClash_bychain[chain]

				# Triangulate the remaining grid points that do not clash with the protein.
				###########################################################################
				print("Start -- Calling serial triangulateRemainingGridPoints...", len(gridVerticesNoClash))
				gridPointMesh = convexHull4D(gridVerticesNoClash, walkDistanceLimit=0.25)
				gridPointTriangulation = gridPointMesh.s1Dict
				print("End -- Calling serial triangulateRemainingGridPoints...", chain)

				self.gridPointTriangulation_bychain[chain] = gridPointTriangulation
				self.gridPointMesh_bychain[chain] = gridPointMesh

	# def samplingReduction(self, minVoidNetworkEdgeLength=0.5):

	# 	reductionLimit = minVoidNetworkEdgeLength

	# 	print("Start -- samplingReduction...")

	# 	# Rank the triangulation nodes by node order...
	# 	rankedNodes = {}
	# 	for node in self.gridPointTriangulation:
	# 		l = len(self.gridPointTriangulation[node].s2s)
	# 		rankedNodes[(l, node)] = None

	# 	# Identify redundant nodes using the reductionLimit...
	# 	keys = sorted(rankedNodes)
	# 	keys.reverse()
	# 	markForDeletion = {}
	# 	for key in keys:
	# 		node = key[1]
	# 		if node in markForDeletion:
	# 			continue
	# 		v1 = self.gridPointTriangulation[node].s1
	# 		for s2 in self.gridPointTriangulation[node].s2s:
	# 			for v2id in s2:
	# 				if v2id != node:
	# 					v2 = self.gridPointTriangulation[v2id].s1
	# 					d = distance(v1, v2)
	# 					if d < reductionLimit:
	# 						markForDeletion[v2id] = None
	# 	# Remove redundant nodes...
	# 	vReduced = []
	# 	for key in keys:
	# 		node = key[1]
	# 		if node not in markForDeletion:
	# 			vReduced.append(self.gridPointTriangulation[node].s1)
	# 	self.gridVerticesNoClash = vReduced

	# 	# Triangulate the refined set of reduced nodes...
	# 	print("Start -- Serial triangulation of reduced nodes...", len(self.gridVerticesNoClash))
	# 	gridPointMesh = convexHull4D(self.gridVerticesNoClash, walkDistanceLimit=0.25)
	# 	gridPointTriangulation = gridPointMesh.s1Dict
	# 	print("End -- Serial triangulation of reduced nodes...")

	# 	self.gridPointTriangulation = gridPointTriangulation
	# 	self.gridPointMesh = gridPointMesh

	def identifyAndParseIndividualSamplingVoids(self, maxVoidNetworkEdgeLength=0, minVoidNetworkEdgeLength=0, minVoidNetworkSize=1, psa=0):#, reduced=0, pruneDistance=0):

		if not maxVoidNetworkEdgeLength:
			maxVoidNetworkEdgeLength = self.maxVoidNetworkEdgeLength

		# Use pruneTriangulationGoFo to identify sub-networks within the void sampling network

		print("Start -- identifyAndParseIndividualSamplingVoids...",)

		# carbon-carbon bond lengths
		# single : 1.54 Angstroms
		# double : 1.33 Angstroms
		# triple : 1.20 Angstroms
		#########################

		if self.group_chains:

			# Prune the global triangulation of void sampling points.
			# This function includes a goFo step that splits the global triangulation into sub-networks.
			############################################################################################
			prunedVoidNetworkTriangulation = deepcopy(self.gridPointTriangulation) # Key feature.
			prunedVoidNetworkTriangulationResult = pruneTriangulationGoFo(prunedVoidNetworkTriangulation, maxVoidNetworkEdgeLength, minNetworkEdgeLength=minVoidNetworkEdgeLength, psa=psa)
			self.prunedVoidNetworkTriangulation = prunedVoidNetworkTriangulationResult[0]
			# self.prunedVoidNetworks = prunedVoidNetworkTriangulationResult[2]
			self.prunedVoidNetworks = prunedVoidNetworkTriangulationResult[1]

			# Iteratively minimize the pruned triangulation.
			# Remove redundant network connections by biasing shorter network edges.
			########################################################################
			minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation)
			self.voidNetworks = getFinalMinimizedNetworks(self.prunedVoidNetworks, minNetworkSize=minVoidNetworkSize)
			if self.voidNetworks:
				writeNetworkToFile(self.voidNetworks, self.prunedVoidNetworkTriangulation, "chemiForms", self.outPath, self.pdbCode, include_residues=0, chain=self.chainString)
		
		else:

			for chain in self.gridPointTriangulation_bychain:

				gridPointTriangulation = self.gridPointTriangulation_bychain[chain]

				# Prune the global triangulation of void sampling points.
				# This function includes a goFo step that splits the global triangulation into sub-networks.
				############################################################################################
				prunedVoidNetworkTriangulation = deepcopy(gridPointTriangulation) # Key feature.
				prunedVoidNetworkTriangulationResult = pruneTriangulationGoFo(prunedVoidNetworkTriangulation, maxVoidNetworkEdgeLength, minNetworkEdgeLength=minVoidNetworkEdgeLength, psa=psa)
				self.prunedVoidNetworkTriangulation_bychain[chain] = prunedVoidNetworkTriangulationResult[0]
				# self.prunedVoidNetworks = prunedVoidNetworkTriangulationResult[2]
				self.prunedVoidNetworks_bychain[chain] = prunedVoidNetworkTriangulationResult[1]

				# Iteratively minimize the pruned triangulation.
				# Remove redundant network connections by biasing shorter network edges.
				########################################################################
				minimizeNetworks(self.prunedVoidNetworks_bychain[chain], self.prunedVoidNetworkTriangulation_bychain[chain])
				self.voidNetworks_bychain[chain] = getFinalMinimizedNetworks(self.prunedVoidNetworks_bychain[chain], minNetworkSize=minVoidNetworkSize)
				if self.voidNetworks_bychain[chain]:
					writeNetworkToFile(self.voidNetworks_bychain[chain], self.prunedVoidNetworkTriangulation_bychain[chain], "chemiForms", self.outPath + chain + sep, self.pdbCode, include_residues=0, chain=chain)



		# # In case it is need by later calculations....
		# self.voidNetworks = self.prunedVoidNetworks

		# i = 1
		# previousResult = None
		# # XXXX minimizeNetworks need to update
		# minimizing = minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
		# while minimizing:
		# 	minimizing = minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
		# 	previousResult = minimizing
		# 	i += 1
		# self.voidNetworks = self.prunedVoidNetworks

		# # Arrange networks from largest to smallest. All prune using self.minNetworkSize...
		# self.rankedVoidNetworks = {}
		# for voidNetwork in self.voidNetworks:
		# 	if len(voidNetwork) > self.minNetworkSize:
		# 		self.rankedVoidNetworks[(len(voidNetwork), voidNetwork)] = voidNetwork
		# rankedVoidNetworksKeys = sorted(self.rankedVoidNetworks)
		# rankedVoidNetworksKeys.reverse()

		# # if not reduced:

		# # Create a directory for the individual chemiform network edge files.
		# #####################################################################
		# if not exists(self.outputDirectory + "chemiformNetworks"):
		# 	mkdir(self.outputDirectory + "chemiformNetworks")
		# else:
		# 	rmtree(self.outputDirectory + "chemiformNetworks")
		# 	mkdir(self.outputDirectory + "chemiformNetworks")

		# # Create a directory for the individual chemiform PDB files.
		# ############################################################
		# if not exists(self.outputDirectory + "chemiformNodes"):
		# 	mkdir(self.outputDirectory + "chemiformNodes")
		# else:
		# 	rmtree(self.outputDirectory + "chemiformNodes")
		# 	mkdir(self.outputDirectory + "chemiformNodes")

		# # else:

		# # 	# Create a directory for the individual chemiform network edge files.
		# # 	#####################################################################
		# # 	if not exists(self.outputDirectory + "chemiformNetworksReduced"):
		# # 		mkdir(self.outputDirectory + "chemiformNetworksReduced")
		# # 	else:
		# # 		rmtree(self.outputDirectory + "chemiformNetworksReduced")
		# # 		mkdir(self.outputDirectory + "chemiformNetworksReduced")

		# # 	# Create a directory for the individual chemiform PDB files.
		# # 	############################################################
		# # 	if not exists(self.outputDirectory + "chemiformNodesReduced"):
		# # 		mkdir(self.outputDirectory + "chemiformNodesReduced")
		# # 	else:
		# # 		rmtree(self.outputDirectory + "chemiformNodesReduced")
		# # 		mkdir(self.outputDirectory + "chemiformNodesReduced")

		# # # Unique repository for aggregating grid points.
		# # ################################################
		# # uniqueChemiformPoints = {}
		# # uniqueNetworkEdgeStrings = {}

		# # Prepare each network to be written to file.
		# #############################################
		# # Void networks first...
		# ########################
		# j, networkName = 1, "networkEdges"
		# #networkString, resStrings, pdbString = "", "", ""
		# for rankedVoidNetworksKey in rankedVoidNetworksKeys:

		# 	networkString, resStrings, pdbString = "", "", ""

		# 	# Unique repository for aggregating grid points.
		# 	################################################
		# 	uniqueChemiformPoints = {}
		# 	uniqueNetworkEdgeStrings = {}

		# 	voidNetwork = self.rankedVoidNetworks[rankedVoidNetworksKey]
		# 	triangulation = self.voidNetworks[voidNetwork]

		# 	# Indicate a new network stack in the file output.
		# 	##################################################
		# 	pdbString = "# Network " + str(j) + " Number of Nodes: " + str(len(voidNetwork)) + "\n"
		# 	networkString += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"
		# 	resStrings += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"

		# 	# Original version...
		# 	for node in voidNetwork:

		# 		pruneS2s = {}
		# 		for s2 in self.prunedVoidNetworkTriangulation[node].s2s:

		# 			if s2[0] not in voidNetwork or s2[1] not in voidNetwork:
		# 				continue

		# 			# Network edge vertices.
		# 			########################
		# 			v1 = self.prunedVoidNetworkTriangulation[s2[0]].s1
		# 			v2 = self.prunedVoidNetworkTriangulation[s2[1]].s1
		# 			uniqueChemiformPoints[(v1.data.atom_serial, v1)] = v1
		# 			uniqueChemiformPoints[(v2.data.atom_serial, v2)] = v2

		# 			# Number each grid point atoms by its network.
		# 			##############################################
		# 			v1.data.residue_sequence_number, v2.data.residue_sequence_number = j, j

		# 			# Format coordinates for writing (a bit redundant: could go with one).
		# 			######################################################################
		# 			format = "%15.6f%15.6f%15.6f "
		# 			resCoords = format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
		# 			networkString += format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
		# 			resStringFormat = "%s-%i-%s"
		# 			resString = "%13s %13s | " % (resStringFormat % (v1.data.residue.chn, v1.data.residue.num, v1.data.residue.name),
		# 											resStringFormat % (v2.data.residue.chn, v2.data.residue.num, v2.data.residue.name))
		# 			if resString not in uniqueNetworkEdgeStrings:
		# 				uniqueNetworkEdgeStrings[resString] = resString + resCoords
		# 			resStrings += resString + resCoords

		# 	if uniqueChemiformPoints:

		# 		uniqueGridPointKeys = sorted(uniqueChemiformPoints)
		# 		conectString = ""
		# 		for uniqueGridPointKey in uniqueGridPointKeys:
		# 			v = uniqueGridPointKey[1]
		# 			conectString += "%-6s%5i" % ("CONECT", v.data.atom_serial)
		# 			for s2 in self.prunedVoidNetworkTriangulation[v.id].s2s:
		# 				for vid in s2:
		# 					if vid != v.id:
		# 						vConect = self.prunedVoidNetworkTriangulation[vid]
		# 						conectString += "%5i" % vConect.s1.data.atom_serial
		# 			conectString += "\n"

		# 		uniqueGridPointKeys = sorted(uniqueChemiformPoints)
		# 		pdbString = "# Network " + str(j) + " Size: " + str(len(uniqueGridPointKeys)) + "\n"
		# 		resNum = 1
		# 		for uniqueGridPointKey in uniqueGridPointKeys:
		# 			atom = uniqueChemiformPoints[uniqueGridPointKey].data
		# 			atom.record_name = "HETATM"
		# 			atom.atom_name = " C"
		# 			atom.residue_sequence_number = resNum
		# 			atom.element_symbol = "C"
		# 			pdbString += str(uniqueChemiformPoints[uniqueGridPointKey].data)
		# 			resNum += 1

		# 		output = open(self.outputDirectory + "chemiformNodes/" + self.pdbCode + ".nodes." + str(j) + ".pdb", "w")
		# 		output.write(pdbString)
		# 		output.write(conectString)
		# 		output.close()

		# 	if uniqueNetworkEdgeStrings:

		# 		uniqueResStringKeys = sorted(uniqueNetworkEdgeStrings)
		# 		resString = "# Network " + str(j) + " Size: " + str(len(uniqueNetworkEdgeStrings)) + "\n"
		# 		for uniqueResStringKey in uniqueResStringKeys:
		# 			resString += str(uniqueNetworkEdgeStrings[uniqueResStringKey])

		# 		output = gzip.open(self.outputDirectory + "chemiformNetworks/" + self.pdbCode + ".network." + str(j) + ".txt.gz", "wt")
		# 		output.write(resString)
		# 		output.close()

		# 	# Indicate the end of a network stack with a blank line.
		# 	########################################################
		# 	networkString += "\n"

		# 	j += 1

	# def reduceChemiformsIteration(self, voidNetworks, prunedVoidNetworkTriangulation, psa=0):

	# 	# Arrange networks from largest to smallest...
	# 	# Maybe not necessary...
	# 	rankedVoidNetworks = {}
	# 	for voidNetwork in voidNetworks:
	# 		rankedVoidNetworks[(len(voidNetwork), voidNetwork)] = voidNetwork
	# 	rankedVoidNetworksKeys = sorted(rankedVoidNetworks)
	# 	rankedVoidNetworksKeys.reverse()
		
	# 	# Collect all vertices from all of the void networks...
	# 	removeChemiformVertices, uniqueChemiformVertices, reducedVoidNetworkNodes = {}, {}, {}
	# 	for rankedVoidNetworksKey in rankedVoidNetworksKeys:

	# 		voidNetwork = rankedVoidNetworks[rankedVoidNetworksKey]

	# 		# Original version...
	# 		for node in voidNetwork:

	# 			for s2 in prunedVoidNetworkTriangulation[node].s2s:

	# 				# Network edge vertices...
	# 				v1 = prunedVoidNetworkTriangulation[s2[0]].s1
	# 				v2 = prunedVoidNetworkTriangulation[s2[1]].s1
	# 				# Save all network edge vertices...
	# 				uniqueChemiformVertices[(v1.data.atom_serial, v1)] = v1
	# 				uniqueChemiformVertices[(v2.data.atom_serial, v2)] = v2
	# 				# Flag those network edge vertices that are "too close to one another"...
	# 				d = distance(v1, v2)
	# 				if d < 1.1:
	# 					removeChemiformVertices[(v1.data.atom_serial, v1)] = v1
	# 					break

	# 	# Proceed with removing network edge vertices that are "too close to one another"...
	# 	# Vertex distancing...
	# 	print("Size of unique vertex list before removal...", len(uniqueChemiformVertices))
	# 	for removeChemiformVertex in removeChemiformVertices:
	# 		del uniqueChemiformVertices[removeChemiformVertex]
	# 	print("Size of unique vertex list after removal...", len(uniqueChemiformVertices))

	# 	# Triangulate the reduced set of unique chemiform vertices...
	# 	reducedVoidNetworkMesh = convexHull4D(uniqueChemiformVertices.values())
	# 	reducedVoidNetworkTriangulation = reducedVoidNetworkMesh.s1Dict
	# 	prunedReducedVoidNetworkTriangulationResult = pruneTriangulationGoFo(reducedVoidNetworkTriangulation, 1.5, psa=psa)#self.maxVoidNetworkEdgeLength)
	# 	prunedReducedVoidNetworkTriangulation = prunedReducedVoidNetworkTriangulationResult[0]
	# 	prunedReducedVoidNetworks = prunedReducedVoidNetworkTriangulationResult[2]

	# 	# Iteratively minimize the pruned triangulation.
	# 	# Remove redundant network connections by biasing shorter network edges....
	# 	minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation)
	# 	reducedVoidNetworks = getFinalMinimizedNetworks(prunedReducedVoidNetworks)

	# 	# i = 1
	# 	# previousResult = None
	# 	# # XXXX minimizeNetworks need to update
	# 	# minimizing = minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# while minimizing:
	# 	# 	minimizing = minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# 	previousResult = minimizing
	# 	# 	i += 1
	# 	# reducedVoidNetworks = prunedReducedVoidNetworks

	# 	return (reducedVoidNetworks, prunedReducedVoidNetworkTriangulation, removeChemiformVertices)


	# def reduceChemiforms(self, psa=0):

	# 	# Create a directory for the individual chemiform network edge files.
	# 	#####################################################################
	# 	if not exists(self.outputDirectory + "chemiformNetworksReduced"):
	# 		mkdir(self.outputDirectory + "chemiformNetworksReduced")
	# 	else:
	# 		rmtree(self.outputDirectory + "chemiformNetworksReduced")
	# 		mkdir(self.outputDirectory + "chemiformNetworksReduced")

	# 	# Create a directory for the individual chemiform PDB files.
	# 	############################################################
	# 	if not exists(self.outputDirectory + "chemiformNodesReduced"):
	# 		mkdir(self.outputDirectory + "chemiformNodesReduced")
	# 	else:
	# 		rmtree(self.outputDirectory + "chemiformNodesReduced")
	# 		mkdir(self.outputDirectory + "chemiformNodesReduced")

	# 	removeChemiformVertices, rcvIteraction = 1, 1
	# 	reducedVoidNetworks, prunedReducedVoidNetworkTriangulation = self.voidNetworks, self.prunedVoidNetworkTriangulation
	# 	while removeChemiformVertices:
	# 		print("Remove chemiform vertices...iteration...", rcvIteraction)
	# 		reduceIteration = self.reduceChemiformsIteration(reducedVoidNetworks, prunedReducedVoidNetworkTriangulation, psa=psa)
	# 		reducedVoidNetworks = reduceIteration[0]
	# 		prunedReducedVoidNetworkTriangulation = reduceIteration[1]
	# 		removeChemiformVertices = reduceIteration[2]
	# 		rcvIteraction += 1

	# 	# Arrange networks from largest to smallest.
	# 	############################################
	# 	rankedVoidNetworks = {}
	# 	for voidNetwork in reducedVoidNetworks:
	# 		rankedVoidNetworks[(len(voidNetwork), voidNetwork)] = voidNetwork
	# 	rankedVoidNetworksKeys = sorted(rankedVoidNetworks)
	# 	rankedVoidNetworksKeys.reverse()

	# 	###################################################################
	# 	# # Only for testing..........XXXX
	# 	# testDirectory = "/Users/danisom/Desktop/chemiTest/"
	# 	# if not exists(testDirectory):
	# 	# 	mkdir(testDirectory)
	# 	# else:
	# 	# 	rmtree(testDirectory) 
	# 	# 	mkdir(testDirectory)
	# 	# # Only for testing..........XXXX
	# 	###################################################################

	# 	# Prepare each network to be written to file.
	# 	#############################################
	# 	# Void networks first...
	# 	########################
	# 	j, networkName = 1, "networkEdges"
	# 	for rankedVoidNetworksKey in rankedVoidNetworksKeys:

	# 		networkString, resStrings, pdbString = "", "", ""

	# 		# Unique repository for aggregating grid points.
	# 		################################################
	# 		uniqueChemiformPoints = {}
	# 		uniqueNetworkEdgeStrings = {}

	# 		voidNetwork = rankedVoidNetworks[rankedVoidNetworksKey]

	# 		if len(voidNetwork) < 3:
	# 			continue

	# 		###################################################################
	# 		# # # Only for testing..........XXXX

	# 		# # Start the goFo collection from the highest degree network node...
	# 		# # Starting from a node of highest degree may result in unexpected goFo leveling...
	# 		# # I think this unexpected leveling occurs when goFo starts in a cycle loop...
	# 		# nStart, nDegree = None, 0
	# 		# byDegree, nodes = {}, []
	# 		# for node in voidNetwork:
	# 		# 	iDegree = len(prunedReducedVoidNetworkTriangulation[node].s2s)
	# 		# 	byDegree[(iDegree,node)] = node
	# 		# 	nodes.append(prunedReducedVoidNetworkTriangulation[node].s1)
	# 		# psa = PseudoAtom()
	# 		# nodeCentroid = centroid3D(nodes, psa)
	# 		# byDegreeKeys = sorted(byDegree.keys())
	# 		# lowestDegree = byDegreeKeys[0][0]
	# 		# nStart, dCom, nDegree = None, 0., 0
	# 		# for byDegreeKey in byDegreeKeys:
	# 		# 	if byDegreeKey[0] == lowestDegree:
	# 		# 		d = distance(nodeCentroid, prunedReducedVoidNetworkTriangulation[byDegreeKey[1]].s1)
	# 		# 		if d > dCom:
	# 		# 			nStart = byDegreeKey[1]
	# 		# 			dCom = d
	# 		# 			nDegree = byDegreeKey[0]

	# 		# print("\n\n start chemiform by level..."	)
	# 		# print("Node", nStart, "Degree", nDegree)

	# 		# #nStart = voidNetwork[0] # This selection might be key...
	# 		# been = {}
	# 		# s2s = prunedReducedVoidNetworkTriangulation[nStart].s2s
	# 		# s1Dict = prunedReducedVoidNetworkTriangulation
	# 		# goFo(nStart, been, s2s, s1Dict, level=1, byEdge=1)
	# 		# beenKeys = sorted(been.keys())
	# 		# byLevel = {}
	# 		# for beenKey in beenKeys:
	# 		# 	#print(beenKey, been[beenKey])
	# 		# 	level = been[beenKey]
	# 		# 	print("*", level, beenKey)
	# 		# 	if level not in byLevel:
	# 		# 		byLevel[level] = [beenKey]
	# 		# 	else:
	# 		# 		byLevel[level].append(beenKey)
	# 		# # for level in sorted(byLevel.keys()):
	# 		# # 	print(level, byLevel[level])

	# 		# testDirectory = "/Users/danisom/Desktop/chemiTest/"
	# 		# fileName = str(nStart) + ".pickle"
	# 		# f = open(testDirectory + fileName, "w")
	# 		# pickle.dump(been, f)
	# 		# f.close()

	# 		# # from graphs import Graph
	# 		# # levelGraph = Graph()
	# 		# # levelGraph.levels = byLevel
	# 		# # levelGraph.makeGraph()

	# 		# # # Only for testing..........XXXX
	# 		# ###################################################################

	# 		# Indicate a new network stack in the file output.
	# 		##################################################
	# 		pdbString = "# Network " + str(j) + " Number of Nodes: " + str(len(voidNetwork)) + "\n"
	# 		networkString += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"
	# 		resStrings += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"

	# 		# Original version...
	# 		for node in voidNetwork:

	# 			for s2 in prunedReducedVoidNetworkTriangulation[node].s2s:

	# 				# Network edge vertices.
	# 				########################
	# 				v1 = prunedReducedVoidNetworkTriangulation[s2[0]].s1
	# 				v2 = prunedReducedVoidNetworkTriangulation[s2[1]].s1
	# 				uniqueChemiformPoints[(v1.data.atom_serial, v1)] = v1
	# 				uniqueChemiformPoints[(v2.data.atom_serial, v2)] = v2

	# 				# Number each grid point atoms by its network.
	# 				##############################################
	# 				v1.data.residue_sequence_number, v2.data.residue_sequence_number = j, j

	# 				# Format coordinates for writing (a bit redundant: could go with one).
	# 				######################################################################
	# 				format = "%15.6f%15.6f%15.6f "
	# 				resCoords = format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
	# 				networkString += format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
	# 				resStringFormat = "%s-%i-%s"
	# 				resString = "%13s %13s | " % (resStringFormat % (v1.data.residue.chn, v1.data.residue.num, v1.data.residue.name),
	# 												resStringFormat % (v2.data.residue.chn, v2.data.residue.num, v2.data.residue.name))
	# 				if resString not in uniqueNetworkEdgeStrings:
	# 					uniqueNetworkEdgeStrings[resString] = resString + resCoords
	# 				resStrings += resString + resCoords
	# 		if uniqueChemiformPoints:

	# 			uniqueGridPointKeys = sorted(uniqueChemiformPoints)
	# 			conectString = ""
	# 			for uniqueGridPointKey in uniqueGridPointKeys:
	# 				v = uniqueGridPointKey[1]
	# 				conectString += "%-6s%5i" % ("CONECT", v.data.atom_serial)
	# 				for s2 in prunedReducedVoidNetworkTriangulation[v.id].s2s:
	# 					for vid in s2:
	# 						if vid != v.id:
	# 							vConect = prunedReducedVoidNetworkTriangulation[vid]
	# 							conectString += "%5i" % vConect.s1.data.atom_serial
	# 				conectString += "\n"


	# 			uniqueGridPointKeys = sorted(uniqueChemiformPoints)
	# 			pdbString = "# Network " + str(j) + " Size: " + str(len(uniqueGridPointKeys)) + "\n"
	# 			resNum = 1
	# 			for uniqueGridPointKey in uniqueGridPointKeys:
	# 				atom = uniqueChemiformPoints[uniqueGridPointKey].data
	# 				atom.record_name = "HETATM"
	# 				atom.atom_name = " C"
	# 				atom.residue_sequence_number = resNum
	# 				atom.element_symbol = "C"
	# 				pdbString += str(uniqueChemiformPoints[uniqueGridPointKey].data)
	# 				resNum += 1

	# 			output = open(self.outputDirectory + "chemiformNodesReduced/" + self.pdbCode + ".nodes." + str(j) + ".pdb", "w")
	# 			output.write(pdbString)
	# 			output.write(conectString)
	# 			output.close()

	# 		if uniqueNetworkEdgeStrings:

	# 			uniqueResStringKeys = sorted(uniqueNetworkEdgeStrings)
	# 			resString = "# Network " + str(j) + " Size: " + str(len(uniqueNetworkEdgeStrings)) + "\n"
	# 			for uniqueResStringKey in uniqueResStringKeys:
	# 				resString += str(uniqueNetworkEdgeStrings[uniqueResStringKey])

	# 			output = gzip.open(self.outputDirectory + "chemiformNetworksReduced/" + self.pdbCode + ".network." + str(j) + ".txt.gz", "wt")
	# 			output.write(resString)
	# 			output.close()

	# 		# Indicate the end of a network stack with a blank line.
	# 		########################################################
	# 		networkString += "\n"

	# 		j += 1

	# def identifyAndParseIndividualSamplingVoidsReduced(self):

	# 	# Use pruneTriangulationGoFo to identify sub-networks within the void sampling network

	# 	print("Start -- identifyAndParseIndividualSamplingVoids...",)

	# 	# carbon-carbon bond lengths
	# 	# single : 1.54 Angstroms
	# 	# double : 1.33 Angstroms
	# 	# triple : 1.20 Angstroms
	# 	#########################

	# 	# Prune the global triangulation of void sampling points.
	# 	# This function includes a goFo step that splits the global triangulation into sub-networks.
	# 	############################################################################################
	# 	prunedVoidNetworkTriangulation = deepcopy(self.gridPointTriangulation) # Key feature.
	# 	prunedVoidNetworkTriangulationResult = pruneTriangulationGoFo(prunedVoidNetworkTriangulation, self.maxVoidNetworkEdgeLength)
	# 	self.prunedVoidNetworkTriangulation = prunedVoidNetworkTriangulationResult[0]
	# 	self.prunedVoidNetworks = prunedVoidNetworkTriangulationResult[2]

	# 	# Iteratively minimize the pruned triangulation.
	# 	# Remove redundant network connections by biasing shorter network edges.
	# 	########################################################################
	# 	minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation)
	# 	self.voidNetworks = getFinalMinimizedNetworks(self.prunedVoidNetworks)
	# 	writeNetworkToFile(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation, "voidNetworks", self.outputDirectory, self.pdbCode)

	# 	# i = 1
	# 	# previousResult = None
	# 	# # XXXX minimizeNetworks need to update
	# 	# minimizing = minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# while minimizing:
	# 	# 	minimizing = minimizeNetworks(self.prunedVoidNetworks, self.prunedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# 	previousResult = minimizing
	# 	# 	i += 1
	# 	# self.voidNetworks = self.prunedVoidNetworks

	# 	# # Arrange networks from largest to smallest.
	# 	# ############################################
	# 	# self.rankedVoidNetworks = {}
	# 	# for voidNetwork in self.voidNetworks:
	# 	# 	self.rankedVoidNetworks[(len(voidNetwork), voidNetwork)] = voidNetwork
	# 	# rankedVoidNetworksKeys = sorted(self.rankedVoidNetworks)
	# 	# rankedVoidNetworksKeys.reverse()

	# 	# # Create a directory for the individual chemiform network edge files.
	# 	# #####################################################################
	# 	# if not exists(self.outputDirectory + "chemiformNetworksReduced"):
	# 	# 	mkdir(self.outputDirectory + "chemiformNetworksReduced")
	# 	# else:
	# 	# 	rmtree(self.outputDirectory + "chemiformNetworksReduced")
	# 	# 	mkdir(self.outputDirectory + "chemiformNetworksReduced")

	# 	# # Create a directory for the individual chemiform PDB files.
	# 	# ############################################################
	# 	# if not exists(self.outputDirectory + "chemiformNodesReduced"):
	# 	# 	mkdir(self.outputDirectory + "chemiformNodesReduced")
	# 	# else:
	# 	# 	rmtree(self.outputDirectory + "chemiformNodesReduced")
	# 	# 	mkdir(self.outputDirectory + "chemiformNodesReduced")

	# 	# # Prepare each network to be written to file.
	# 	# #############################################
	# 	# # Void networks first...
	# 	# ########################
	# 	# j, networkName = 1, "networkEdges"
	# 	# networkString, resStrings, pdbString = "", "", ""
	# 	# for rankedVoidNetworksKey in rankedVoidNetworksKeys:

	# 	# 	voidNetwork = self.rankedVoidNetworks[rankedVoidNetworksKey]
	# 	# 	triangulation = self.voidNetworks[voidNetwork]

	# 	# 	# Indicate a new network stack in the file output.
	# 	# 	##################################################
	# 	# 	pdbString = "# Network " + str(j) + " Number of Nodes: " + str(len(voidNetwork)) + "\n"
	# 	# 	networkString += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"
	# 	# 	resStrings += "# Network " + str(j) + " Size: " + str(len(voidNetwork)) + "\n"

	# 	# 	### Start new...could be deleted with no consequences...

	# 	# 	# Iteratively reduce the void network nodes to create more compound-like chemiforms...
	# 	# 	reducing, loseNodes, reducedVoidNetworkNodes, n = 1, {}, {}, 1
	# 	# 	while reducing:
	# 	# 		print("Void reduction round", n, len(reducedVoidNetworkNodes))
	# 	# 		reducing, reducedVoidNetworkNodes = 0, {}
	# 	# 		for node in voidNetwork:

	# 	# 			if node in loseNodes:
	# 	# 				continue

	# 	# 			# for s2 in self.gridPointTriangulation[node].s2s:
	# 	# 			#for s2 in triangulation[node].s2s:
	# 	# 			pruneNodes = []
	# 	# 			for s2 in self.prunedVoidNetworkTriangulation[node].s2s:

	# 	# 				if s2[0] not in voidNetwork or s2[1] not in voidNetwork:
	# 	# 					continue

	# 	# 				# Network edge vertices.
	# 	# 				########################
	# 	# 				v1 = self.prunedVoidNetworkTriangulation[s2[0]].s1
	# 	# 				v2 = self.prunedVoidNetworkTriangulation[s2[1]].s1

	# 	# 				# Keep original nodes...
	# 	# 				d = distance(v1, v2)
	# 	# 				if d >= 1.0:
	# 	# 					reducedVoidNetworkNodes[(v1.x, v1.y, v1.z, v1.u)] = v1
	# 	# 					reducedVoidNetworkNodes[(v2.x, v2.y, v2.z, v2.u)] = v2
	# 	# 				else:
	# 	# 					# Lose the nodes...
	# 	# 					print("less than 1.0......")
	# 	# 					reducing = 1
	# 	# 					loseNodes[v1] = None
	# 	# 					loseNodes[v2] = None
	# 	# 					pruneNodes.append(v2)

	# 	# 					nodeCom = centroid4D([v1, v2], v1.data) # Use v1 as the PSA...
	# 	# 					nodeCom.id = v1.id
	# 	# 					reducedVoidNetworkNodes[(nodeCom.x, nodeCom.y, nodeCom.z, nodeCom.u)] = nodeCom
	# 	# 					self.prunedVoidNetworkTriangulation[s2[0]].s1 = nodeCom # Alters the pruned triangulation....

	# 	# 					# In the case the node was saved because although it's in this "short" network edge, it is also coincident in at least one "long" network edge...
	# 	# 					if (v1.x, v1.y, v1.z, v1.u) in reducedVoidNetworkNodes:
	# 	# 						del reducedVoidNetworkNodes[(v1.x, v1.y, v1.z, v1.u)]
	# 	# 					if (v2.x, v2.y, v2.z, v2.u) in reducedVoidNetworkNodes:
	# 	# 						del reducedVoidNetworkNodes[(v2.x, v2.y, v2.z, v2.u)]
				
	# 	# 			# Reduces and alters the pruned triangulation....
	# 	# 			pruneS2s = {}
	# 	# 			for pruneNode in pruneNodes:
	# 	# 				for s2 in self.prunedVoidNetworkTriangulation[node].s2s:
	# 	# 					if pruneNode.id in s2:
	# 	# 						pruneS2s[s2] = pruneNode
	# 	# 			for pruneS2 in pruneS2s:
	# 	# 				self.prunedVoidNetworkTriangulation[node].s2s.remove(pruneS2)

	# 	# 		if reducing:
	# 	# 			n += 1

	# 	# 	print("Number of reducedVoidNetworkNodes", len(reducedVoidNetworkNodes.values()), self.maxVoidNetworkEdgeLength)
	# 	# 	#Not enough nodes to calculate a triangulation...
	# 	# 	if len(reducedVoidNetworkNodes) < 6:
	# 	# 		continue
	# 	# 	reducedVoidNetworkMesh = convexHull4D(reducedVoidNetworkNodes.values())
	# 	# 	reducedVoidNetworkTriangulation = reducedVoidNetworkMesh.s1Dict
	# 	# 	# Chemiform size is very sensitive to the second argument: maxVoidNetworkEdgeLength: 2.0 seems like a good general value
	# 	# 	prunedReducedVoidNetworkTriangulationResult = pruneTriangulationGoFo(reducedVoidNetworkTriangulation, 2.0)
	# 	# 	prunedReducedVoidNetworkTriangulation = prunedReducedVoidNetworkTriangulationResult[0]
	# 	# 	prunedReducedVoidNetworks = prunedReducedVoidNetworkTriangulationResult[2]

	# 	# 	# Iteratively minimize the pruned triangulation.
	# 	# 	# Remove redundant network connections by biasing shorter network edges.
	# 	# 	########################################################################
	# 	# 	minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation)
	# 	# 	reducedVoidNetworks = getFinalMinimizedNetworks(prunedReducedVoidNetworks)

	# 	# 	# i = 1
	# 	# 	# previousResult = None
	# 	# 	# # XXXX minimizeNetworks need to update
	# 	# 	# minimizing = minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# 	# while minimizing:
	# 	# 	# 	minimizing = minimizeNetworks(prunedReducedVoidNetworks, prunedReducedVoidNetworkTriangulation, previousResult=previousResult, allowSymmetryLoops=1)
	# 	# 	# 	previousResult = minimizing
	# 	# 	# 	i += 1
	# 	# 	# reducedVoidNetworks = prunedReducedVoidNetworks

	# 	# 	for reducedVoidNetwork in reducedVoidNetworks:

	# 	# 		# Unique repository for aggregating grid points.
	# 	# 		################################################
	# 	# 		uniqueChemiformPoints = {}
	# 	# 		uniqueNetworkEdgeStrings = {}

	# 	# 		for node in reducedVoidNetwork:

	# 	# 			for s2 in prunedReducedVoidNetworkTriangulation[node].s2s:

	# 	# 				if s2[0] not in reducedVoidNetwork or s2[1] not in reducedVoidNetwork:
	# 	# 					continue

	# 	# 				# Network edge vertices.
	# 	# 				########################
	# 	# 				v1 = prunedReducedVoidNetworkTriangulation[s2[0]].s1
	# 	# 				uniqueChemiformPoints[(v1.data.atom_serial, v1)] = v1
	# 	# 				v2 = prunedReducedVoidNetworkTriangulation[s2[1]].s1
	# 	# 				uniqueChemiformPoints[(v2.data.atom_serial, v2)] = v2

	# 	# 				print("Chemiform edge distance:", distance(v1, v2))

	# 	# 				# Number each grid point atoms by its network.
	# 	# 				##############################################
	# 	# 				v1.data.residue_sequence_number, v2.data.residue_sequence_number = j, j

	# 	# 				# Format coordinates for writing (a bit redundant: could go with one).
	# 	# 				######################################################################
	# 	# 				format = "%15.6f%15.6f%15.6f "
	# 	# 				resCoords = format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
	# 	# 				networkString += format % (v1.x,v1.y,v1.z) + format % (v2.x,v2.y,v2.z) + "\n"
	# 	# 				resStringFormat = "%s-%i-%s"
	# 	# 				resString = "%13s %13s | " % (resStringFormat % (v1.data.residue.chn, v1.data.residue.num, v1.data.residue.name),
	# 	# 												resStringFormat % (v2.data.residue.chn, v2.data.residue.num, v2.data.residue.name))
	# 	# 				if resString not in uniqueNetworkEdgeStrings:
	# 	# 					uniqueNetworkEdgeStrings[resString] = resString + resCoords
	# 	# 				resStrings += resString + resCoords

	# 	# 		if uniqueChemiformPoints:

	# 	# 			uniqueGridPointKeys = sorted(uniqueChemiformPoints)
	# 	# 			conectString = ""
	# 	# 			for uniqueGridPointKey in uniqueGridPointKeys:
	# 	# 				v = uniqueGridPointKey[1]
	# 	# 				conectString += "%-6s%5i" % ("CONECT", v.data.atom_serial)
	# 	# 				for s2 in prunedReducedVoidNetworkTriangulation[v.id].s2s:
	# 	# 					for vid in s2:
	# 	# 						if vid != v.id:
	# 	# 							vConect = prunedReducedVoidNetworkTriangulation[vid]
	# 	# 							conectString += "%5i" % vConect.s1.data.atom_serial
	# 	# 				conectString += "\n"

	# 	# 			uniqueGridPointKeys = sorted(uniqueChemiformPoints)
	# 	# 			pdbString = "# Network " + str(j) + " Size: " + str(len(uniqueGridPointKeys)) + "\n"
	# 	# 			resNum = 1
	# 	# 			for uniqueGridPointKey in uniqueGridPointKeys:
	# 	# 				atom = uniqueChemiformPoints[uniqueGridPointKey].data
	# 	# 				atom.record_name = "HETATM"
	# 	# 				atom.atom_name = " C"
	# 	# 				atom.residue_sequence_number = resNum
	# 	# 				atom.element_symbol = "C"
	# 	# 				pdbString += str(uniqueChemiformPoints[uniqueGridPointKey].data)
	# 	# 				resNum += 1

	# 	# 			output = open(self.outputDirectory + "chemiformNodesReduced/" + self.pdbCode + ".nodes." + str(j) + ".pdb", "w")
	# 	# 			output.write(pdbString)
	# 	# 			output.write(conectString)
	# 	# 			output.close()

	# 	# 		if uniqueNetworkEdgeStrings:

	# 	# 			uniqueResStringKeys = sorted(uniqueNetworkEdgeStrings)
	# 	# 			resString = "# Network " + str(j) + " Size: " + str(len(uniqueNetworkEdgeStrings)) + "\n"
	# 	# 			for uniqueResStringKey in uniqueResStringKeys:
	# 	# 				resString += str(uniqueNetworkEdgeStrings[uniqueResStringKey])

	# 	# 			output = gzip.open(self.outputDirectory + "chemiformNetworksReduced/" + self.pdbCode + ".network." + str(j) + ".txt.gz", "wt")
	# 	# 			output.write(resString)
	# 	# 			output.close()

	# 	# 		# Indicate the end of a network stack with a blank line.
	# 	# 		########################################################
	# 	# 		networkString += "\n"

	# 	# 	j += 1

	# 	print("End -- identifyAndParseIndividualSamplingVoids...")
		
	# def identifyCreviceFacets(self, facetDistanceLimit=2.0):

	# 	print("Start -- Identifying crevice facets in identifyCreviceFacets...")

	# 	voidNetworkVertices = {}
	# 	for voidNetwork in self.voidNetworks:

	# 		triangulation = self.voidNetworks[voidNetwork]
	# 		for node in voidNetwork:
	# 			for s2 in triangulation[node].s2s:
	# 				if s2[0] not in voidNetwork or s2[1] not in voidNetwork:
	# 					continue

	# 				# Network edge vertices.
	# 				########################
	# 				v1 = self.gridPointTriangulation[s2[0]].s1
	# 				voidNetworkVertices[v1] = None
	# 				v2 = self.gridPointTriangulation[s2[1]].s1
	# 				voidNetworkVertices[v2] = None

	# 	quadSurfaceByCentroid = {}
	# 	for edgeKey in self.quadSurface:
	# 		centroid = centroid4D([self.quadSurface[edgeKey].origin_vertex, self.quadSurface[edgeKey].destination_vertex], PseudoAtom())
	# 		quadSurfaceByCentroid[edgeKey] = centroid
			

	# 	quadSurfaceCrevice = {}
	# 	for v in voidNetworkVertices:
	# 		for edgeKey in quadSurfaceByCentroid:
	# 			centroid = quadSurfaceByCentroid[edgeKey]
	# 			d = distance(v, centroid)
	# 			if d < facetDistanceLimit:
	# 				quadSurfaceCrevice[edgeKey] = self.quadSurface[edgeKey]
	# 	self.quadSurfaceCrevice = quadSurfaceCrevice


	# 	# Write crevice surface facets to file.
	# 	#######################################
	# 	surfaceFacetString = ""
	# 	for edgeKey in self.quadSurfaceCrevice:
	# 		lf = self.quadSurfaceCrevice[edgeKey].lf
	# 		surfaceFacetString += s3StringOutput(lf.v1, lf.v2, lf.v3)
	# 		rf = self.quadSurfaceCrevice[edgeKey].rf
	# 		surfaceFacetString += s3StringOutput(rf.v1, rf.v2, rf.v3)
	# 		centroid = quadSurfaceByCentroid[edgeKey]

	# 	outputSurf = gzip.open(self.outputDirectory + "creviceSurface.txt.gz","wt")
	# 	outputSurf.write(surfaceFacetString)
	# 	outputSurf.close()

	# 	print("End -- Identifying crevice facets in identifyCreviceFacets...")

	def calculateSamplingVoidSurfaces(self, extend_sampling=False):

		if self.group_chains:

			# Arrange networks from largest to smallest. Also, prune using self.minNetworkSize...
			# This condition is relevant in the case identifyAndParseIndividualSamplingVoids() has not been called...
			print("Start -- Calculating void surfaces in calculateSamplingVoidSurfaces...")
			if not self.rankedVoidNetworks:
				for voidNetwork in self.voidNetworks:
					if len(voidNetwork) > self.minVoidNetworkSize:
						self.rankedVoidNetworks[(len(voidNetwork), voidNetwork)] = voidNetwork
			keys = sorted(self.rankedVoidNetworks)
			keys.reverse()

			# Calculate the individual network surfaces.
			############################################
			i, chemiSurfaces = 1, {}
			for key in keys:

				if len(self.rankedVoidNetworks[key]) >= 5:

					vertexList = []
					for simplex1Id in self.rankedVoidNetworks[key]:
						vertexList.append(self.prunedVoidNetworkTriangulation[simplex1Id].s1)
					vertexDict = {}
					for v in vertexList:
						vertexDict[v.id] = v
					last_id = sorted(vertexDict)[-1]

					if extend_sampling:
						from sphere import Sphere
						atom_serial = last_id + 1
						extended_vertexList = []
						for v in vertexList:
							s = Sphere()
							s.radius = 1.
							s.origin_x = v.x
							s.origin_y = v.y
							s.origin_z = v.z
							s.generate_surface_points()
							extended_vertices = s.get_pseudoatoms(atom_serial=atom_serial, return_vertices=True)
							for ev in extended_vertices:
								ev4d = Vertex4D((ev.x, ev.y, ev.z, ev.x**2 + ev.y**2 + ev.z**2), data=v.data, unique_id=ev.id)
								extended_vertexList.append(ev4d)
							atom_serial += len(extended_vertices) + 1
						vertexList.extend(extended_vertexList)
						vertexDict = {}
						for v in vertexList:
							vertexDict[v.id] = v			

					print("Start -- calculating void chemiSurface...", i)
					voidSurfaceResultTuple = calculateSurface(vertexList, 2.0, minArea=1.0, allowSmallSurfaces=50)
					# Have to have a surface....
					############################
					if voidSurfaceResultTuple:
						chemiSurfaces[key] = voidSurfaceResultTuple[1] # should be quadSurface

					print("End -- calculating void chemiSurface...", i, "\n\n" )
					i += 1

			pymolScriptContent = """from pymol.cgo import *
from pymol import cmd
"""

			i = 1
			chemiSurfacesKeys = sorted(chemiSurfaces)
			chemiSurfacesKeys.reverse()
			for chemiSurfacesKey in chemiSurfacesKeys:

				surfaceFacets = {}
				quadSurface = chemiSurfaces[chemiSurfacesKey]
				# A quadSurface is not calculated for voids having < 5 sampling points 
				if quadSurface:
					for e in quadSurface:
						surfaceFacets.update({quadSurface[e].lf.id:quadSurface[e].lf})
						surfaceFacets.update({quadSurface[e].rf.id:quadSurface[e].rf})
		
					pymolScriptContent += """
obj = [
	BEGIN, LINES,
"""


					for fKey in surfaceFacets:
						f = surfaceFacets[fKey]
						v1, v2, v3 = f.v1, f.v2, f.v3
						pymolScriptContent += f"\tVERTEX, {v1.x}, {v1.y}, {v1.z},\n"
						pymolScriptContent += f"\tVERTEX, {v2.x}, {v2.y}, {v2.z},\n"
						pymolScriptContent += f"\tVERTEX, {v2.x}, {v2.y}, {v2.z},\n"
						pymolScriptContent += f"\tVERTEX, {v3.x}, {v3.y}, {v3.z},\n"
						pymolScriptContent += f"\tVERTEX, {v3.x}, {v3.y}, {v3.z},\n"
						pymolScriptContent += f"\tVERTEX, {v1.x}, {v1.y}, {v1.z},\n"

					pymolScriptContent += f"""
	END
]

cmd.load_cgo(obj, 'cs_{i}_{self.chainString}')"""

					i += 1

			# Save the PyMOL script
			outFile = self.outPath + f"{self.pdbCode}.chemiSurface-" + self.chainString + ".py"
			with open(outFile, "w") as pymolScriptFile:
				pymolScriptFile.write(pymolScriptContent)

			print("End -- Calculating void surfaces in calculateSamplingVoidSurfaces...")

		else:

			for chain in self.voidNetworks_bychain:

				voidNetworks = self.voidNetworks_bychain[chain]
				self.rankedVoidNetworks_bychain[chain] = {}

				# Arrange networks from largest to smallest. Also, prune using self.minNetworkSize...
				# This condition is relevant in the case identifyAndParseIndividualSamplingVoids() has not been called...
				print("Start -- Calculating void surfaces in calculateSamplingVoidSurfaces...")
				if not self.rankedVoidNetworks_bychain[chain]:
					for voidNetwork in voidNetworks:
						if len(voidNetwork) > self.minVoidNetworkSize:
							self.rankedVoidNetworks_bychain[chain][(len(voidNetwork), voidNetwork)] = voidNetwork
				keys = sorted(self.rankedVoidNetworks_bychain[chain])
				keys.reverse()

				# Calculate the individual network surfaces.
				############################################
				i, chemiSurfaces = 1, {}
				for key in keys:

					if len(self.rankedVoidNetworks_bychain[chain][key]) >= 5:

						vertexList = []
						for simplex1Id in self.rankedVoidNetworks_bychain[chain][key]:
							vertexList.append(self.prunedVoidNetworkTriangulation_bychain[chain][simplex1Id].s1)
						vertexDict = {}
						for v in vertexList:
							vertexDict[v.id] = v
						last_id = sorted(vertexDict)[-1]

						if extend_sampling:
							from sphere import Sphere
							atom_serial = last_id + 1
							extended_vertexList = []
							for v in vertexList:
								s = Sphere()
								s.radius = 1.
								s.origin_x = v.x
								s.origin_y = v.y
								s.origin_z = v.z
								s.generate_surface_points()
								extended_vertices = s.get_pseudoatoms(atom_serial=atom_serial, return_vertices=True)
								for ev in extended_vertices:
									ev4d = Vertex4D((ev.x, ev.y, ev.z, ev.x**2 + ev.y**2 + ev.z**2), data=v.data, unique_id=ev.id)
									extended_vertexList.append(ev4d)
								atom_serial += len(extended_vertices) + 1
							vertexList.extend(extended_vertexList)
							vertexDict = {}
							for v in vertexList:
								vertexDict[v.id] = v			

						print("Start -- calculating void chemiSurface...", i)
						voidSurfaceResultTuple = calculateSurface(vertexList, 2.0, minArea=1.0, allowSmallSurfaces=50)
						# Have to have a surface....
						############################
						if voidSurfaceResultTuple:
							chemiSurfaces[key] = voidSurfaceResultTuple[1] # should be quadSurface

						print("End -- calculating void chemiSurface...", i, "\n\n" )
						i += 1

				pymolScriptContent = """from pymol.cgo import *
from pymol import cmd
"""

				i = 1
				chemiSurfacesKeys = sorted(chemiSurfaces)
				chemiSurfacesKeys.reverse()
				for chemiSurfacesKey in chemiSurfacesKeys:

					surfaceFacets = {}
					quadSurface = chemiSurfaces[chemiSurfacesKey]
					# A quadSurface is not calculated for voids having < 5 sampling points 
					if quadSurface:
						for e in quadSurface:
							surfaceFacets.update({quadSurface[e].lf.id:quadSurface[e].lf})
							surfaceFacets.update({quadSurface[e].rf.id:quadSurface[e].rf})
			
						pymolScriptContent += """
obj = [
	BEGIN, LINES,
"""


						for fKey in surfaceFacets:
							f = surfaceFacets[fKey]
							v1, v2, v3 = f.v1, f.v2, f.v3
							pymolScriptContent += f"\tVERTEX, {v1.x}, {v1.y}, {v1.z},\n"
							pymolScriptContent += f"\tVERTEX, {v2.x}, {v2.y}, {v2.z},\n"
							pymolScriptContent += f"\tVERTEX, {v2.x}, {v2.y}, {v2.z},\n"
							pymolScriptContent += f"\tVERTEX, {v3.x}, {v3.y}, {v3.z},\n"
							pymolScriptContent += f"\tVERTEX, {v3.x}, {v3.y}, {v3.z},\n"
							pymolScriptContent += f"\tVERTEX, {v1.x}, {v1.y}, {v1.z},\n"

						pymolScriptContent += f"""
	END
]

cmd.load_cgo(obj, 'cs_{i}_{chain}')"""

						i += 1

				# Save the PyMOL script
				outFile = self.outPath + chain + sep + f"{self.pdbCode}.chemiSurface-" + chain +".py"
				with open(outFile, "w") as pymolScriptFile:
					pymolScriptFile.write(pymolScriptContent)

				print("End -- Calculating void surfaces in calculateSamplingVoidSurfaces...")

	def calculateSamplingVoidHulls(self):
		pass
