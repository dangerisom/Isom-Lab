# Copyright (c) 2025 Isom Lab
# This Source Code Form is subject to the terms of the Mozilla Public License, v. 2.0.
# If a copy of the MPL was not distributed with this file, You can obtain one at
# https://mozilla.org/MPL/2.0/.


#!/usr/bin/env python3
"""
GPCR TSV filter & Excel writer (openpyxl) with tidy tqdm progress bars,
sequence-column exclusion, flexible 7TM+GPCR matching (Plan B),
and GCR+digits family rule.

Dependencies:
  pip install pandas openpyxl tqdm
"""

from pathlib import Path
from typing import List, Optional
import pandas as pd
import re
from collections import Counter

# ========= User Config =========
input_tsv = Path("/Volumes/isomlab_1/bulk_interpro/output_with_interpro.tsv")
output_dir = input_tsv.parent / "gpcr_outputs"
chunksize = 100_000
excel_max_rows = 1_048_576  # Excel per-sheet max (including header)
precount_lines = True       # Set to False to skip first-pass line count

gpcr_patterns = [
    "gpcr", "g protein-coupled receptor", "g-protein coupled receptor",
    "g-protein-coupled receptor", "opsin", "rhodopsin", "olfactory",
    "vomeronasal", "taste", "pheromone", "frizzled", "purinoceptor",
    "odorant", "tastant", "7tm_gpcr_serpentine",  # Plan B will generalize this
    "serpentine",
    "transmembrane 7 superfamily member", "G_PROTEIN_RECEP",
    "Gustatory", "OPSD", "CHEMOKINE RECEPTOR", "G PROTEIN-COUPLED",
    "FOLLICLE-STIMULATING HORMONE RECEPTOR", "CRE-SRX-10 PROTEIN", "-SRX",
    "GLUCOSE RECEPTOR", 
    "GCR",  # NEW: triggers GCR-?\d+ rule (e.g., GCR128, GCR-128)
]
skip_patterns = []

# ========= tqdm handling =========
try:
    from tqdm import tqdm
    _TQDM_AVAILABLE = True
except Exception:
    _TQDM_AVAILABLE = False
    def tqdm(iterable=None, total=None, desc=None, unit=None, **kwargs):
        return iterable if iterable is not None else None

def pbar_start(total=None, desc="", unit=" rows"):
    if _TQDM_AVAILABLE:
        return tqdm(total=total, desc=desc, unit=unit, dynamic_ncols=True, smoothing=0.1, leave=True)
    return None

def pbar_update(pbar, n):
    if _TQDM_AVAILABLE and (pbar is not None):
        pbar.update(n)

def pbar_close(pbar):
    if _TQDM_AVAILABLE and (pbar is not None):
        pbar.close()

# ========= Regex builder with Plan B + GCR family =========
def build_regex_from_list(terms: List[str]) -> re.Pattern:
    """
    Build a regex matching ANY term (case-insensitive via UPPERCASE haystack).
    - Special handling for "7TM ... GPCR" family: 7TM[_ -]?GPCR with optional suffix tokens.
    - Special handling for "GCR": GCR-?\d+ (e.g., GCR128, GCR-128).
    - Short single-word tokens (2–6 chars, [A-Z0-9_]+) use word boundaries.
    - Longer tokens/phrases match as substrings (escaped).
    """
    patterns = []
    for t in terms:
        t = str(t).strip()
        if not t:
            continue
        T = t.upper()

        # Plan B: flexible 7TM+GPCR
        if T in {"7TM_GPCR", "7TM GPCR", "7TM-GPCR", "7TM_GPCR_SERPENTINE"}:
            patterns.append(r"\b7TM[_\-\s]?GPCR(?:\b|[_\-\s][A-Z0-9_]+)*")
            continue

        # GCR family: GCR + optional hyphen + digits
        if T == "GCR":
            patterns.append(r"\bGCR-?\d+\b")
            continue

        # Default
        if re.fullmatch(r"[A-Z0-9_]+", T) and 2 <= len(T) <= 6:
            patterns.append(rf"\b{re.escape(T)}\b")
        else:
            patterns.append(re.escape(T))

    if not patterns:
        return re.compile(r"(?!)")
    return re.compile("|".join(patterns))

def detect_superkingdom_column(columns: List[str]) -> Optional[str]:
    for c in columns:
        key = c.lower().replace(" ", "").replace("_", "")
        if "superkingdom" in key:
            return c
    return None

# ---- Sequence column detection ----
_AA_SET = set("ACDEFGHIKLMNPQRSTVWY")
_AA_EXTRAS = set("BXZJUO*")

def _looks_like_protein_sequence(s: str, min_len: int = 60, max_non_aa_frac: float = 0.0) -> bool:
    """
    Heuristic: value looks like a protein sequence if:
    - length >= min_len, and
    - nearly all characters are in AA sets.
    """
    if s is None:
        return False
    s = str(s).strip()
    if len(s) < min_len:
        return False
    up = s.upper()
    allowed = _AA_SET | _AA_EXTRAS
    non_aa = sum(ch not in allowed for ch in up)
    return non_aa / max(1, len(up)) <= max_non_aa_frac

def detect_sequence_columns(df: pd.DataFrame, sample_rows: int = 200, threshold: float = 0.8) -> List[str]:
    """
    Return column names where >= threshold of sampled values look like AA sequences.
    """
    seq_cols: List[str] = []
    if df.empty:
        return seq_cols
    sample = df.head(sample_rows)
    for col in df.columns:
        vals = sample[col].astype(str).tolist()
        if not vals:
            continue
        hits = sum(_looks_like_protein_sequence(v) for v in vals)
        if hits / len(vals) >= threshold:
            seq_cols.append(col)
    return seq_cols

# ========= Excel stream writer (openpyxl) =========
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class ExcelStreamWriter:
    """
    Stream-append DataFrame rows into an .xlsx file using openpyxl,
    rolling over to a new sheet when per-sheet row limit is reached.
    """
    def __init__(self, out_path: Path, excel_max_rows: int = 1_048_576, sheet_base: str = "sheet"):
        self.out_path = Path(out_path)
        self.excel_max_rows = excel_max_rows
        self.sheet_base = sheet_base

        self.wb = Workbook(write_only=False)
        default_ws = self.wb.active
        self.wb.remove(default_ws)

        self.header_font = Font(bold=True)
        self.current_sheet_idx = 1
        self.ws = None

        self.columns: Optional[List[str]] = None
        self.header_written = False
        self.current_row = 0  # 1-based

        self._open_new_sheet()

    def _open_new_sheet(self):
        name = f"{self.sheet_base}_{self.current_sheet_idx}"
        if len(name) > 31:
            name = name[:31]
        self.ws = self.wb.create_sheet(title=name)
        self.header_written = False
        self.current_row = 0

    def _write_header(self):
        if not self.columns:
            return
        self.current_row = 1
        for c_idx, col_name in enumerate(self.columns, start=1):
            cell = self.ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = self.header_font
        self.header_written = True
        self.current_row = 2

    def _rows_available(self) -> int:
        if not self.header_written:
            return self.excel_max_rows - 1
        else:
            written_data_rows = (self.current_row - 2) if self.current_row >= 2 else 0
            return (self.excel_max_rows - 1) - written_data_rows

    def write_chunk(self, df: pd.DataFrame):
        if df.empty:
            return

        if self.columns is None:
            self.columns = list(df.columns)

        df = df.reindex(columns=self.columns)

        remaining = len(df)
        start = 0
        while remaining > 0:
            if not self.header_written:
                self._write_header()

            space_left = self._rows_available()
            if space_left <= 0:
                self.current_sheet_idx += 1
                self._open_new_sheet()
                continue

            n_write = min(space_left, remaining)
            end = start + n_write
            part = df.iloc[start:end]

            values = part.values.tolist()
            for r_off, row_vals in enumerate(values):
                row_idx = self.current_row + r_off
                for c_idx, v in enumerate(row_vals, start=1):
                    self.ws.cell(row=row_idx, column=c_idx, value=v)

            self.current_row += n_write
            remaining -= n_write
            start = end

    def autofit_headers_only(self):
        if not self.columns:
            return
        try:
            for c_idx, name in enumerate(self.columns, start=1):
                width = min(max(10, len(str(name)) + 2), 50)
                self.ws.column_dimensions[get_column_letter(c_idx)].width = width
        except Exception:
            pass

    def close(self):
        self.autofit_headers_only()
        self.wb.save(self.out_path)

# ========= Utilities =========
def count_lines_fast(path: Path) -> int:
    """Count lines for an accurate total rows pbar."""
    total = 0
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for _ in f:
            total += 1
    return max(0, total - 1)  # subtract header

# ========= Main =========
def main():
    output_dir.mkdir(parents=True, exist_ok=True)

    gpcr_re = build_regex_from_list(gpcr_patterns)
    skip_re = build_regex_from_list(skip_patterns)

    gpcr_xlsx = ExcelStreamWriter(output_dir / f"{input_tsv.stem}__GPCR.xlsx",
                                  excel_max_rows=excel_max_rows, sheet_base="gpcr")
    stmp_xlsx = ExcelStreamWriter(output_dir / f"{input_tsv.stem}__STMP.xlsx",
                                  excel_max_rows=excel_max_rows, sheet_base="stmp")

    gpcr_counts = Counter()
    stmp_counts = Counter()
    combined_counts = Counter()

    superkingdom_col: Optional[str] = None
    seq_cols: List[str] = []

    # Precount (optional) for polished total on the rows pbar
    total_rows_estimate = None
    if precount_lines:
        precount_pbar = pbar_start(desc="[0/3] Counting lines", unit=" lines")
        total_rows_estimate = count_lines_fast(input_tsv)
        pbar_update(precount_pbar, total_rows_estimate)
        pbar_close(precount_pbar)

    read_pbar = pbar_start(total=None, desc="[1/3] Reading chunks", unit=" chunks")
    rows_pbar = pbar_start(total=total_rows_estimate, desc="[2/3] Scanning & splitting", unit=" rows")

    # Stream by chunks
    for chunk_idx, df in enumerate(pd.read_csv(
        input_tsv, sep="\t", dtype=str, keep_default_na=False,
        chunksize=chunksize, on_bad_lines="skip"
    )):
        if chunk_idx == 0:
            superkingdom_col = detect_superkingdom_column(df.columns.tolist())
            seq_cols = detect_sequence_columns(df)
            # Also exclude common explicit sequence column names if present
            for n in ("sequence", "seq", "aa_sequence", "protein_sequence"):
                if (n in df.columns) and (n not in seq_cols):
                    seq_cols.append(n)

        # Uppercase haystack across NON-sequence columns only
        cols_to_scan = [c for c in df.columns if c not in seq_cols]
        if cols_to_scan:
            df_upper = df[cols_to_scan].map(lambda x: str(x).upper() if x is not None else "")
            haystack = df_upper.apply(lambda row: " ".join(row.values.tolist()), axis=1)
        else:
            haystack = pd.Series([""] * len(df))

        gpcr_hit = haystack.str.contains(gpcr_re, regex=True, na=False)
        skip_hit = haystack.str.contains(skip_re, regex=True, na=False)
        gpcr_mask = gpcr_hit & ~skip_hit
        stmp_mask = ~gpcr_mask

        gpcr_df = df.loc[gpcr_mask]
        stmp_df = df.loc[stmp_mask]

        gpcr_xlsx.write_chunk(gpcr_df)
        stmp_xlsx.write_chunk(stmp_df)

        if superkingdom_col and superkingdom_col in df.columns:
            combined_counts.update(df[superkingdom_col].astype(str).tolist())
            if not gpcr_df.empty:
                gpcr_counts.update(gpcr_df[superkingdom_col].astype(str).tolist())
            if not stmp_df.empty:
                stmp_counts.update(stmp_df[superkingdom_col].astype(str).tolist())

        pbar_update(read_pbar, 1)
        pbar_update(rows_pbar, len(df))

    pbar_close(read_pbar)
    pbar_close(rows_pbar)

    # Finalize Excel outputs
    finish_pbar = pbar_start(total=3, desc="[3/3] Finalizing", unit=" steps")
    gpcr_xlsx.close();            pbar_update(finish_pbar, 1)
    stmp_xlsx.close();            pbar_update(finish_pbar, 1)

    # Write counts workbook (openpyxl via pandas)
    counts_path = output_dir / f"{input_tsv.stem}__SUPERKINGDOM_COUNTS.xlsx"
    with pd.ExcelWriter(counts_path, engine="openpyxl") as writer:
        def write_counts_sheet(name: str, counter: Counter):
            if counter:
                cdf = (pd.DataFrame(counter.items(), columns=["superkingdom", "count"])
                       .sort_values(["count", "superkingdom"], ascending=[False, True])
                       .reset_index(drop=True))
            else:
                cdf = pd.DataFrame(columns=["superkingdom", "count"])
            cdf.to_excel(writer, sheet_name=name, index=False)

        write_counts_sheet("gpcr", gpcr_counts)
        write_counts_sheet("stmp", stmp_counts)
        write_counts_sheet("combined", combined_counts)

        summary = pd.DataFrame([
            {"metric": "superkingdom_col", "value": superkingdom_col or "(not detected)"},
            {"metric": "seq_columns_excluded", "value": ", ".join(seq_cols) if seq_cols else "(none)"},
        ])
        summary.to_excel(writer, sheet_name="summary", index=False)

    pbar_update(finish_pbar, 1)
    pbar_close(finish_pbar)

if __name__ == "__main__":
    main()
