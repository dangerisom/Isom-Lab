# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# Copyright (c) 2026 Daniel G. Isom

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import cv2
import numpy as np
import csv
from compGeometry import Vertex
from convexHull3D_2_1 import convexHull3D
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image, ImageTk


class Data:
    def __init__(self):
        self.label = ""
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.mixed = False

from openpyxl.utils import get_column_letter

def _autofit_worksheet_columns(ws, min_width=8, max_width=50, padding=2):
    for col_idx, column_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            cell_len = len(str(cell.value))
            if cell_len > max_length:
                max_length = cell_len

        adjusted_width = min(max(max_length + padding, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


class ImageTriangulatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Vesicle Triangulator")
        self.geometry("1100x1100")
        self.minsize(1300, 1100)

        self.main_container = tk.Frame(self)
        self.main_container.pack(fill="both", expand=True)
        self.left_panel = tk.Frame(self.main_container)
        self.left_panel.pack(side="left", fill="both", expand=True)

        # Data containers
        self.csv_paths: dict[str, str] = {}
        self.component_image_paths: dict[str, str] = {}
        self.component_image_overlay_path: str = ""

        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.default_dir = self.script_dir

        # Editable in GUI
        self.t_min_ch1 = 25
        self.t_mid_ch1 = 120
        self.t_max_ch1 = 255
        self.t_min_ch2 = 25
        self.t_mid_ch2 = 120
        self.t_max_ch2 = 255
        # self.component_thickness = 1
        self.component_area_min = 1
        self.component_area_max = 1000
        self.max_edge_length = 20

        # internal default colors still used by processing, but not editable in GUI
        self.component_color_channel_1 = (255, 255, 0)
        self.component_color_channel_2 = (200, 130, 255)

        # Path state
        self.channel1_path = ""
        self.channel2_path = ""
        self.original_overlay_path = ""
        self.save_dir = ""

        # Whether user explicitly picked each field after defaults were propagated
        self.channel2_is_explicit = False
        self.overlay_is_explicit = False
        self.save_is_explicit = False

        # === Image Channel 1 ===
        self.frame_channel1 = ttk.LabelFrame(self.left_panel, text="Image Channel 1")
        self.frame_channel1.pack(padx=20, pady=(20, 10), fill="x")
        self.btn_channel1 = ttk.Button(
            self.frame_channel1,
            text="Load Channel 1",
            command=self.load_channel1
        )
        self.btn_channel1.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.lbl_channel1 = ttk.Label(
            self.frame_channel1,
            text="No file selected",
            foreground="gray"
        )
        self.lbl_channel1.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # === Image Channel 2 ===
        self.frame_channel2 = ttk.LabelFrame(self.left_panel, text="Image Channel 2")
        self.frame_channel2.pack(padx=20, pady=(0, 10), fill="x")
        self.btn_channel2 = ttk.Button(
            self.frame_channel2,
            text="Load Channel 2",
            command=self.load_channel2
        )
        self.btn_channel2.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.lbl_channel2 = ttk.Label(
            self.frame_channel2,
            text="No file selected",
            foreground="gray"
        )
        self.lbl_channel2.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # === Original Overlay Image ===
        self.frame_overlay = ttk.LabelFrame(self.left_panel, text="Original Overlay Image")
        self.frame_overlay.pack(padx=20, pady=(0, 10), fill="x")
        self.btn_overlay = ttk.Button(
            self.frame_overlay,
            text="Load Overlay",
            command=self.load_original_overlay
        )
        self.btn_overlay.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.lbl_overlay = ttk.Label(
            self.frame_overlay,
            text="No file selected",
            foreground="gray"
        )
        self.lbl_overlay.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        self.original_overlay_path = ""

        # === Save Output ===
        self.frame_save = ttk.LabelFrame(self.left_panel, text="Save Output Directory")
        self.frame_save.pack(padx=20, pady=(0, 20), fill="x")
        self.btn_save = ttk.Button(
            self.frame_save,
            text="Select Save Location",
            command=self.select_save_directory
        )
        self.btn_save.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.lbl_save = ttk.Label(
            self.frame_save,
            text="No directory selected",
            foreground="gray"
        )
        self.lbl_save.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # === Threshold & Component Settings ===
        self.frame_params = ttk.LabelFrame(self.left_panel, text="Threshold & Component Settings")
        self.frame_params.pack(padx=20, pady=(0, 10), fill="x")

        tk.Label(self.frame_params, text="Ch1 Threshold Min:") \
            .grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.entry_threshold_min_ch1 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_min_ch1.insert(0, str(self.t_min_ch1))
        self.entry_threshold_min_ch1.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Ch1 Threshold Mid:") \
            .grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.entry_threshold_mid_ch1 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_mid_ch1.insert(0, str(self.t_mid_ch1))
        self.entry_threshold_mid_ch1.grid(row=0, column=3, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Ch1 Threshold Max:") \
            .grid(row=0, column=4, padx=10, pady=5, sticky="w")
        self.entry_threshold_max_ch1 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_max_ch1.insert(0, str(self.t_max_ch1))
        self.entry_threshold_max_ch1.grid(row=0, column=5, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Ch2 Threshold Min:") \
            .grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.entry_threshold_min_ch2 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_min_ch2.insert(0, str(self.t_min_ch2))
        self.entry_threshold_min_ch2.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Ch2 Threshold Mid:") \
            .grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.entry_threshold_mid_ch2 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_mid_ch2.insert(0, str(self.t_mid_ch2))
        self.entry_threshold_mid_ch2.grid(row=1, column=3, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Ch2 Threshold Max:") \
            .grid(row=1, column=4, padx=10, pady=5, sticky="w")
        self.entry_threshold_max_ch2 = ttk.Entry(self.frame_params, width=8)
        self.entry_threshold_max_ch2.insert(0, str(self.t_max_ch2))
        self.entry_threshold_max_ch2.grid(row=1, column=5, padx=10, pady=5, sticky="w")

        # tk.Label(self.frame_params, text="Component Thickness:")\
        #     .grid(row=2, column=0, padx=10, pady=5, sticky="w")
        # self.entry_component_thickness = ttk.Entry(self.frame_params, width=8)
        # self.entry_component_thickness.insert(0, str(self.component_thickness))
        # self.entry_component_thickness.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Area Min (pixels):")\
            .grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.entry_area_min = ttk.Entry(self.frame_params, width=8)
        self.entry_area_min.insert(0, str(self.component_area_min))
        self.entry_area_min.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        tk.Label(self.frame_params, text="Area Max (pixels):")\
            .grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.entry_area_max = ttk.Entry(self.frame_params, width=8)
        self.entry_area_max.insert(0, str(self.component_area_max))
        self.entry_area_max.grid(row=3, column=3, padx=10, pady=5, sticky="w")

        self.btn_threshold_ch1 = ttk.Button(
            self.frame_params,
            text="Threshold Ch1 Only",
            command=self.on_threshold_ch1_only,
        )
        self.btn_threshold_ch1.grid(
            row=4, column=0, sticky="w", padx=10, pady=(10, 5)
        )

        self.btn_threshold_ch2 = ttk.Button(
            self.frame_params,
            text="Threshold Ch2 Only",
            command=self.on_threshold_ch2_only,
        )
        self.btn_threshold_ch2.grid(
            row=4, column=1, sticky="w", padx=10, pady=(10, 5)
        )

        # === Triangulation Settings ===
        self.frame_triangulation = ttk.LabelFrame(self.left_panel, text="Triangulation")
        self.frame_triangulation.pack(padx=20, pady=(10, 10), fill="x")

        tk.Label(self.frame_triangulation, text="Max Edge Length:") \
            .grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.entry_max_edge_length = ttk.Entry(self.frame_triangulation, width=8)
        self.entry_max_edge_length.insert(0, str(self.max_edge_length))
        self.entry_max_edge_length.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        self.btn_triangulate = ttk.Button(
            self.frame_triangulation,
            text="Triangulate",
            command=self.on_triangulate
        )
        self.btn_triangulate.grid(row=1, column=0, sticky="w", padx=10, pady=(10, 5))

        # === Progress / Status BELOW triangulation ===
        self.status_frame = ttk.LabelFrame(self.left_panel, text="Progress / Status")
        self.status_frame.pack(padx=20, pady=(0, 10), fill="both", expand=True)
        self.status_text = tk.Text(
            self.status_frame,
            height=12,
            state="disabled",
            wrap="word"
        )
        self.status_text.pack(fill="both", expand=True, padx=5, pady=5)

    def log_status(self, msg: str):
        self.status_text.config(state="normal")
        self.status_text.insert("end", msg + "\n")
        self.status_text.see("end")
        self.status_text.config(state="disabled")
        self.update_idletasks()

    def _make_timestamped_output_dir(self, base_dir: str) -> str:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return os.path.join(base_dir, f"output_{timestamp}")

    def write_parameter_log(self, output_root: str):
        """
        Write a neatly formatted XLSX with all relevant input parameters
        into the parent directory (output_root).
        """
        param_xlsx = os.path.join(output_root, "parameter_log.xlsx")

        ch1_color_str = (
            f"{self.component_color_channel_1[0]}, "
            f"{self.component_color_channel_1[1]}, "
            f"{self.component_color_channel_1[2]}"
        )
        ch2_color_str = (
            f"{self.component_color_channel_2[0]}, "
            f"{self.component_color_channel_2[1]}, "
            f"{self.component_color_channel_2[2]}"
        )

        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        data = {
            "Parameter": [
                "Timestamp",
                "Ch1 Threshold Min",
                "Ch1 Threshold Mid",
                "Ch1 Threshold Max",
                "Ch2 Threshold Min",
                "Ch2 Threshold Mid",
                "Ch2 Threshold Max",
                # "Component Thickness",
                "Component Area Min",
                "Component Area Max",
                "Max Edge Length",
                "Component Color Channel 1 (BGR)",
                "Component Color Channel 2 (BGR)",
                "Channel 1 Path",
                "Channel 2 Path",
                "Original Overlay Path",
                "Output Root Directory",
            ],
            "Value": [
                timestamp_str,
                self.t_min_ch1,
                self.t_mid_ch1,
                self.t_max_ch1,
                self.t_min_ch2,
                self.t_mid_ch2,
                self.t_max_ch2,
                # self.component_thickness,
                self.component_area_min,
                self.component_area_max,
                self.max_edge_length,
                ch1_color_str,
                ch2_color_str,
                self.channel1_path,
                self.channel2_path,
                self.original_overlay_path,
                output_root,
            ]
        }

        df = pd.DataFrame(data)
        df.to_excel(param_xlsx, index=False, sheet_name="Parameter Log")

        from openpyxl import load_workbook

        wb = load_workbook(param_xlsx)
        ws = wb["Parameter Log"]
        _autofit_worksheet_columns(ws, min_width=8, max_width=80, padding=2)
        wb.save(param_xlsx)

    def show_cv2_image(self, img, title="Preview"):
        if img is None:
            return
        self.log_status(f"IMAGE: {title}\n")

        imgrgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        im = Image.fromarray(imgrgb)

        self.update_idletasks()
        mainw = max(self.winfo_width(), 1)
        mainh = max(self.winfo_height(), 1)

        w, h = im.size
        scale = min(mainw / w, mainh / h)
        neww, newh = int(w * scale), int(h * scale)

        im = im.resize((neww, newh), Image.Resampling.LANCZOS)

        photo = ImageTk.PhotoImage(im)
        win = tk.Toplevel(self)
        win.title(title)

        label = ttk.Label(win, image=photo)
        label.image = photo
        label.pack(fill="both", expand=True)

        mainx = self.winfo_x()
        mainy = self.winfo_y()
        x = mainx + max((mainw - neww) // 2, 0)
        y = mainy + max((mainh - newh) // 2, 0)
        win.geometry(f"{neww}x{newh}+{x}+{y}")

        win.transient(self)
        win.grab_set()
        self.wait_window(win)

    def _display_last_parts(self, path: str, n_parts: int = 2) -> str:
        if not path:
            return ""

        norm = os.path.normpath(path)
        parts = [p for p in norm.split(os.sep) if p]

        if not parts:
            return norm

        try:
            n_parts = int(n_parts)
        except (TypeError, ValueError):
            n_parts = 2

        if n_parts <= 0:
            return os.path.basename(norm)

        if len(parts) >= n_parts:
            return os.path.join(*parts[-n_parts:])
        else:
            return os.path.join(*parts)

    def load_channel1(self):
        path = filedialog.askopenfilename(
            title="Select Image Channel 1",
            filetypes=[("Image files", "*.png *.jpg *.tif *.tiff *.bmp")],
            initialdir=self.default_dir
        )
        if path:
            self.channel1_path = path
            ch1_dir = os.path.dirname(path)
            self.default_dir = ch1_dir

            self.lbl_channel1.config(
                text=self._display_last_parts(path),
                foreground="green"
            )

            if not self.channel2_is_explicit:
                self.lbl_channel2.config(
                    text=self._display_last_parts(ch1_dir),
                    foreground="green"
                )

            if not self.overlay_is_explicit:
                self.lbl_overlay.config(
                    text=self._display_last_parts(ch1_dir),
                    foreground="green"
                )

            if not self.save_is_explicit:
                self.save_dir = self._make_timestamped_output_dir(ch1_dir)
                self.lbl_save.config(
                    text=self._display_last_parts(self.save_dir, 3),
                    foreground="green"
                )

            self.log_status(f"Loaded channel 1: {path}")
            self.log_status(f"default_dir set to: {self.default_dir}")


    def load_channel2(self):
        path = filedialog.askopenfilename(
            title="Select Image Channel 2",
            filetypes=[("Image files", "*.png *.jpg *.tif *.tiff *.bmp")],
            initialdir=self.default_dir
        )
        if path:
            self.channel2_path = path
            self.channel2_is_explicit = True
            self.lbl_channel2.config(
                text=self._display_last_parts(path),
                foreground="green"
            )
            self.log_status(f"Loaded channel 2: {path}")


    def load_original_overlay(self):
        path = filedialog.askopenfilename(
            title="Select Original Overlay Image",
            filetypes=[("Image files", "*.png *.jpg *.tif *.tiff *.bmp")],
            initialdir=self.default_dir
        )
        if path:
            self.original_overlay_path = path
            self.overlay_is_explicit = True
            self.lbl_overlay.config(
                text=self._display_last_parts(path),
                foreground="green"
            )
            self.log_status(f"Loaded original overlay: {path}")

            if not self.save_is_explicit:
                overlay_dir = os.path.dirname(path)
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                self.save_dir = os.path.join(overlay_dir, f"output_{timestamp}")
                self.lbl_save.config(
                    text=self._display_last_parts(self.save_dir),
                    foreground="green"
                )
                self.log_status(f"Planned output directory (auto): {self.save_dir}")


    def select_save_directory(self):
        directory = filedialog.askdirectory(
            title="Select Save Location",
            initialdir=self.default_dir
        )
        if directory:
            self.save_is_explicit = True
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            self.save_dir = os.path.join(directory, f"output_{timestamp}")
            self.lbl_save.config(
                text=self._display_last_parts(self.save_dir, 3),
                foreground="green"
            )
            self.log_status(f"Planned output directory (override): {self.save_dir}")
        else:
            self.log_status("Save directory override canceled; using current setting.")

    def _prepare_threshold_params_ch1(self):
        self.t_min_ch1 = int(self.entry_threshold_min_ch1.get())
        self.t_mid_ch1 = int(self.entry_threshold_mid_ch1.get())
        self.t_max_ch1 = int(self.entry_threshold_max_ch1.get())
        # self.component_thickness = int(self.entry_component_thickness.get())
        self.component_area_min = int(self.entry_area_min.get())
        self.component_area_max = int(self.entry_area_max.get())

    def _prepare_threshold_params_ch2(self):
        self.t_min_ch2 = int(self.entry_threshold_min_ch2.get())
        self.t_mid_ch2 = int(self.entry_threshold_mid_ch2.get())
        self.t_max_ch2 = int(self.entry_threshold_max_ch2.get())
        # self.component_thickness = int(self.entry_component_thickness.get())
        self.component_area_min = int(self.entry_area_min.get())
        self.component_area_max = int(self.entry_area_max.get())

    def threshold_one_channel(
        self,
        path: str,
        color: tuple[int, int, int],
        out_subdir: str,
        title: str,
        short_tag: str,
    ) -> tuple[str, str]:
        if not path:
            raise ValueError(f"Missing image path for {short_tag}")
        if not self.save_dir:
            raise ValueError("No save directory set")

        ch_dir = os.path.join(self.save_dir, out_subdir)
        os.makedirs(ch_dir, exist_ok=True)

        if short_tag == "ch1":
            self.t_min_ch1 = int(self.entry_threshold_min_ch1.get())
            self.t_mid_ch1 = int(self.entry_threshold_mid_ch1.get())
            self.t_max_ch1 = int(self.entry_threshold_max_ch1.get())
        else:
            self.t_min_ch2 = int(self.entry_threshold_min_ch2.get())
            self.t_mid_ch2 = int(self.entry_threshold_mid_ch2.get())
            self.t_max_ch2 = int(self.entry_threshold_max_ch2.get())

        if short_tag == "ch1":
            self.threshold_and_components(
                path,
                color,
                ch_dir,
                t_min=self.t_min_ch1,
                t_mid=self.t_mid_ch1,
                t_max=self.t_max_ch1,
                title=title,
                short_tag=short_tag,
            )
        else:
            self.threshold_and_components(
                path,
                color,
                ch_dir,
                t_min=self.t_min_ch2,
                t_mid=self.t_mid_ch2,
                t_max=self.t_max_ch2,
                title=title,
                short_tag=short_tag,
            )

        out_components = os.path.join(ch_dir, f"{short_tag}_components.png")
        out_thresh = os.path.join(ch_dir, f"{short_tag}_thresh.png")
        return out_components, out_thresh

    def on_threshold_ch1_only(self):
        if not self.channel1_path:
            messagebox.showwarning(
                "Missing Channel 1",
                "Please load image channel 1 before processing."
            )
            return
        if not self.save_dir:
            messagebox.showwarning(
                "No Save Directory",
                "Please load an overlay or select a save directory before processing."
            )
            return

        self.log_status("Starting thresholding: channel 1 only...")
        try:
            self._prepare_threshold_params_ch1()
        except ValueError as e:
            messagebox.showerror(
                "Invalid Input",
                f"Please enter valid numeric values:\n{e}"
            )
            self.log_status(f"[ERROR] Invalid input (Ch1): {e}")
            return

        try:
            self.threshold_one_channel(
                self.channel1_path,
                self.component_color_channel_1,
                out_subdir="ch1",
                title="threshold channel 1",
                short_tag="ch1",
            )
            self.log_status("Thresholding complete for channel 1.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_status(f"[ERROR] Thresholding Ch1: {e}")
        finally:
            cv2.destroyAllWindows()

    def on_threshold_ch2_only(self):
        if not self.channel2_path:
            messagebox.showwarning(
                "Missing Channel 2",
                "Please load image channel 2 before processing."
            )
            return
        if not self.save_dir:
            messagebox.showwarning(
                "No Save Directory",
                "Please load an overlay or select a save directory before processing."
            )
            return

        self.log_status("Starting thresholding: channel 2 only...")
        try:
            self._prepare_threshold_params_ch2()
        except ValueError as e:
            messagebox.showerror(
                "Invalid Input",
                f"Please enter valid numeric values:\n{e}"
            )
            self.log_status(f"[ERROR] Invalid input (Ch2): {e}")
            return

        try:
            self.threshold_one_channel(
                self.channel2_path,
                self.component_color_channel_2,
                out_subdir="ch2",
                title="threshold channel 2",
                short_tag="ch2",
            )
            self.log_status("Thresholding complete for channel 2.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_status(f"[ERROR] Thresholding Ch2: {e}")
        finally:
            cv2.destroyAllWindows()

    def on_triangulate(self):
        self.max_edge_length = float(self.entry_max_edge_length.get())

        if not self.channel1_path or not self.channel2_path:
            messagebox.showwarning(
                "Missing Channels",
                "Please load both image channels before triangulating."
            )
            return
        if not self.save_dir:
            messagebox.showwarning(
                "No Save Directory",
                "Please select a save directory before triangulating."
            )
            return
        if not self.original_overlay_path:
            messagebox.showwarning(
                "Missing Original Overlay",
                "Please load the original overlay image before triangulating."
            )
            return

        self.log_status("Starting triangulation (mixed only)...")
        try:
            self.triangulate()
            self.log_status("Triangulation complete.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log_status(f"[ERROR] Triangulation: {e}")

    def threshold_and_components(
        self,
        path,
        component_color,
        output_dir,
        t_min: int,
        t_mid: int,
        t_max: int | None = None,
        title: str = "",
        short_tag: str = "",
    ):
        try:
            del self.csv_paths[short_tag]
        except Exception:
            pass

        self.log_status(f"Thresholding {os.path.basename(path)} as {short_tag}...")
        img_gray = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        img_color = cv2.imread(path)
        if img_gray is None or img_color is None:
            raise ValueError(f"Failed to read image: {path}")

        if t_max is None:
            t_max = 255

        centroids = []
        overlay = img_color.copy()
        component_only = np.zeros_like(img_color)

        centroids_pass1 = 0
        centroids_pass2 = 0

        h, w = img_gray.shape[:2]
        image_area = float(h * w)

        min_area = self.component_area_min
        max_area = self.component_area_max

        self.log_status(f"{short_tag}: min_area={min_area}, max_area={max_area}")

        def add_component_centroid(cx, cy, source_name):
            centroids.append({
                "filename": source_name,
                "x": float(cx),
                "y": float(cy),
            })

        def draw_component_mask(mask_8u):
            component_only[mask_8u > 0] = component_color
            overlay_mask = np.zeros_like(img_color)
            overlay_mask[mask_8u > 0] = component_color
            cv2.addWeighted(overlay_mask, 0.5, overlay, 1.0, 0, dst=overlay)

        _, thresh1 = cv2.threshold(img_gray, t_min, 255, cv2.THRESH_BINARY)

        nlabels1, labels1, stats1, centroids1 = cv2.connectedComponentsWithStats(
            thresh1, connectivity=8
        )

        big_masks = []

        for label in range(1, nlabels1):
            area = stats1[label, cv2.CC_STAT_AREA]

            if area >= max_area:
                mask = np.uint8(labels1 == label) * 255
                big_masks.append(mask)
                continue

            if area < min_area:
                continue

            cx, cy = centroids1[label]
            add_component_centroid(cx, cy, os.path.basename(path))
            centroids_pass1 += 1

            mask = np.uint8(labels1 == label) * 255
            draw_component_mask(mask)

        if t_mid is not None and t_min < t_mid < t_max:
            for big_mask in big_masks:
                roi_gray = cv2.bitwise_and(img_gray, img_gray, mask=big_mask)
                _, thresh2 = cv2.threshold(roi_gray, t_mid, 255, cv2.THRESH_BINARY)

                thresh2 = cv2.bitwise_and(thresh2, big_mask)

                nlabels2, labels2, stats2, centroids2 = cv2.connectedComponentsWithStats(
                    thresh2, connectivity=8
                )

                for label in range(1, nlabels2):
                    area = stats2[label, cv2.CC_STAT_AREA]

                    if area < min_area or area >= max_area:
                        continue

                    cx, cy = centroids2[label]
                    add_component_centroid(cx, cy, os.path.basename(path))
                    centroids_pass2 += 1

                    mask = np.uint8(labels2 == label) * 255
                    draw_component_mask(mask)

        os.makedirs(output_dir, exist_ok=True)
        base = short_tag
        out_overlay = os.path.join(output_dir, f"{base}_overlay.png")
        out_components = os.path.join(output_dir, f"{base}_components.png")
        out_thresh = os.path.join(output_dir, f"{base}_thresh.png")
        csv_path = os.path.join(output_dir, f"{base}_points.csv")

        self.csv_paths[short_tag] = csv_path
        self.component_image_paths[short_tag] = out_components

        thresh_combined = np.zeros_like(img_gray, dtype=np.uint8)
        gray_components = cv2.cvtColor(component_only, cv2.COLOR_BGR2GRAY)
        thresh_combined[gray_components > 0] = 255

        color_mask = np.zeros_like(img_color)
        color_mask[thresh_combined > 0] = component_color

        for c in centroids:
            cx = int(round(c["x"]))
            cy = int(round(c["y"]))
            cv2.circle(overlay, (cx, cy), 1, (255, 255, 255), -1)

        cv2.imwrite(out_overlay, overlay)
        cv2.imwrite(out_components, component_only)
        cv2.imwrite(out_thresh, color_mask)

        with open(csv_path, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["filename", "x", "y"])
            for c in centroids:
                writer.writerow([c["filename"], c["x"], c["y"]])

        total = len(centroids)
        self.log_status(
            f"{short_tag}: pass1 centroids={centroids_pass1}, "
            f"pass2 centroids={centroids_pass2}, total={total}, "
            f"min_area={min_area}, max_area={max_area}"
        )

        assert total == centroids_pass1 + centroids_pass2, \
            f"Centroid count mismatch for {short_tag}"

        self.show_cv2_image(overlay, title=title)

    def plot_histogram(
        self,
        values,
        bins: int = 10,
        png_filename: str = "hist.png",
        excel_filename: str = "hist.xlsx",
        output_path: str = ".",
        xlabel="Value"
    ):
        os.makedirs(output_path, exist_ok=True)
        counts, bin_edges = np.histogram(values, bins=bins)
        plt.figure()
        plt.hist(values, bins=bins)
        plt.title("Histogram")
        plt.xlabel(xlabel)
        plt.ylabel("Frequency")
        png_full = os.path.join(output_path, png_filename)
        plt.savefig(png_full, dpi=300)
        plt.close()
        df = pd.DataFrame({
            "bin_left": bin_edges[:-1],
            "bin_right": bin_edges[1:],
            "count": counts
        })
        xlsx_full = os.path.join(output_path, excel_filename)
        df.to_excel(xlsx_full, index=False)
        self.log_status(f"Histogram saved: {png_full, xlsx_full}")

    def autocontrast(self, img: np.ndarray) -> np.ndarray:
        chans = cv2.split(img)
        out_chans = []
        for ch in chans:
            lo, hi = int(ch.min()), int(ch.max())
            if hi > lo:
                stretched = ((ch - lo) * 255.0 / (hi - lo)).clip(0, 255).astype(np.uint8)
            else:
                stretched = ch
            out_chans.append(stretched)
        return cv2.merge(out_chans)

    def _ensure_bgr_image(self, img_or_path):
        if isinstance(img_or_path, np.ndarray):
            return img_or_path
        img = cv2.imread(img_or_path)
        if img is None:
            raise FileNotFoundError(f"Could not load image: {img_or_path}")
        return img

    def overlay_evenly(
        self,
        img1,
        img2,
        output_path: str = None,
        alpha: float = 0.5,
        title=""
    ) -> any:
        self.log_status("Creating combined component overlay...")
        img1 = self._ensure_bgr_image(img1)
        img2 = self._ensure_bgr_image(img2)

        if img1.shape[:2] != img2.shape[:2]:
            img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]), interpolation=cv2.INTER_AREA)
        blended = cv2.addWeighted(img1, alpha, img2, 1 - alpha, 0)
        auto = self.autocontrast(blended)
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            cv2.imwrite(output_path, auto)
            self.component_image_overlay_path = output_path
            self.log_status(f"Combined overlay saved: {output_path}")
        if title:
            self.show_cv2_image(auto, title=title)
        return auto

    def overlay_edges_with_vertices(
        self,
        image,
        edges: list[tuple[tuple[float, float], tuple[float, float]]],
        output_path: str = None,
        vertex_color: tuple[int, int, int] = (0, 0, 0),
        edge_color: tuple[int, int, int] = (0, 0, 255),
        vertex_radius: int = 4,
        edge_thickness: int = 1,
        title: str = ""
    ) -> np.ndarray:
        name = image if isinstance(image, str) else ""
        self.log_status(f"Drawing edges on {os.path.basename(name)} ({title})...")
        img = self._ensure_bgr_image(image)

        for (x1, y1), (x2, y2) in edges:
            cv2.line(
                img,
                (int(x1), int(y1)),
                (int(x2), int(y2)),
                edge_color,
                thickness=edge_thickness,
                lineType=cv2.LINE_AA,
            )

        unique_pts = {(p[0], p[1]) for edge in edges for p in edge}
        for x, y in unique_pts:
            cv2.circle(
                img,
                (int(x), int(y)),
                radius=vertex_radius,
                color=vertex_color,
                thickness=-1,
                lineType=cv2.LINE_AA,
            )

        if output_path is not None:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            cv2.imwrite(output_path, img)
            self.log_status(f"Overlay saved: {output_path}")
        else:
            self.log_status("Overlay generated (not saved to file).")

        if title:
            self.show_cv2_image(img, title=title)
        return img

    def overlay_edges_with_vertices_outline(
        self,
        image,
        edges,
        output_path: str = None,
        vertex_color=(0, 0, 0),
        edge_color=(0, 0, 255),
        vertex_radius: int = 4,
        edge_thickness: int = 1,
        title: str = "",
        edge_outline_color=None,
        edge_outline_thickness: int | None = None,
        edge_alpha: float = 1.0,
    ):
        name = image if isinstance(image, str) else ""
        self.log_status(f"Drawing edges on {os.path.basename(name)} ({title})...")
        base = self._ensure_bgr_image(image)

        if edge_outline_color is not None and edge_outline_thickness is not None:
            for (x1, y1), (x2, y2) in edges:
                cv2.line(
                    base,
                    (int(x1), int(y1)),
                    (int(x2), int(y2)),
                    edge_outline_color,
                    thickness=edge_outline_thickness,
                    lineType=cv2.LINE_AA,
                )

        overlay = base.copy()
        for (x1, y1), (x2, y2) in edges:
            cv2.line(
                overlay,
                (int(x1), int(y1)),
                (int(x2), int(y2)),
                edge_color,
                thickness=edge_thickness,
                lineType=cv2.LINE_AA,
            )

        alpha = max(0.0, min(1.0, float(edge_alpha)))
        img = cv2.addWeighted(overlay, alpha, base, 1 - alpha, 0)

        unique_pts = {(p[0], p[1]) for edge in edges for p in edge}
        for x, y in unique_pts:
            cv2.circle(
                img,
                (int(x), int(y)),
                radius=vertex_radius,
                color=vertex_color,
                thickness=-1,
                lineType=cv2.LINE_AA,
            )

        if output_path is not None:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            cv2.imwrite(output_path, img)
            self.log_status(f"Overlay saved: {output_path}")
        else:
            self.log_status("Overlay generated (not saved to file).")

        if title:
            self.show_cv2_image(img, title=title)
        return img

    def get_edges_tuples(self, edge_dict):
        edge_tuples = []
        for e in edge_dict:
            e = edge_dict[e]
            o = e.origin_vertex
            d = e.destination_vertex
            edge_tuple = ((o.x, o.y), (d.x, d.y))
            edge_tuples.append(edge_tuple)
        return edge_tuples

    def triangulate(self):
        mixed_dir = os.path.join(self.save_dir, "mixed")
        os.makedirs(mixed_dir, exist_ok=True)

        def _build_triangulation(vertices):
            h = convexHull3D(vertices, for_2d_triangulation=True)
            t = h.get_triangulation()
            return h, t

        def _analyze_triangulation(t, labels, label_tag):
            self.log_status("Triangulation in progress: computing edge lengths and histogram...")
            t.get_all_edges(d_cutoff=0)
            os.makedirs(mixed_dir, exist_ok=True)

            self.log_status("Triangulation in progress: drawing full edge network...")
            base_overlay_path = self.original_overlay_path
            edges = self.get_edges_tuples(t.all_edges)
            self.overlay_edges_with_vertices(
                base_overlay_path,
                edges,
                vertex_radius=1,
                edge_thickness=2,
                vertex_color=(255, 255, 255),
                edge_color=(120, 120, 120),
                output_path=os.path.join(mixed_dir, "tri_full.png"),
                title="triangulation full"
            )

            self.log_status("Triangulation in progress: drawing trimmed edges...")
            all_edges_with_cutoff = t.get_all_edges(d_cutoff=self.max_edge_length)
            edges = self.get_edges_tuples(all_edges_with_cutoff)
            self.overlay_edges_with_vertices(
                base_overlay_path,
                edges,
                vertex_radius=1,
                edge_thickness=2,
                vertex_color=(255, 255, 255),
                edge_color=(120, 120, 120),
                output_path=os.path.join(mixed_dir, "tri_trimmed.png"),
                title="triangulation trimmed"
            )

            self.log_status("Triangulation in progress: drawing mixed vs same edges...")
            all_edges_with_cutoff_clustered = t.get_all_edges_clustered(
                d_cutoff=self.max_edge_length,
                min_cluster_size=2
            )
            t.get_edges_by_labels(
                labels,
                d_cutoff=self.max_edge_length,
                get_edge_clusters=True,
                min_cluster_size=2
            )

            same_e, mixed_e = {}, {}
            same_v, mixed_v = {}, {}

            for key in all_edges_with_cutoff_clustered:
                e = all_edges_with_cutoff_clustered[key]

                lo = getattr(e.origin_vertex.data, "label", None)
                ld = getattr(e.destination_vertex.data, "label", None)

                is_mixed_edge = (lo is not None and ld is not None and lo != ld)

                if is_mixed_edge:
                    mixed_e[e.id] = e
                    mixed_v[e.origin_vertex] = None
                    mixed_v[e.destination_vertex] = None
                else:
                    same_e[e.id] = e
                    same_v[e.origin_vertex] = None
                    same_v[e.destination_vertex] = None

            same_edges_tuple = self.get_edges_tuples(same_e)
            mixed_edges_tuple = self.get_edges_tuples(mixed_e)

            mixed_and_same_edges_overlay_image = self.overlay_edges_with_vertices_outline(
                base_overlay_path,
                mixed_edges_tuple,
                vertex_color=(255, 255, 255),
                edge_color=(0, 255, 255),
                vertex_radius=1,
                edge_thickness=1,
                edge_outline_color=(0, 0, 255),
                edge_outline_thickness=3,
                edge_alpha=0.9,
            )

            self.overlay_edges_with_vertices(
                mixed_and_same_edges_overlay_image,
                same_edges_tuple,
                vertex_color=(255, 255, 255),
                vertex_radius=1,
                edge_color=(120, 120, 120),
                edge_thickness=2,
                output_path=os.path.join(mixed_dir, "tri_trimmed_mixed_vs_same_edges.png"),
                title="triangulation mixed vs same edges",
            )

            mixed_edges_overlay_image = self.overlay_edges_with_vertices_outline(
                base_overlay_path,
                mixed_edges_tuple,
                vertex_color=(255, 255, 255),
                edge_color=(0, 255, 255),
                vertex_radius=1,
                edge_thickness=1,
                edge_outline_color=(0, 0, 255),
                edge_outline_thickness=3,
                edge_alpha=0.9,
                output_path=os.path.join(mixed_dir, "tri_trimmed_mixed_edges.png"),
                title="triangulation mixed edges",
            )

            # mixed_edges_overlay_image = self.overlay_edges_with_vertices_outline(
            #     base_overlay_path,
            #     mixed_edges_tuple,
            #     vertex_color=(255, 255, 255),
            #     edge_color=(0, 255, 255),
            #     vertex_radius=1,
            #     edge_thickness=1,
            #     edge_outline_color=(0, 0, 255),
            #     edge_outline_thickness=3,
            #     edge_alpha=0.9,
            # )

            # self.overlay_edges_with_vertices(
            #     mixed_edges_overlay_image,
            #     same_edges_tuple,
            #     vertex_color=(255, 255, 255),
            #     vertex_radius=1,
            #     edge_color=(120, 120, 120),
            #     edge_thickness=2,
            #     output_path=os.path.join(mixed_dir, "tri_trimmed_mixed.png"),
            #     title="triangulation mixed vs same",
            # )

            vertex_stats = {}
            for e in t.all_edges.values():
                for v in (e.origin_vertex, e.destination_vertex):
                    if v not in vertex_stats:
                        vertex_stats[v] = {
                            "id": v.id,
                            "x": v.x,
                            "y": v.y,
                            "label": getattr(v.data, "label", ""),
                            "is_mixed_vertex": bool(getattr(v.data, "mixed", False)),
                            "n_mixed_edges": 0,
                            "n_same_edges": 0,
                        }

                    lbl_o = getattr(e.origin_vertex.data, "label", None)
                    lbl_d = getattr(e.destination_vertex.data, "label", None)
                    is_mixed_edge = (
                        lbl_o is not None and
                        lbl_d is not None and
                        lbl_o != lbl_d
                    )

                    if is_mixed_edge:
                        vertex_stats[v]["n_mixed_edges"] += 1
                    else:
                        vertex_stats[v]["n_same_edges"] += 1

            for v_stat in vertex_stats.values():
                total_edges = v_stat["n_mixed_edges"] + v_stat["n_same_edges"]
                if total_edges > 0:
                    v_stat["fraction_mixed_edges"] = v_stat["n_mixed_edges"] / total_edges
                else:
                    v_stat["fraction_mixed_edges"] = 0.0

            n_vertices_with_mixed = sum(
                1 for vs in vertex_stats.values() if vs["n_mixed_edges"] > 0
            )

            n_vertices_total = len(vertex_stats)
            fraction_vertices_with_mixed_edges = (
                n_vertices_with_mixed / n_vertices_total if n_vertices_total > 0 else 0.0
            )

            self.log_status(
                f"{label_tag}: vertices_with_mixed={n_vertices_with_mixed}, "
                f"total_vertices={n_vertices_total}, "
                f"fraction_vertices_with_mixed_edges={fraction_vertices_with_mixed_edges:.3f}"
            )

            n_vertices_without_mixed = n_vertices_total - n_vertices_with_mixed
            fraction_vertices_without_mixed_edges = (
                n_vertices_without_mixed / n_vertices_total if n_vertices_total > 0 else 0.0
            )

            summary_xlsx = os.path.join(mixed_dir, f"{label_tag}_vertex_mixing_summary.xlsx")

            data = {
                "Metric": [
                    "Total Vertices",
                    "Vertices with Mixed Edges",
                    "Fraction with Mixed Edges",
                    "Vertices without Mixed Edges",
                    "Fraction without Mixed Edges",
                ],
                "Value": [
                    n_vertices_total,
                    n_vertices_with_mixed,
                    fraction_vertices_with_mixed_edges,
                    n_vertices_without_mixed,
                    fraction_vertices_without_mixed_edges,
                ]
            }

            df = pd.DataFrame(data)
            df.to_excel(summary_xlsx, index=False, sheet_name="Vertex Mixing Summary")

            from openpyxl import load_workbook

            wb = load_workbook(summary_xlsx)
            ws = wb["Vertex Mixing Summary"]

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value is None:
                        continue
                    max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(summary_xlsx)

            self.log_status(f"{label_tag}: vertex mixing summary saved: {summary_xlsx}")

            mixed_edges_per_vertex = [vs["n_mixed_edges"] for vs in vertex_stats.values()]
            same_edges_per_vertex = [vs["n_same_edges"] for vs in vertex_stats.values()]

            self.plot_histogram(
                mixed_edges_per_vertex,
                bins=20,
                png_filename=f"{label_tag}_hist_mixed_edges_per_vertex.png",
                excel_filename=f"{label_tag}_hist_mixed_edges_per_vertex.xlsx",
                output_path=mixed_dir,
                xlabel="number of mixed edges"
            )

            self.plot_histogram(
                same_edges_per_vertex,
                bins=20,
                png_filename=f"{label_tag}_hist_same_edges_per_vertex.png",
                excel_filename=f"{label_tag}_hist_same_edges_per_vertex.xlsx",
                output_path=mixed_dir,
                xlabel="number of same edges"
            )

            n_same_v, n_mixed_v = len(same_v), len(mixed_v)
            fraction_mixed_vertices = (
                n_mixed_v / (n_mixed_v + n_same_v) if (n_same_v + n_mixed_v) > 0 else 0.0
            )

            self.log_status(
                f"{label_tag}: same_v={n_same_v}, mixed_v={n_mixed_v}, "
                f"fraction_mixed_vertices={fraction_mixed_vertices:.3f}"
            )

            return {
                "same_v": n_same_v,
                "mixed_v": n_mixed_v,
                "same_e": len(same_e),
                "mixed_e": len(mixed_e),
                "fraction_mixed_vertices": fraction_mixed_vertices,
            }

        # self.log_status("Triangulation in progress: building vertex set and hull...")
        # all_vertices = []
        # vid_global = 1
        # all_labels = []

        # for csv_path_key in self.csv_paths:
        #     csv_path = self.csv_paths[csv_path_key]
        #     with open(csv_path, newline='') as csvfile:
        #         label = os.path.basename(csvfile.name).split("_points")[0]
        #         all_labels.append(label)
        #         reader = csv.reader(csvfile)
        #         next(reader, None)
        #         for fname, x_str, y_str in reader:
        #             x, y = float(x_str), float(y_str)
        #             z = x**2 + y**2
        #             data = Data()
        #             data.label = label
        #             data.x, data.y, data.z = x, y, z
        #             v = Vertex((x, y, z), data=data, unique_id=vid_global)
        #             all_vertices.append(v)
        #             vid_global += 1

        self.log_status("Triangulation in progress: building vertex set and hull...")
        all_vertices = []
        vid_global = 1
        all_labels = []

        if not self.csv_paths:
            msg = "No threshold point files found. Run thresholding for Channel 1 and Channel 2 before triangulating."
            self.log_status(f"[ERROR] {msg}")
            messagebox.showwarning("No points to triangulate", msg)
            return

        required_tags = {"ch1", "ch2"}
        missing_tags = required_tags - set(self.csv_paths.keys())
        if missing_tags:
            msg = f"Missing threshold results for: {', '.join(sorted(missing_tags))}. Run both threshold operations before triangulating."
            self.log_status(f"[ERROR] {msg}")
            messagebox.showwarning("Missing threshold results", msg)
            return

        for csv_path_key in self.csv_paths:
            csv_path = self.csv_paths[csv_path_key]

            if not os.path.exists(csv_path):
                self.log_status(f"[WARNING] Missing CSV, skipping: {csv_path}")
                continue

            with open(csv_path, newline='') as csvfile:
                label = os.path.basename(csvfile.name).split("_points")[0]
                all_labels.append(label)
                reader = csv.reader(csvfile)
                next(reader, None)

                for fname, x_str, y_str in reader:
                    x, y = float(x_str), float(y_str)
                    z = x**2 + y**2
                    data = Data()
                    data.label = label
                    data.x, data.y, data.z = x, y, z
                    v = Vertex((x, y, z), data=data, unique_id=vid_global)
                    all_vertices.append(v)
                    vid_global += 1

        if len(all_vertices) == 0:
            msg = "No points were found in the threshold CSV files. Run thresholding for both channels before triangulating."
            self.log_status(f"[ERROR] {msg}")
            messagebox.showwarning("No points to triangulate", msg)
            return

        if len(all_vertices) < 3:
            msg = f"Not enough points to triangulate. Need at least 3 points, found {len(all_vertices)}."
            self.log_status(f"[ERROR] {msg}")
            messagebox.showwarning("Not enough points", msg)
            return

        h_all, t_all = _build_triangulation(all_vertices)
        _ = _analyze_triangulation(t_all, all_labels, label_tag="mixed")

        self.write_parameter_log(self.save_dir)


if __name__ == "__main__":
    app = ImageTriangulatorApp()
    app.mainloop()